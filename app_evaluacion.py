import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import streamlit as st
import pandas as pd
import numpy as np
import math
from typing import Dict, List, Optional, Tuple
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import copy
# ============================================================
# IMPORTAR CONFIGURACIÓN CENTRALIZADA (Fase 2)
# ============================================================
from core.config import ConfiguracionProyecto
from core.data_manager import DataManager
from core.excel_engine import ExcelReportEngine
from core.project_manager import ProjectManager

# ============================================================
# FUNCIONES MATEMÁTICAS — CON FALLBACK DE BISECCIÓN
# ============================================================

def calcular_van(tasa: float, flujos: List[float]) -> float:
    if not flujos:
        return 0.0
    return sum(flujo / ((1 + tasa) ** i) for i, flujo in enumerate(flujos))


def calcular_tir(flujos: List[float], estimado: float = 0.1) -> Optional[float]:
    """
    Calcula TIR con Newton-Raphson. Si falla, usa bisección.
    Si no hay cambio de signo en los flujos, retorna None.
    """
    if len(flujos) < 2:
        return None
    
    # Validar cambio de signo (condición necesaria para TIR real)
    signos = [f >= 0 for f in flujos if f != 0]
    if len(set(signos)) < 2:
        # No hay cambio de signo: no existe TIR real (proyecto siempre positivo o negativo)
        return 0.0
    
    def npv(rate):
        return sum(f / (1 + rate) ** i for i, f in enumerate(flujos))
    
    # --- Intento 1: Newton-Raphson ---
    def derivative(rate):
        return sum(-i * f / (1 + rate) ** (i + 1) for i, f in enumerate(flujos) if i > 0)
    
    rate = estimado
    for _ in range(100):
        try:
            npv_val = npv(rate)
            deriv_val = derivative(rate)
            if abs(deriv_val) < 1e-10:
                break
            new_rate = rate - npv_val / deriv_val
            if abs(new_rate - rate) < 1e-6:
                if -0.99 <= new_rate <= 3.0:  # TIR razonable
                    return new_rate
                break
            rate = new_rate
        except:
            break
    
    # --- Intento 2: Bisección (fallback) ---
    try:
        # Buscar intervalo [a, b] donde NPV cambia de signo
        a, b = -0.99, 3.0
        npv_a, npv_b = npv(a), npv(b)
        
        # Expandir hacia abajo si es necesario
        while npv_a * npv_b > 0 and a > -0.99:
            a -= 0.1
            npv_a = npv(a)
        
        if npv_a * npv_b > 0:
            return None  # No se encontró intervalo con cambio de signo
        
        # Asegurar que a tenga NPV positivo y b negativo (o viceversa)
        if npv_a > 0:
            a, b = b, a
        
        for _ in range(100):
            c = (a + b) / 2
            npv_c = npv(c)
            if abs(npv_c) < 1e-6:
                return c
            if npv_c < 0:
                a = c
            else:
                b = c
        
        tir_biseccion = (a + b) / 2
        if -0.5 <= tir_biseccion <= 2.0:
            return tir_biseccion
    except:
        pass
    
    return 0.0
# ============================================================
# PROYECCIÓN POBLACIONAL
# ============================================================
class ProyeccionPoblacional:
    def __init__(self, config: ConfiguracionProyecto):
        self.config = config
        self.anios = config.generar_anios()
    
    def calcular_lineal(self) -> List[float]:
        return [self.config.poblacion_base * (1 + self.config.tasa_crecimiento * (anio - self.config.anio_inicio)) 
                for anio in self.anios]
    
    def calcular_geometrica(self) -> List[float]:
        return [self.config.poblacion_base * (1 + self.config.tasa_crecimiento) ** (anio - self.config.anio_inicio) 
                for anio in self.anios]
    
    def calcular_exponencial(self) -> List[float]:
        return [self.config.poblacion_base * math.exp(self.config.tasa_crecimiento * (anio - self.config.anio_inicio)) 
                for anio in self.anios]
    
    def generar_dataframe(self) -> pd.DataFrame:
        df = pd.DataFrame({
            'Año': self.anios,
            'Lineal': self.calcular_lineal(),
            'Geométrico': self.calcular_geometrica(),
            'Exponencial': self.calcular_exponencial()
        })
        df['Promedio'] = df[['Lineal', 'Geométrico', 'Exponencial']].mean(axis=1)
        return df
# ============================================================
# GENERADOR DE REPORTE PRODUCCIÓN AGRÍCOLA  —  MAPEO CORREGIDO
# ============================================================
def generar_df_valor_neto_produccion(datos_cult: Dict) -> pd.DataFrame:
    """
    Genera tabla consolidada de Valor Neto de Producción (SP vs CP vs Incremental).
    Todos los valores son TOTALES escalados por superficie cultivada (Ha).
    """
    df = datos_cult['df'].copy()

    # Asegurar columnas numéricas
    numeric_cols = ['sup_sp_ha', 'sup_cp_ha', 'ingreso_sp', 'ingreso_cp',
                    'costototal_sp', 'costototal_cp',
                    'bt_sp', 'bt_cp', 'bnt_sp', 'bnt_cp',
                    'moc_sp', 'moc_cp', 'mos_sp', 'mos_cp',
                    'monu_sp', 'monu_cp', 'monr_sp', 'monr_cp']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Escalar por superficie (valores TOTALES)
    df['ingreso_sp_total'] = df['ingreso_sp'] * df['sup_sp_ha']
    df['ingreso_cp_total'] = df['ingreso_cp'] * df['sup_cp_ha']
    df['costo_sp_total'] = df['costototal_sp'] * df['sup_sp_ha']
    df['costo_cp_total'] = df['costototal_cp'] * df['sup_cp_ha']

    # Valor Neto = Ingreso − Costo
    df['vneto_sp'] = df['ingreso_sp_total'] - df['costo_sp_total']
    df['vneto_cp'] = df['ingreso_cp_total'] - df['costo_cp_total']

    # Incrementales (CP − SP)
    df['delta_ingreso'] = df['ingreso_cp_total'] - df['ingreso_sp_total']
    df['delta_costo'] = df['costo_cp_total'] - df['costo_sp_total']
    df['delta_vneto'] = df['vneto_cp'] - df['vneto_sp']

    # Construir DataFrame de salida
    data = {
        'Cultivo': df['nombre'].values,
        'Sup. SP (Ha)': df['sup_sp_ha'].values,
        'Sup. CP (Ha)': df['sup_cp_ha'].values,
        'Ingreso SP (Bs)': df['ingreso_sp_total'].values,
        'Ingreso CP (Bs)': df['ingreso_cp_total'].values,
        'Δ Ingreso (Bs)': df['delta_ingreso'].values,
        'Costo SP (Bs)': df['costo_sp_total'].values,
        'Costo CP (Bs)': df['costo_cp_total'].values,
        'Δ Costo (Bs)': df['delta_costo'].values,
        'V.Neto SP (Bs)': df['vneto_sp'].values,
        'V.Neto CP (Bs)': df['vneto_cp'].values,
        'Δ V.Neto (Bs)': df['delta_vneto'].values,
    }

    df_out = pd.DataFrame(data)

    # Fila TOTAL
    totales = {'Cultivo': 'TOTAL'}
    for col in df_out.columns[1:]:
        totales[col] = df_out[col].sum()
    df_out = pd.concat([df_out, pd.DataFrame([totales])], ignore_index=True)

    return df_out

def generar_df_costos_rpc_incremental(datos_cult: Dict) -> pd.DataFrame:
    """
    Desglose de costos incrementales por clasificación RPC (RM 115/2015).
    Valores TOTALES escalados por la diferencia de superficie.
    """
    df = datos_cult['df'].copy()

    numeric_cols = ['sup_sp_ha', 'sup_cp_ha',
                    'bt_sp', 'bt_cp', 'bnt_sp', 'bnt_cp',
                    'moc_sp', 'moc_cp', 'mos_sp', 'mos_cp',
                    'monu_sp', 'monu_cp', 'monr_sp', 'monr_cp']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Costos totales SP y CP
    df['bt_sp_total'] = df['bt_sp'] * df['sup_sp_ha']
    df['bt_cp_total'] = df['bt_cp'] * df['sup_cp_ha']
    df['bnt_sp_total'] = df['bnt_sp'] * df['sup_sp_ha']
    df['bnt_cp_total'] = df['bnt_cp'] * df['sup_cp_ha']
    df['moc_sp_total'] = df['moc_sp'] * df['sup_sp_ha']
    df['moc_cp_total'] = df['moc_cp'] * df['sup_cp_ha']
    df['mos_sp_total'] = df['mos_sp'] * df['sup_sp_ha']
    df['mos_cp_total'] = df['mos_cp'] * df['sup_cp_ha']
    df['monu_sp_total'] = df['monu_sp'] * df['sup_sp_ha']
    df['monu_cp_total'] = df['monu_cp'] * df['sup_cp_ha']
    df['monr_sp_total'] = df['monr_sp'] * df['sup_sp_ha']
    df['monr_cp_total'] = df['monr_cp'] * df['sup_cp_ha']

    # Incrementales
    data = {
        'Cultivo': df['nombre'].values,
        'Δ BT (Bs)': (df['bt_cp_total'] - df['bt_sp_total']).values,
        'Δ BNT (Bs)': (df['bnt_cp_total'] - df['bnt_sp_total']).values,
        'Δ MOC (Bs)': (df['moc_cp_total'] - df['moc_sp_total']).values,
        'Δ MOS (Bs)': (df['mos_cp_total'] - df['mos_sp_total']).values,
        'Δ MONU (Bs)': (df['monu_cp_total'] - df['monu_sp_total']).values,
        'Δ MONR (Bs)': (df['monr_cp_total'] - df['monr_sp_total']).values,
        'Δ Total (Bs)': (
            (df['bt_cp_total'] - df['bt_sp_total']) +
            (df['bnt_cp_total'] - df['bnt_sp_total']) +
            (df['moc_cp_total'] - df['moc_sp_total']) +
            (df['mos_cp_total'] - df['mos_sp_total']) +
            (df['monu_cp_total'] - df['monu_sp_total']) +
            (df['monr_cp_total'] - df['monr_sp_total'])
        ).values,
    }

    df_out = pd.DataFrame(data)

    # Fila TOTAL
    totales = {'Cultivo': 'TOTAL'}
    for col in df_out.columns[1:]:
        totales[col] = df_out[col].sum()
    df_out = pd.concat([df_out, pd.DataFrame([totales])], ignore_index=True)

    return df_out
# ============================================================
# ESTILOS OPENPYXL COMUNES
# ============================================================
def _estilos_base():
    return {
        'header_fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        'bold_font': Font(bold=True, size=11),
        'subheader_font': Font(bold=True, size=11, color="1F4E78"),
        'title_font': Font(bold=True, size=14, color="1F4E78"),
        'thin_border': Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin')),
        'center_align': Alignment(horizontal='center', vertical='center', wrap_text=True),
    }

def _write_row(ws, row_idx, values, bold=False, fill=None, border=True, align=None, number_format='#,##0.00'):
    styles = _estilos_base()
    font = styles['bold_font'] if bold else Font()
    for c_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=c_idx, value=val)
        cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = styles['thin_border']
        if align:
            cell.alignment = align
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            cell.number_format = number_format
    return row_idx + 1

def _ajustar_anchos(ws):
    for idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(idx)
        max_len = 0
        for row in ws.iter_rows(min_col=idx, max_col=idx):
            cell = row[0]
            try:
                if cell.value is not None:
                    val_len = len(str(cell.value))
                    if val_len > max_len:
                        max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

def _escribir_tabla_prep(ws, df, header_fill, bold_font, thin_border, center_align):
    nivel1 = [c[0] for c in df.columns]
    nivel2 = [c[1] for c in df.columns]
    
    ws.append(nivel1)
    row_hdr1 = ws.max_row
    for c_idx, val in enumerate(nivel1, 1):
        cell = ws.cell(row=row_hdr1, column=c_idx, value=val if val else None)
        cell.fill = header_fill
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align
    
    current_group = None
    start_col = 1
    for c_idx, val in enumerate(nivel1, 1):
        if val != current_group and val != '':
            if current_group is not None:
                ws.merge_cells(start_row=row_hdr1, start_column=start_col, 
                               end_row=row_hdr1, end_column=c_idx-1)
            current_group = val
            start_col = c_idx
    if current_group:
        ws.merge_cells(start_row=row_hdr1, start_column=start_col,
                       end_row=row_hdr1, end_column=len(nivel1))
    
    ws.append(nivel2)
    row_hdr2 = ws.max_row
    for c_idx, val in enumerate(nivel2, 1):
        cell = ws.cell(row=row_hdr2, column=c_idx, value=val)
        cell.fill = header_fill
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align
    
    for r_idx, row in enumerate(df.values, row_hdr2 + 1):
        ws.append(list(row))
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
            if val == 'TOTAL':
                cell.font = bold_font

# ============================================================
# GENERADOR EXCEL HOJA PREP
# ============================================================
def _generar_hoja_prep_engine(engine: ExcelReportEngine, config, datos):
    """Genera hoja PREP con el motor genérico."""
    # ProyeccionPoblacional está definida en este mismo archivo (líneas ~100-130)
    
    # 1. Datos Generales
    fila = engine.escribir_datos_generales(1, config)
    
    # 2. Población
    fila = engine.escribir_subtitulo(fila, "2. ESTIMACION DE POBLACION OBJETIVO PARA EL PERIODO DE DISEÑO", 6)
    
    proj = ProyeccionPoblacional(config)  # ← Usa la clase local, sin import
    df_pob = proj.generar_dataframe()
    
    # Headers de población
    hdr = ["METODOS DE ESTIMACION"] + [str(a) for a in df_pob['Año'].tolist()]
    for col_idx, h in enumerate(hdr, 1):
        engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
    fila += 1
    
    metodos = ['Lineal', 'Geométrico', 'Exponencial', 'Promedio']
    for met in metodos:
        if met == 'Promedio':
            fila_datos = [met] + [round(v, 2) for v in df_pob['Promedio'].tolist()]
        else:
            fila_datos = [met] + [round(v, 2) for v in df_pob[met].tolist()]
        engine.escribir_fila_datos(fila, fila_datos, bold=(met=='Promedio'))
        fila += 1
    
    fila += 1

    # 3. Valor Neto de Producción (consolidado)
    fila = engine.escribir_subtitulo(fila, "3. VALOR NETO DE PRODUCCIÓN (Bs/año)", 13)
    df_vn = generar_df_valor_neto_produccion(datos['cultivos'])
    fila = engine.escribir_tabla_simple(fila, df_vn)
    
    fila += 1
    
    # 4. Desglose RPC de costos incrementales
    fila = engine.escribir_subtitulo(fila, "4. COMPOSICIÓN DE COSTOS INCREMENTALES POR RPC (Bs/año)", 8)
    df_rpc = generar_df_costos_rpc_incremental(datos['cultivos'])
    fila = engine.escribir_tabla_simple(fila, df_rpc)

    fila += 1

    # 5. COMPOSICIÓN DE LA INVERSIÓN
    fila = engine.escribir_subtitulo(fila, "5. COMPOSICIÓN DE LA INVERSIÓN (Bs)", 8)
    
    if 'inversion' in datos and datos['inversion'].get('ok') and not datos['inversion']['df'].empty:
        df_inv = datos['inversion']['df'].copy()
        
        # Asegurar nombres de columnas limpios para el motor de reportes
        df_inv.columns = [str(c).strip() for c in df_inv.columns]
        
        # Renombrar columnas si vienen con variaciones para mantener consistencia visual
        col_map = {}
        for c in df_inv.columns:
            c_upper = c.upper()
            if 'CATEGOR' in c_upper:
                col_map[c] = 'Categoría'
            elif c == 'TOTAL':
                col_map[c] = 'TOTAL'
            else:
                col_map[c] = c  # BT, BNT, MOC, MOS, MONU, MONR se mantienen
        if col_map:
            df_inv.rename(columns=col_map, inplace=True)
        
        # Formatear valores numéricos
        numeric_cols = [c for c in df_inv.columns if c != 'Categoría']
        for col in numeric_cols:
            df_inv[col] = pd.to_numeric(df_inv[col], errors='coerce').fillna(0)
        
        fila = engine.escribir_tabla_simple(fila, df_inv)
    else:
        engine.escribir_celda(fila, 1, "No hay datos de inversión disponibles en el puente.")
        fila += 1
# ============================================================
# GENERADOR EXCEL HOJA E_FIN — VERSIÓN PROFESIONAL
# ============================================================
def _generar_hoja_efin_engine(engine: ExcelReportEngine, config, df_fin, ind_fin):
    """Genera hoja E_FIN con estructura profesional para EDTP Riego."""
    anios = config.generar_anios()
    n = len(anios)
    ws = engine.ws_active

    # --- TÍTULO PRINCIPAL ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+3)
    engine.escribir_celda(1, 1, "EVALUACIÓN FINANCIERA (Precios de Mercado)",
                         bold=True, size=14, color=engine.COLOR_TITULO, align_h='center')
    engine.escribir_celda(2, 1, f"Proyecto: {config.nombre}  |  Horizonte: {anios[0]} - {anios[-1]}  |  TPD: {config.tasa_privada_descuento*100:.2f}%",
                         size=10, align_h='center')
    fila = 4

    # --- 1. PARÁMETROS ---
    fila = engine.escribir_titulo_seccion(fila, "1. PARÁMETROS DE LA EVALUACIÓN", n+3)
    params = [
        ("Tasa Privada de Descuento (TPD):", f"{config.tasa_privada_descuento*100:.4f} %"),
        ("Duración Inversión:", f"{config.duracion_inversion} años"),
        ("Superficie Incremental:", f"{config.area_incremental:,.2f} Ha"),
        ("Total Familias Beneficiadas:", f"{config.total_familias:,}"),
    ]
    for label, val in params:
        engine.escribir_celda(fila, 1, label, bold=True, align_h='left')
        engine.escribir_celda(fila, 3, val, align_h='left')
        fila += 1
    fila += 1

    # --- 2. INGRESOS INCREMENTALES ---
    fila = engine.escribir_titulo_seccion(fila, "2. INGRESOS INCREMENTALES DEL PROYECTO (VNPI)", n+3)
    headers = ["Concepto", "Unidad"] + [str(a) for a in anios] + ["Total"]
    for col_idx, h in enumerate(headers, 1):
        engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
    fila += 1

    # Cultivos individuales
    cult_rows = [idx for idx in df_fin.index if idx.startswith('1.1 ')]
    for cult_idx in sorted(cult_rows):
        vals = df_fin.loc[cult_idx, anios].values.tolist()
        tot = df_fin.loc[cult_idx, 'Total']
        nombre_cult = cult_idx[4:]
        engine.escribir_fila_datos(fila, [nombre_cult, "Bs/año"] + vals + [tot])
        fila += 1

    # Total ingresos
    ingresos = df_fin.loc['1. INGRESOS POR VENTAS', anios].values.tolist()
    total_ing = df_fin.loc['1. INGRESOS POR VENTAS', 'Total']
    engine.escribir_fila_datos(fila, ["TOTAL INGRESOS", "Bs/año"] + ingresos + [total_ing],
                              bold=True, fill_color=engine.COLOR_TOTAL)
    fila += 2

    # --- 3. COSTOS DEL PROYECTO ---
    # 3.1 INVERSIÓN EN INFRAESTRUCTURA (Obras Civiles)
    fila = engine.escribir_subtitulo(fila, "3.1 INVERSIÓN EN INFRAESTRUCTURA", n+3)
    inv_items = [
        ('Bienes Transables', '2.1.1 Bienes Transables (Inv)'),
        ('Bienes no Transables', '2.1.2 Bienes no Transables (Inv)'),
        ('Mano de Obra Calificada', '2.1.3 Mano de Obra Calificada (Inv)'),
        ('Mano de Obra Semicalificada', '2.1.4 Mano de Obra Semicalificada (Inv)'),
        ('M.O. No Calificada Urbana', '2.1.5 M.O. No Calificada Urbana (Inv)'),
        ('M.O. No Calificada Rural', '2.1.6 M.O. No Calificada Rural (Inv)'),
    ]
    subtotal_obras = [0.0]*n
    for label, idx_df in inv_items:
        if idx_df in df_fin.index:
            vals = df_fin.loc[idx_df, anios].values.tolist()
            tot = df_fin.loc[idx_df, 'Total']
        else:
            vals = [0]*n; tot = 0
        engine.escribir_fila_datos(fila, [label, "Bs"] + vals + [tot])
        for i in range(n): subtotal_obras[i] += vals[i]
        fila += 1
    engine.escribir_fila_datos(fila, ["Subtotal Obras Civiles", "Bs"] + subtotal_obras + [sum(subtotal_obras)],
                              bold=True, fill_color="E2EFDA")
    fila += 2

    # 3.2 SUPERVISIÓN Y OTROS SERVICIOS (ATI + Supervisión)
    fila = engine.escribir_subtitulo(fila, "3.2 SUPERVISIÓN Y OTROS SERVICIOS", n+3)
    serv_items = [
        ('Bienes Transables', '2.2.1 Bienes Transables (InvServ)'),
        ('Bienes no Transables', '2.2.2 Bienes no Transables (InvServ)'),
        ('Mano de Obra Calificada', '2.2.3 Mano de Obra Calificada (InvServ)'),
        ('Mano de Obra Semicalificada', '2.2.4 Mano de Obra Semicalificada (InvServ)'),
        ('M.O. No Calificada Urbana', '2.2.5 M.O. No Calificada Urbana (InvServ)'),
        ('M.O. No Calificada Rural', '2.2.6 M.O. No Calificada Rural (InvServ)'),
    ]
    subtotal_serv = [0.0]*n
    for label, idx_df in serv_items:
        if idx_df in df_fin.index:
            vals = df_fin.loc[idx_df, anios].values.tolist()
            tot = df_fin.loc[idx_df, 'Total']
        else:
            vals = [0]*n; tot = 0
        engine.escribir_fila_datos(fila, [label, "Bs"] + vals + [tot])
        for i in range(n): subtotal_serv[i] += vals[i]
        fila += 1
    engine.escribir_fila_datos(fila, ["Subtotal Servicios", "Bs"] + subtotal_serv + [sum(subtotal_serv)],
                              bold=True, fill_color="E2EFDA")
    fila += 2

    # 3.3 O&M
    fila = engine.escribir_subtitulo(fila, "3.3 OPERACIÓN Y MANTENIMIENTO", n+3)
    om_items = [
        ('Bienes Transables', '3.1.1 Bienes Transables (O&M)'),
        ('Bienes no Transables', '3.1.2 Bienes no Transables (O&M)'),
        ('Mano de Obra Calificada', '3.1.3 Mano de Obra Calificada (O&M)'),
        ('Mano de Obra Semicalificada', '3.1.4 Mano de Obra Semicalificada (O&M)'),
        ('M.O. No Calificada Urbana', '3.1.5 M.O. No Calificada Urbana (O&M)'),
        ('M.O. No Calificada Rural', '3.1.6 M.O. No Calificada Rural (O&M)'),
    ]
    subtotal_om = [0.0]*n
    for label, idx_df in om_items:
        if idx_df in df_fin.index:
            vals = df_fin.loc[idx_df, anios].values.tolist()
            tot = df_fin.loc[idx_df, 'Total']
        else:
            vals = [0]*n; tot = 0
        engine.escribir_fila_datos(fila, [label, "Bs/año"] + vals + [tot])
        for i in range(n): subtotal_om[i] += vals[i]
        fila += 1
    engine.escribir_fila_datos(fila, ["Subtotal O&M", "Bs/año"] + subtotal_om + [sum(subtotal_om)],
                              bold=True, fill_color="E2EFDA")
    fila += 2

    # 3.4 COSTOS DE PRODUCCIÓN INCREMENTAL
    fila = engine.escribir_subtitulo(fila, "3.4 COSTOS DE PRODUCCIÓN INCREMENTAL", n+3)
    prod_items = [
        ('Bienes Transables', '4.1.1 Bienes Transables (Prod)'),
        ('Bienes no Transables', '4.1.2 Bienes no Transables (Prod)'),
        ('Mano de Obra Calificada', '4.1.3 Mano de Obra Calificada (Prod)'),
        ('Mano de Obra Semicalificada', '4.1.4 Mano de Obra Semicalificada (Prod)'),
        ('M.O. No Calificada Urbana', '4.1.5 M.O. No Calificada Urbana (Prod)'),
        ('M.O. No Calificada Rural', '4.1.6 M.O. No Calificada Rural (Prod)'),
    ]
    subtotal_prod = [0.0]*n
    for label, idx_df in prod_items:
        if idx_df in df_fin.index:
            vals = df_fin.loc[idx_df, anios].values.tolist()
            tot = df_fin.loc[idx_df, 'Total']
        else:
            vals = [0]*n; tot = 0
        engine.escribir_fila_datos(fila, [label, "Bs/año"] + vals + [tot])
        for i in range(n): subtotal_prod[i] += vals[i]
        fila += 1
    engine.escribir_fila_datos(fila, ["Subtotal Producción", "Bs/año"] + subtotal_prod + [sum(subtotal_prod)],
                              bold=True, fill_color="E2EFDA")
    fila += 2

    # 3.5 MITIGACIÓN AMBIENTAL Y GESTIÓN DE RIESGOS
    fila = engine.escribir_subtitulo(fila, "3.5 MITIGACIÓN AMBIENTAL Y GESTIÓN DE RIESGOS", n+3)
    amb_vals = df_fin.loc['6. TOTAL COSTOS AMBIENTALES', anios].values.tolist() if '6. TOTAL COSTOS AMBIENTALES' in df_fin.index else [0]*n
    amb_tot = df_fin.loc['6. TOTAL COSTOS AMBIENTALES', 'Total'] if '6. TOTAL COSTOS AMBIENTALES' in df_fin.index else 0
    engine.escribir_fila_datos(fila, ["Costo de Mitigación / Manejo Ambiental", "Bs"] + amb_vals + [amb_tot])
    fila += 1
    engine.escribir_fila_datos(fila, ["Subtotal Ambiental", "Bs"] + amb_vals + [amb_tot],
                              bold=True, fill_color="E2EFDA")
    fila += 2

    # 3.6 TOTAL COSTOS
    total_costos_vals = df_fin.loc['7. TOTAL COSTOS', anios].values.tolist() if '7. TOTAL COSTOS' in df_fin.index else [0]*n
    total_costos_tot = df_fin.loc['7. TOTAL COSTOS', 'Total'] if '7. TOTAL COSTOS' in df_fin.index else 0
    engine.escribir_fila_datos(fila, ["TOTAL COSTOS DEL PROYECTO", "Bs/año"] + total_costos_vals + [total_costos_tot],
                              bold=True, fill_color=engine.COLOR_HEADER)
    fila += 3

    # --- 4. FLUJO DE FONDOS NETO FINANCIERO ---
    fila = engine.escribir_titulo_seccion(fila, "4. FLUJO DE FONDOS NETO FINANCIERO", n+3)
    for col_idx, h in enumerate(headers, 1):
        engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
    fila += 1

    # Extraer flujos base
    flujo_inv = df_fin.loc['2. TOTAL COSTOS DE INVERSIÓN', anios].values.tolist() if '2. TOTAL COSTOS DE INVERSIÓN' in df_fin.index else [0]*n
    flujo_om = df_fin.loc['3. TOTAL COSTOS DE OPERACIÓN', anios].values.tolist() if '3. TOTAL COSTOS DE OPERACIÓN' in df_fin.index else [0]*n
    flujo_prod = df_fin.loc['4. TOTAL COSTOS DE PRODUCCIÓN', anios].values.tolist() if '4. TOTAL COSTOS DE PRODUCCIÓN' in df_fin.index else [0]*n
    flujo_amb = df_fin.loc['6. TOTAL COSTOS AMBIENTALES', anios].values.tolist() if '6. TOTAL COSTOS AMBIENTALES' in df_fin.index else [0]*n
    salvamento = [0.0]*n
    if n > 0:
        salvamento[-1] = config.valor_salvamento if hasattr(config, 'valor_salvamento') else 0.0

    # NUEVO: Préstamo e Intereses desde configuración
    prestamo = [0.0]*n
    intereses = [0.0]*n
    if hasattr(config, 'prestamo') and config.prestamo != 0:
        prestamo[0] = config.prestamo  # Desembolso en año 0
    if hasattr(config, 'costo_financiero') and config.costo_financiero != 0:
        intereses = [config.costo_financiero] * n

    # Calcular flujo neto
    flujo_neto = [0.0]*n
    for i in range(n):
        flujo_neto[i] = (ingresos[i] - flujo_inv[i] - flujo_om[i] -
                        flujo_prod[i] - flujo_amb[i] + salvamento[i] +
                        prestamo[i] - intereses[i])

    filas_flujo = [
        ("(+) Ingresos por Ventas (VNPI)", ingresos, False),
        ("(-) Inversión en Infraestructura", [-v for v in flujo_inv], False),
        ("(-) Operación y Mantenimiento", [-v for v in flujo_om], False),
        ("(-) Costos de Producción", [-v for v in flujo_prod], False),
        ("(-) Mitigación Ambiental", [-v for v in flujo_amb], False),
        ("(+) Valor de Salvamento", salvamento, False),
        ("(+) Préstamo", prestamo, False),
        ("(-) Intereses del Préstamo", [-v for v in intereses], False),
        ("(=) FLUJO DE FONDOS NETO", flujo_neto, True),
    ]

    for label, vals, is_total in filas_flujo:
        display_vals = [v if v != 0 else 0 for v in vals]
        engine.escribir_fila_datos(fila, [label, "Bs/año"] + display_vals + [sum(display_vals)],
                                  bold=is_total,
                                  fill_color=engine.COLOR_TOTAL if is_total else None)
        fila += 1
    fila += 2

    # --- 5. INDICADORES FINANCIEROS ---
    fila = engine.escribir_titulo_seccion(fila, "5. INDICADORES FINANCIEROS", 4)
    inds = [
        ("Valor Presente Neto Financiero (VANF)", ind_fin['VAN'], "#,##0.00"),
        ("Tasa Interna de Retorno Financiera (TIRF)", ind_fin['TIR']*1 if ind_fin['TIR'] else 0, "0.00%"),
        ("Relación Beneficio / Costo (B/C)", ind_fin['Relacion_BC'], "0.0000"),
        ("Costo Anual Equivalente Financiero (CAEF)", ind_fin['CAE'], "#,##0.00"),
    ]
    for label, val, fmt in inds:
        engine.escribir_celda(fila, 1, label, bold=True, align_h='left')
        cell = engine.escribir_celda(fila, 2, val, align_h='right')
        cell.number_format = fmt
        fila += 1

    # Nota metodológica
    fila += 2
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n+3)
    engine.escribir_celda(fila, 1,
        "Nota: La evaluación financiera utiliza precios de mercado. No se incluyen impuestos, depreciaciones ni servicio de deuda "
        "al tratarse de un proyecto de inversión pública en riego (RM 115/2015).",
        size=9, align_h='left', color="666666")

# ============================================================
# GENERADOR EXCEL HOJA E_ECO — VERSIÓN PROFESIONAL
# ============================================================
def _generar_hoja_eeco_engine(engine: ExcelReportEngine, config, df_eco, ind_eco):
    """Genera hoja E_ECO con estructura profesional y matriz RPC visible."""
    anios = config.generar_anios()
    n = len(anios)
    ws = engine.ws_active

    # --- TÍTULO ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+3)
    engine.escribir_celda(1, 1, "EVALUACIÓN ECONÓMICA (Precios de Cuenta / Sombra)",
                         bold=True, size=14, color=engine.COLOR_TITULO, align_h='center')
    engine.escribir_celda(2, 1, f"Proyecto: {config.nombre}  |  Horizonte: {anios[0]} - {anios[-1]}  |  TSD: {config.tasa_social_descuento*100:.2f}%",
                         size=10, align_h='center')
    fila = 4

    # --- 1. MATRIZ DE RAZONES PRECIO CUENTA (RPC) ---
    fila = engine.escribir_titulo_seccion(fila, "1. FACTORES DE CORRECCIÓN RPC — VIPFE", 4)
    rpc_data = [
        ("Recurso", "RPC", "Precio Económico = Financiero × RPC", "Aplicación"),
        ("Divisas / Bienes Transables", config.rpc['divisa'], "Insumos importados o exportables", "Inversión, O&M, Producción"),
        ("Bienes No Transables / Locales", 1.00, "Materiales de origen nacional", "Inversión, O&M, Producción"),
        ("Mano de Obra Calificada", config.rpc['mo_calificada'], "Costo social alta calificación", "Inversión, O&M, Producción"),
        ("Mano de Obra Semicalificada", config.rpc['mo_semicalificada'], "Costo social media calif.", "Inversión, O&M, Producción"),
        ("M.O. No Calificada Urbana", config.rpc['mo_no_calif_urbana'], "Costo social MO urbana", "Inversión, O&M, Producción"),
        ("M.O. No Calificada Rural", config.rpc['mo_no_calif_rural'], "Costo social MO rural", "Inversión, O&M, Producción"),
    ]
    for i, row_data in enumerate(rpc_data):
        is_header = (i == 0)
        for col_idx, val in enumerate(row_data, 1):
            cell = engine.escribir_celda(fila, col_idx, val, bold=is_header,
                                          fill_color=engine.COLOR_HEADER if is_header else None,
                                          align_h='center' if is_header else 'left')
            if col_idx == 2 and not is_header and isinstance(val, float):
                cell.number_format = '0.00'
        fila += 1
    fila += 1

    # Factor de corrección de ingresos
    factor_ing = (1 - config.pct_produccion_transable) + (config.pct_produccion_transable * config.rpc['divisa'])
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    engine.escribir_celda(fila, 1,
        f"Factor de corrección a Ingresos (producción transable {config.pct_produccion_transable*100:.0f}%): {factor_ing:.4f}",
        bold=True, fill_color="FFF2CC", align_h='left')
    fila += 2

    # --- 2. BENEFICIOS ECONÓMICOS ---
    fila = engine.escribir_titulo_seccion(fila, "2. BENEFICIOS ECONÓMICOS INCREMENTALES", n+3)
    headers = ["Concepto", "Unidad"] + [str(a) for a in anios] + ["Total"]
    for col_idx, h in enumerate(headers, 1):
        engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
    fila += 1

    cult_rows = [idx for idx in df_eco.index if idx.startswith('1.1 ')]
    for cult_idx in sorted(cult_rows):
        vals = df_eco.loc[cult_idx, anios].values.tolist()
        tot = df_eco.loc[cult_idx, 'Total']
        nombre_cult = cult_idx[4:]
        engine.escribir_fila_datos(fila, [nombre_cult, "Bs/año"] + vals + [tot])
        fila += 1

    ingresos_eco = df_eco.loc['1. INGRESOS POR VENTAS', anios].values.tolist()
    total_ing_eco = df_eco.loc['1. INGRESOS POR VENTAS', 'Total']
    engine.escribir_fila_datos(fila, ["TOTAL BENEFICIOS ECONÓMICOS", "Bs/año"] + ingresos_eco + [total_ing_eco],
                              bold=True, fill_color=engine.COLOR_TOTAL)
    fila += 2

    # --- 3. COSTOS ECONÓMICOS ---
    # 3.1 Inversión Económica Ajustada (Obras Civiles)
    fila = engine.escribir_subtitulo(fila, "3.1 INVERSIÓN ECONÓMICA AJUSTADA (OBRAS CIVILES)", n+3)
    obras_items_eco = [
        ('Bienes Transables', '2.1.1 Bienes Transables (Inv)'),
        ('Materiales Locales', '2.1.2 Bienes no Transables (Inv)'),
        ('Mano de Obra Calificada', '2.1.3 Mano de Obra Calificada (Inv)'),
        ('Mano de Obra Semicalificada', '2.1.4 Mano de Obra Semicalificada (Inv)'),
        ('M.O. No Calificada Urbana', '2.1.5 M.O. No Calificada Urbana (Inv)'),
        ('M.O. No Calificada Rural', '2.1.6 M.O. No Calificada Rural (Inv)'),
    ]
    subtotal_obras_eco = [0.0]*n
    for label, idx_df in obras_items_eco:
        if idx_df in df_eco.index:
            vals = df_eco.loc[idx_df, anios].values.tolist()
            tot = df_eco.loc[idx_df, 'Total']
        else:
            vals = [0]*n; tot = 0
        engine.escribir_fila_datos(fila, [label, "Bs"] + vals + [tot])
        for i in range(n): subtotal_obras_eco[i] += vals[i]
        fila += 1
    engine.escribir_fila_datos(fila, ["Subtotal Obras Civiles Económico", "Bs"] + subtotal_obras_eco + [sum(subtotal_obras_eco)],
                              bold=True, fill_color="E2EFDA")
    fila += 2

    # 3.2 SUPERVISIÓN Y OTROS SERVICIOS ECONÓMICOS
    fila = engine.escribir_subtitulo(fila, "3.2 SUPERVISIÓN Y OTROS SERVICIOS ECONÓMICOS", n+3)
    serv_items_eco = [
        ('Bienes Transables', '2.2.1 Bienes Transables (InvServ)'),
        ('Materiales Locales', '2.2.2 Bienes no Transables (InvServ)'),
        ('Mano de Obra Calificada', '2.2.3 Mano de Obra Calificada (InvServ)'),
        ('Mano de Obra Semicalificada', '2.2.4 Mano de Obra Semicalificada (InvServ)'),
        ('M.O. No Calificada Urbana', '2.2.5 M.O. No Calificada Urbana (InvServ)'),
        ('M.O. No Calificada Rural', '2.2.6 M.O. No Calificada Rural (InvServ)'),
    ]
    subtotal_serv_eco = [0.0]*n
    for label, idx_df in serv_items_eco:
        if idx_df in df_eco.index:
            vals = df_eco.loc[idx_df, anios].values.tolist()
            tot = df_eco.loc[idx_df, 'Total']
        else:
            vals = [0]*n; tot = 0
        engine.escribir_fila_datos(fila, [label, "Bs"] + vals + [tot])
        for i in range(n): subtotal_serv_eco[i] += vals[i]
        fila += 1
    engine.escribir_fila_datos(fila, ["Subtotal Servicios Económico", "Bs"] + subtotal_serv_eco + [sum(subtotal_serv_eco)],
                              bold=True, fill_color="E2EFDA")
    fila += 2

    # 3.3 O&M Económico
    fila = engine.escribir_subtitulo(fila, "3.3 OPERACIÓN Y MANTENIMIENTO ECONÓMICO", n+3)
    om_items_eco = [
        ('Bienes Transables', '3.1.1 Bienes Transables (O&M)'),
        ('Materiales Locales', '3.1.2 Bienes no Transables (O&M)'),
        ('Mano de Obra Calificada', '3.1.3 Mano de Obra Calificada (O&M)'),
        ('Mano de Obra Semicalificada', '3.1.4 Mano de Obra Semicalificada (O&M)'),
        ('M.O. No Calificada Urbana', '3.1.5 M.O. No Calificada Urbana (O&M)'),
        ('M.O. No Calificada Rural', '3.1.6 M.O. No Calificada Rural (O&M)'),
    ]
    subtotal_om_eco = [0.0]*n
    for label, idx_df in om_items_eco:
        if idx_df in df_eco.index:
            vals = df_eco.loc[idx_df, anios].values.tolist()
            tot = df_eco.loc[idx_df, 'Total']
        else:
            vals = [0]*n; tot = 0
        engine.escribir_fila_datos(fila, [label, "Bs/año"] + vals + [tot])
        for i in range(n): subtotal_om_eco[i] += vals[i]
        fila += 1
    engine.escribir_fila_datos(fila, ["Subtotal O&M Económico", "Bs/año"] + subtotal_om_eco + [sum(subtotal_om_eco)],
                              bold=True, fill_color="E2EFDA")
    fila += 2

    # 3.4 Producción Económica
    fila = engine.escribir_subtitulo(fila, "3.4 COSTOS DE PRODUCCIÓN ECONÓMICOS", n+3)
    prod_items_eco = [
        ('Bienes Transables', '4.1.1 Bienes Transables (Prod)'),
        ('Materiales Locales', '4.1.2 Bienes no Transables (Prod)'),
        ('Mano de Obra Calificada', '4.1.3 Mano de Obra Calificada (Prod)'),
        ('Mano de Obra Semicalificada', '4.1.4 Mano de Obra Semicalificada (Prod)'),
        ('M.O. No Calificada Urbana', '4.1.5 M.O. No Calificada Urbana (Prod)'),
        ('M.O. No Calificada Rural', '4.1.6 M.O. No Calificada Rural (Prod)'),
    ]
    subtotal_prod_eco = [0.0]*n
    for label, idx_df in prod_items_eco:
        if idx_df in df_eco.index:
            vals = df_eco.loc[idx_df, anios].values.tolist()
            tot = df_eco.loc[idx_df, 'Total']
        else:
            vals = [0]*n; tot = 0
        engine.escribir_fila_datos(fila, [label, "Bs/año"] + vals + [tot])
        for i in range(n): subtotal_prod_eco[i] += vals[i]
        fila += 1
    engine.escribir_fila_datos(fila, ["Subtotal Producción Económica", "Bs/año"] + subtotal_prod_eco + [sum(subtotal_prod_eco)],
                              bold=True, fill_color="E2EFDA")
    fila += 2

    # 3.5 Ambiental Económico
    fila = engine.escribir_subtitulo(fila, "3.5 MITIGACIÓN AMBIENTAL Y GESTIÓN DE RIESGOS", n+3)
    amb_vals_eco = df_eco.loc['6. TOTAL COSTOS AMBIENTALES', anios].values.tolist() if '6. TOTAL COSTOS AMBIENTALES' in df_eco.index else [0]*n
    amb_tot_eco = df_eco.loc['6. TOTAL COSTOS AMBIENTALES', 'Total'] if '6. TOTAL COSTOS AMBIENTALES' in df_eco.index else 0
    engine.escribir_fila_datos(fila, ["Mitigación Ambiental / RRD+ACC", "Bs"] + amb_vals_eco + [amb_tot_eco])
    fila += 1
    engine.escribir_fila_datos(fila, ["Subtotal Ambiental Económico", "Bs"] + amb_vals_eco + [amb_tot_eco],
                              bold=True, fill_color="E2EFDA")
    fila += 2

    # 3.6 Total Costos Económicos
    total_costos_eco_vals = df_eco.loc['7. TOTAL COSTOS', anios].values.tolist() if '7. TOTAL COSTOS' in df_eco.index else [0]*n
    total_costos_eco_tot = df_eco.loc['7. TOTAL COSTOS', 'Total'] if '7. TOTAL COSTOS' in df_eco.index else 0
    engine.escribir_fila_datos(fila, ["TOTAL COSTOS ECONÓMICOS", "Bs/año"] + total_costos_eco_vals + [total_costos_eco_tot],
                              bold=True, fill_color=engine.COLOR_HEADER)
    fila += 3

    # --- 4. FLUJO DE FONDOS NETO ECONÓMICO ---
    fila = engine.escribir_titulo_seccion(fila, "4. FLUJO DE FONDOS NETO ECONÓMICO", n+3)
    for col_idx, h in enumerate(headers, 1):
        engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
    fila += 1

    salvamento_eco = [0.0]*n
    if n > 0:
        factor_salv = 0.85
        salvamento_eco[-1] = (config.valor_salvamento if hasattr(config, 'valor_salvamento') else 0.0) * factor_salv

    # NUEVO: Préstamo e Intereses desde configuración
    prestamo_eco = [0.0]*n
    intereses_eco = [0.0]*n
    if hasattr(config, 'prestamo') and config.prestamo != 0:
        prestamo_eco[0] = config.prestamo
    if hasattr(config, 'costo_financiero') and config.costo_financiero != 0:
        intereses_eco = [config.costo_financiero] * n

    flujo_neto_eco = [0.0]*n
    for i in range(n):
        flujo_neto_eco[i] = (ingresos_eco[i] - subtotal_obras_eco[i] - subtotal_serv_eco[i] -
                            subtotal_om_eco[i] - subtotal_prod_eco[i] - amb_vals_eco[i] +
                            salvamento_eco[i] + prestamo_eco[i] - intereses_eco[i])

    filas_flujo_eco = [
        ("(+) Beneficios Económicos Incrementales", ingresos_eco, False),
        ("(-) Inversión Económica Ajustada (Obras Civiles)", [-v for v in subtotal_obras_eco], False),
        ("(-) Supervisión y Servicios Económicos", [-v for v in subtotal_serv_eco], False),
        ("(-) O&M Económico", [-v for v in subtotal_om_eco], False),
        ("(-) Producción Económica", [-v for v in subtotal_prod_eco], False),
        ("(-) Mitigación Ambiental", [-v for v in amb_vals_eco], False),
        ("(+) Valor de Salvamento Económico", salvamento_eco, False),
        ("(+) Préstamo", prestamo_eco, False),
        ("(-) Intereses del Préstamo", [-v for v in intereses_eco], False),
        ("(=) FLUJO DE FONDOS NETO ECONÓMICO", flujo_neto_eco, True),
    ]

    for label, vals, is_total in filas_flujo_eco:
        display_vals = [v if v != 0 else 0 for v in vals]
        engine.escribir_fila_datos(fila, [label, "Bs/año"] + display_vals + [sum(display_vals)],
                                  bold=is_total,
                                  fill_color=engine.COLOR_TOTAL if is_total else None)
        fila += 1
    fila += 2
    
    # --- 5. INDICADORES ECONÓMICOS ---
    fila = engine.escribir_titulo_seccion(fila, "5. INDICADORES ECONÓMICOS Y SOCIALES", 4)
    inds_eco = [
        ("Valor Presente Neto Económico (VANE)", ind_eco['VAN'], "#,##0.00"),
        ("Tasa Interna de Retorno Económica (TIRE)", ind_eco['TIR']*1 if ind_eco['TIR'] else 0, "0.00%"),
        ("Relación Beneficio / Costo Económica (B/CE)", ind_eco['Relacion_BC'], "0.0000"),
        ("Costo Anual Equivalente Económico (CAEE)", ind_eco['CAE'], "#,##0.00"),
        ("Valor Actualizado de Ingresos (VAIS)", ind_eco['VA_Ingresos'], "#,##0.00"),
        ("Valor Actualizado de Costos (VACS)", ind_eco['VA_Costos'], "#,##0.00"),
    ]
    for label, val, fmt in inds_eco:
        engine.escribir_celda(fila, 1, label, bold=True, align_h='left')
        cell = engine.escribir_celda(fila, 2, val, align_h='right')
        cell.number_format = fmt
        fila += 1

    # Nota metodológica
    fila += 2
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n+3)
    engine.escribir_celda(fila, 1,
        "Nota: La evaluación económica utiliza precios de cuenta (precios sombra) según RPC oficial VIPFE. "
        "Los beneficios se corrigen por el porcentaje de producción transable. RM 115/2015, Art. 20.",
        size=9, align_h='left', color="666666")

# ============================================================
# GENERADOR EXCEL HOJA SOSTENIBILIDAD
# ============================================================
def _generar_hoja_sostenibilidad_engine(engine: ExcelReportEngine, config, df_fin):
    """Genera hoja Sostenibilidad Operativa."""
    anios = config.generar_anios()
    n = len(anios)
    ind_sost = calcular_sostenibilidad_operativa(config, df_fin, anios)
    
    # Título
    ws = engine.ws_active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    engine.escribir_celda(1, 1, "SOSTENIBILIDAD OPERATIVA", bold=True, size=14,
                         color=engine.COLOR_TITULO, align_h='center')
    
    # Subtítulo normativo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    engine.escribir_celda(2, 1, "Art. 16.9 RM 115/2015 — Verificación de cobertura de costos de operación y mantenimiento",
                         size=10, align_h='center')
    
    fila = 4
    
    # Indicadores agregados
    fila = engine.escribir_subtitulo(fila, "INDICADORES DE SOSTENIBILIDAD", 2)
    
    indicadores = [
        ("Ratio de Cobertura Promedio", f"{ind_sost['ratio_cobertura_promedio']:.2f}x",
         "Ingresos / Costos operativos promedio"),
        ("Año de Punto de Equilibrio", 
         str(ind_sost['anio_equilibrio']) if ind_sost['anio_equilibrio'] else "No alcanza",
         "Primer año con superávit operativo ≥ 0"),
        ("Años de Déficit Inicial", str(ind_sost['anos_deficit_consecutivos']),
         "Años consecutivos con pérdida desde inicio de operación"),
        ("Costo Operativo Promedio Anual", f"Bs {ind_sost['costo_operativo_promedio']:,.0f}",
         "Promedio de O&M + Producción + Comercialización"),
    ]
    
    for label, val, desc in indicadores:
        engine.escribir_celda(fila, 1, label, bold=True, align_h='left')
        engine.escribir_celda(fila, 2, val, align_h='right')
        engine.escribir_celda(fila, 3, desc, align_h='left')
        fila += 1
    
    fila += 1
    
    # Tabla anual detallada
    fila = engine.escribir_subtitulo(fila, "FLUJO OPERATIVO ANUAL", 2)
    
    headers = ["Año", "Ingresos (Bs)", "Costos Operativos (Bs)", 
               "Superávit / Déficit (Bs)", "Ratio Cobertura"]
    for col_idx, h in enumerate(headers, 1):
        engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
    fila += 1
    
    for i, anio in enumerate(anios):
        vals = [
            anio,
            ind_sost['ingresos'][i],
            ind_sost['costos_operativos'][i],
            ind_sost['superavit_operativo'][i],
            ind_sost['ratio_cobertura'][i] if not np.isnan(ind_sost['ratio_cobertura'][i]) else "-"
        ]
        
        # Color condicional para superávit
        fill = None
        if ind_sost['superavit_operativo'][i] > 0:
            fill = "D4EDDA"  # verde claro
        elif ind_sost['superavit_operativo'][i] < 0:
            fill = "F8D7DA"  # rojo claro
        
        engine.escribir_fila_datos(fila, vals, fill_color=fill)
        fila += 1
    
    fila += 1
    
    # Conclusión
    fila = engine.escribir_subtitulo(fila, "CONCLUSIÓN", 2)
    if ind_sost['sostenible']:
        if ind_sost['anos_deficit_consecutivos'] == 0:
            conclusion = "SOSTENIBILIDAD PLENA: El proyecto genera superávit operativo desde el primer año."
        else:
            conclusion = (f"SOSTENIBILIDAD DIFERIDA: Equilibrio alcanzado en {ind_sost['anio_equilibrio']}. "
                         f"Requiere {ind_sost['anos_deficit_consecutivos']} años de apoyo inicial.")
    else:
        conclusion = "INSOSTENIBLE OPERATIVAMENTE: Los ingresos no cubren costos operativos en ningún año."
    
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    engine.escribir_celda(fila, 1, conclusion, bold=True, align_h='left')

# ============================================================
# GENERADOR EXCEL HOJA INDICADORES
# ============================================================
def _generar_hoja_indicadores_engine(engine, config, ind_fin, ind_eco, datos_puente):
    """Genera hoja Indicadores con el motor genérico."""
    fila = 1
    ws = engine.ws_active
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    engine.escribir_celda(fila, 1, "INDICADORES FINANCIEROS Y ECONÓMICOS",
                         bold=True, size=14, align_h='center')
    fila += 2
    
    # --- INDICADORES FINANCIEROS ---
    fila = engine.escribir_indicadores(fila, {
        "VANF": ind_fin['VAN'],
        "TIRF (%)": ind_fin['TIR']*100 if ind_fin['TIR'] else None,
        "Relación B/C Financiero": ind_fin['Relacion_BC'],
        "CAEF": ind_fin['CAE']
    }, "INDICADORES FINANCIEROS")
    fila += 1
    
    # --- INDICADORES ECONÓMICOS ---
    fila = engine.escribir_indicadores(fila, {
        "VANE": ind_eco['VAN'],
        "TIRE (%)": ind_eco['TIR']*100 if ind_eco['TIR'] else None,
        "Relación B/C Económico": ind_eco['Relacion_BC'],
        "CAEE": ind_eco['CAE']
    }, "INDICADORES ECONÓMICOS")
    fila += 1
    
    # --- INDICADORES DE COSTO-EFICIENCIA ---
    area_inc = config.area_incremental if config.area_incremental > 0 else 1
    familias = config.total_familias if config.total_familias > 0 else 1
    poblacion = config.poblacion_base if config.poblacion_base > 0 else 1
    
    # CORRECCIÓN: acceder a través de datos_puente['inversion']
    inv_total = datos_puente['inversion']['inversion_total']
    
    ce_data = {
        'Indicador': [
            'Inversión por Hectárea',
            'Inversión por Familia',
            'CAEF por Hectárea',
            'CAEE por Hectárea',
            'CAEF por Población',
            'CAEE por Población',
        ],
        'Valor': [
            inv_total / area_inc,
            inv_total / familias,
            ind_fin['CAE'] / area_inc,
            ind_eco['CAE'] / area_inc,
            ind_fin['CAE'] / poblacion,
            ind_eco['CAE'] / poblacion,
        ],
        'Unidad': [
            'Bs/Ha',
            'Bs/Familia',
            'Bs/Ha/año',
            'Bs/Ha/año',
            'Bs/hab/año',
            'Bs/hab/año',
        ]
    }
    
    fila = engine.escribir_subtitulo(fila, "INDICADORES DE COSTO-EFICIENCIA", 3)
    
    # Headers
    for col_idx, h in enumerate(['Indicador', 'Valor', 'Unidad'], 1):
        engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
    fila += 1
    
    # Datos
    for i in range(len(ce_data['Indicador'])):
        engine.escribir_celda(fila, 1, ce_data['Indicador'][i], align_h='left')
        cell = engine.escribir_celda(fila, 2, ce_data['Valor'][i], align_h='right')
        cell.number_format = '#,##0.00'
        engine.escribir_celda(fila, 3, ce_data['Unidad'][i], align_h='left')
        fila += 1
# ============================================================   
# GENERADOR EXCEL HOJA SENSIBILIDAD
# ============================================================
def _generar_hoja_sensibilidad_engine(engine, config, datos_puente, df_fin_base,
                                      ind_fin_base, df_eco_base, ind_eco_base):
    """Genera hoja Sensibilidad sin gráficos — solo tablas profesionales."""
    fila = 1
    ws = engine.ws_active
    anios = config.generar_anios()

    # --- TÍTULO ---
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=10)
    engine.escribir_celda(fila, 1, "ANÁLISIS DE SENSIBILIDAD",
                         bold=True, size=14, color=engine.COLOR_TITULO, align_h='center')
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=10)
    engine.escribir_celda(fila, 1,
        f"Proyecto: {config.nombre}  |  TPD: {config.tasa_privada_descuento*100:.2f}%  |  TSD: {config.tasa_social_descuento*100:.2f}%",
        size=10, align_h='center')
    fila += 2

    # --- 1. ESCENARIOS ESTÁNDAR ---
    fila = engine.escribir_titulo_seccion(fila, "1. ESCENARIOS ESTÁNDAR DE SENSIBILIDAD", 10)
    fila += 1

    escenarios = [
        ("Caso Base", 0.0, 0.0, 0.0, 0.0, 0.0, 0.0),
        ("Inversión +20%", 0.2, 0.0, 0.0, 0.0, 0.0, 0.0),
        ("Inversión -20%", -0.2, 0.0, 0.0, 0.0, 0.0, 0.0),
        ("Costos O&M +20%", 0.0, 0.2, 0.0, 0.0, 0.0, 0.0),
        ("Costos Producción +20%", 0.0, 0.0, 0.2, 0.0, 0.0, 0.0),
        ("Ingresos -20%", 0.0, 0.0, 0.0, -0.2, 0.0, 0.0),
        ("Ingresos +20%", 0.0, 0.0, 0.0, 0.2, 0.0, 0.0),
        ("Tasa Privada +2 pp", 0.0, 0.0, 0.0, 0.0, 0.02, 0.0),
        ("Tasa Social +2 pp", 0.0, 0.0, 0.0, 0.0, 0.0, 0.02),
        ("Combinado Pesimista", 0.2, 0.2, 0.2, -0.2, 0.02, 0.02),
    ]

    headers = ["Escenario", "VANF (Bs)", "TIRF (%)", "B/C Fin", "CAEF (Bs)",
               "VANE (Bs)", "TIRE (%)", "B/C Eco", "CAEE (Bs)"]
    for col_idx, h in enumerate(headers, 1):
        engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
    fila += 1

    for nombre, vi, vom, vp, ving, vtp, vts in escenarios:
        df_fin, df_eco, ind_fin, ind_eco = recalcular_con_variacion(
            config, datos_puente, vi, vom, vp, ving, vtp, vts
        )
        fila_datos = [
            nombre,
            ind_fin['VAN'],
            ind_fin['TIR'] * 100 if ind_fin['TIR'] else 0,
            ind_fin['Relacion_BC'],
            ind_fin['CAE'],
            ind_eco['VAN'],
            ind_eco['TIR'] * 100 if ind_eco['TIR'] else 0,
            ind_eco['Relacion_BC'],
            ind_eco['CAE'],
        ]
        for col_idx, val in enumerate(fila_datos, 1):
            is_bold = (col_idx == 1)
            nf = '#,##0.00' if col_idx > 1 and isinstance(val, (int, float)) else None
            engine.escribir_celda(fila, col_idx, val, bold=is_bold, number_format=nf)
        fila += 1

    fila += 2

    # --- 2. IMPACTO POR VARIABLE INDIVIDUAL (reemplaza al gráfico tornado) ---
    fila = engine.escribir_titulo_seccion(fila, "2. IMPACTO INDIVIDUAL SOBRE EL VANF", 6)
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
    engine.escribir_celda(fila, 1,
        "Cada fila muestra el efecto aislado de modificar una sola variable (las demás permanecen en valor base).",
        size=9, align_h='left', color="666666")
    fila += 1

    headers2 = ["Variable Modificada", "Variación", "VANF Resultado (Bs)", "Δ VANF (Bs)",
                "Δ VANF (%)", "Umbral de Quiebre"]
    for col_idx, h in enumerate(headers2, 1):
        engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
    fila += 1

    vanf_base = ind_fin_base['VAN']

    # Definir variaciones individuales a evaluar
    variaciones_individuales = [
        ("Inversión", 0.20, lambda c, d, v: recalcular_con_variacion(c, d, v, 0, 0, 0, 0, 0)),
        ("Inversión", -0.20, lambda c, d, v: recalcular_con_variacion(c, d, v, 0, 0, 0, 0, 0)),
        ("Costos O&M", 0.20, lambda c, d, v: recalcular_con_variacion(c, d, 0, v, 0, 0, 0, 0)),
        ("Costos Producción", 0.20, lambda c, d, v: recalcular_con_variacion(c, d, 0, 0, v, 0, 0, 0)),
        ("Ingresos", -0.20, lambda c, d, v: recalcular_con_variacion(c, d, 0, 0, 0, v, 0, 0)),
        ("Ingresos", 0.20, lambda c, d, v: recalcular_con_variacion(c, d, 0, 0, 0, v, 0, 0)),
        ("Tasa Privada Descuento", 0.02, lambda c, d, v: recalcular_con_variacion(c, d, 0, 0, 0, 0, v, 0)),
        ("Tasa Social Descuento", 0.02, lambda c, d, v: recalcular_con_variacion(c, d, 0, 0, 0, 0, 0, v)),
    ]

    for nombre_var, var_val, recalc_fn in variaciones_individuales:
        _, _, ind_fin_var, _ = recalc_fn(config, datos_puente, var_val)
        vanf_var = ind_fin_var['VAN']
        delta = vanf_var - vanf_base
        delta_pct = (delta / abs(vanf_base) * 100) if vanf_base != 0 else 0

        # Umbral de quiebre: cuánto más puede cambiar esta variable antes de VAN=0
        # Aproximación lineal: si ΔVAN = m·Δvar, entonces quiebre adicional = -VAN_base / m
        # donde m = delta / var_val. Quiebre total desde base = -VAN_base / (delta/var_val)
        if delta != 0:
            pendiente = delta / var_val  # cambio de VAN por unidad de variación
            quiebre = (-vanf_base / pendiente) if pendiente != 0 else "N/A"
            if isinstance(quiebre, float):
                signo = "+" if quiebre >= 0 else ""
                umbral_str = f"{signo}{quiebre*100:.1f} pp" if "Tasa" in nombre_var else f"{signo}{quiebre*100:.1f}%"
            else:
                umbral_str = "N/A"
        else:
            umbral_str = "N/A"

        fila_vals = [
            nombre_var,
            f"+{var_val*100:.0f}%" if var_val > 0 else f"{var_val*100:.0f}%",
            vanf_var,
            delta,
            delta_pct,
            umbral_str,
        ]
        for col_idx, val in enumerate(fila_vals, 1):
            nf = '#,##0.00' if col_idx in [3, 4, 5] and isinstance(val, (int, float)) else None
            is_neg = isinstance(val, (int, float)) and val < 0
            color = "F8D7DA" if is_neg and col_idx in [4, 5] else None
            engine.escribir_celda(fila, col_idx, val, number_format=nf, fill_color=color)
        fila += 1

    fila += 2

    # --- 3. ESCENARIO PERSONALIZADO (si existe en la interfaz) ---
    if 'sensibilidad_personalizada' in st.session_state:
        sp = st.session_state['sensibilidad_personalizada']
        fila = engine.escribir_titulo_seccion(fila, "3. ESCENARIO PERSONALIZADO DESDE INTERFAZ", 5)
        fila += 1

        # Parámetros aplicados
        params_aplicados = [
            ("Variación Inversión", f"{sp.get('var_inv', 0)*100:.1f}%"),
            ("Variación O&M", f"{sp.get('var_om', 0)*100:.1f}%"),
            ("Variación Producción", f"{sp.get('var_prod', 0)*100:.1f}%"),
            ("Variación Ingresos", f"{sp.get('var_ing', 0)*100:.1f}%"),
            ("Variación Tasa Privada", f"{sp.get('var_tpriv', 0)*100:.2f} pp"),
            ("Variación Tasa Social", f"{sp.get('var_tsoc', 0)*100:.2f} pp"),
        ]
        for label, val in params_aplicados:
            engine.escribir_celda(fila, 1, label, bold=True, align_h='left')
            engine.escribir_celda(fila, 2, val, align_h='left')
            fila += 1
        fila += 1

        # Resultados personalizados
        headers3 = ["Indicador", "Valor Base", "Valor Modificado", "Variación"]
        for col_idx, h in enumerate(headers3, 1):
            engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
        fila += 1

        indicadores_comp = [
            ("VAN Financiero (Bs)", sp.get('vanf_base', 0), sp.get('vanf_mod', 0)),
            ("TIR Financiera (%)", sp.get('tirf_base', 0), sp.get('tirf_mod', 0)),
            ("B/C Financiero", sp.get('bcf_base', 0), sp.get('bcf_mod', 0)),
            ("CAEF (Bs)", sp.get('caef_base', 0), sp.get('caef_mod', 0)),
            ("VAN Económico (Bs)", sp.get('vane_base', 0), sp.get('vane_mod', 0)),
            ("TIR Económica (%)", sp.get('tire_base', 0), sp.get('tire_mod', 0)),
            ("B/C Económico", sp.get('bce_base', 0), sp.get('bce_mod', 0)),
            ("CAEE (Bs)", sp.get('caee_base', 0), sp.get('caee_mod', 0)),
        ]

        for nombre, base, mod in indicadores_comp:
            variacion = mod - base if isinstance(base, (int, float)) and isinstance(mod, (int, float)) else "N/A"
            fila_vals = [nombre, base, mod, variacion]
            for col_idx, val in enumerate(fila_vals, 1):
                nf = '#,##0.00' if col_idx > 1 and isinstance(val, (int, float)) else None
                engine.escribir_celda(fila, col_idx, val, number_format=nf)
            fila += 1
        fila += 2

    # --- NOTA METODOLÓGICA ---
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
    engine.escribir_celda(fila, 1,
        "Nota: El análisis de sensibilidad evalúa la robustez del proyecto ante cambios en variables clave. "
        "La sección 'Impacto Individual' permite identificar qué variable tiene mayor efecto sobre la viabilidad financiera.",
        size=9, align_h='left', color="666666")

    engine.ajustar_anchos()

def _generar_hoja_riesgo_engine(engine, config, datos_puente, df_fin=None, ind_fin=None, df_eco=None, ind_eco=None):
    """
    Genera hoja Riesgo según el método activo en session_state.
    Si no hay método ejecutado, genera ambas secciones con defaults.
    """
    ws = engine.ws_active
    metodo = st.session_state.get('metodo_riesgo_activo', "Probabilístico")

    # --- ENCABEZADO GENERAL ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    engine.escribir_celda(1, 1, "ANÁLISIS DE RIESGOS", bold=True, size=14,
                           color=engine.COLOR_TITULO, align_h='center')
    engine.escribir_celda(2, 1, f"Proyecto: {config.nombre}  |  Método activo: {metodo}",
                           size=10, align_h='center')
    fila = 4

    # ============================================================
    # SECCIÓN A: RM 115/2015 (siempre generada si hay datos o por defecto)
    # ============================================================
    fila = engine.escribir_titulo_seccion(fila, "A. METODOLOGÍA PROBABILÍSTICA", 10)
    fila += 1

    # Recuperar parámetros
    riesgos = copy.deepcopy(RIESGOS_DEFAULT)
    prob_opt, prob_esp, prob_adv = 0.20, 0.55, 0.25
    res_rm115 = None

    if 'riesgo_rm115_params' in st.session_state:
        rp = st.session_state['riesgo_rm115_params']
        riesgos = rp.get('riesgos', copy.deepcopy(RIESGOS_DEFAULT))
        prob_opt = rp.get('prob_opt', 0.20)
        prob_esp = rp.get('prob_esp', 0.55)
        prob_adv = rp.get('prob_adv', 0.25)

    if 'riesgo_rm115_resultado' in st.session_state and st.session_state['riesgo_rm115_resultado']:
        res_rm115 = st.session_state['riesgo_rm115_resultado']
    elif df_fin is not None and ind_fin is not None:
        res_rm115 = analizar_riesgo_rm115(config, df_fin, ind_fin, riesgos, prob_opt, prob_esp, prob_adv)

    if res_rm115:
        anios = config.generar_anios()
        n = len(anios)

        # Paso 1: Matriz
        fila = engine.escribir_subtitulo(fila, "A.1 Matriz de Riesgos", 7)
        hdr = ["Riesgo", "Probabilidad", "Impacto", "Nivel", "Clasificación", "Afecta", "Fase"]
        for col_idx, h in enumerate(hdr, 1):
            engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
        fila += 1
        for r in res_rm115['matriz_riesgos']:
            color = calcular_color_riesgo(r['clasificacion'])
            vals = [r['nombre'], r['probabilidad'], r['impacto'], r['nivel_riesgo'],
                    r['clasificacion'], r['afecta'], r['fase']]
            for col_idx, val in enumerate(vals, 1):
                nf = '0.00%' if col_idx in [2,3] else ('0.0000' if col_idx == 4 else None)
                engine.escribir_celda(fila, col_idx, val,
                                       fill_color=color if col_idx == 5 else None,
                                       number_format=nf)
            fila += 1
        fila += 1

        # Paso 2: Pérdida esperada
        fila = engine.escribir_subtitulo(fila, "A.2 Pérdida Esperada Anual (Bs)", n+2)
        h_pe = ["Concepto"] + [str(a) for a in anios] + ["Total"]
        for col_idx, h in enumerate(h_pe, 1):
            engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
        fila += 1
        pe = res_rm115['perdida_esperada_anual']
        engine.escribir_fila_datos(fila, ["Pérdida Ingresos"] + pe['ingresos'] + [sum(pe['ingresos'])])
        fila += 1
        engine.escribir_fila_datos(fila, ["Pérdida Costos"] + pe['costos'] + [sum(pe['costos'])])
        fila += 1
        engine.escribir_fila_datos(fila, ["PÉRDIDA TOTAL"] + pe['total'] + [sum(pe['total'])],
                                  bold=True, fill_color=engine.COLOR_TOTAL)
        fila += 2

        # Paso 3: Escenarios
        fila = engine.escribir_subtitulo(fila, "A.3 Escenarios de VAN", 5)
        h_esc = ["Escenario", "VAN (Bs)", "Probabilidad", "Afectación", "Descripción"]
        for col_idx, h in enumerate(h_esc, 1):
            engine.escribir_celda(fila, col_idx, h, bold=True, fill_color=engine.COLOR_HEADER)
        fila += 1
        esc_rows = [
            ("Optimista", res_rm115['van_base'], res_rm115['probabilidades_escenarios']['optimista'],
             "Ninguna", "Sin materialización de riesgos"),
            ("Esperado", res_rm115['escenarios']['Esperado (pérdida ponderada por probabilidad)']['van'],
             res_rm115['probabilidades_escenarios']['esperado'], "Pérdida esperada", "Impacto ponderado"),
            ("Adverso", res_rm115['van_peor'], res_rm115['probabilidades_escenarios']['adverso'],
             "Impacto pleno", "Todos los riesgos al 100%"),
        ]
        for nombre, van, prob, afecta, desc in esc_rows:
            engine.escribir_fila_datos(fila, [nombre, van, prob, afecta, desc])
            fila += 1
        fila += 1
        engine.escribir_celda(fila, 1, "VAN ESPERADO PONDERADO:", bold=True)
        cell = engine.escribir_celda(fila, 2, res_rm115['van_esperado'], bold=True)
        cell.number_format = '#,##0.00'
        fila += 2

        # Paso 4: Sensibilidad
        fila = engine.escribir_subtitulo(fila, "A.4 Sensibilidad de Quiebre", 3)
        s = res_rm115['sensibilidad_umbral']
        engine.escribir_celda(fila, 1, "Umbral Ingresos:", bold=True)
        engine.escribir_celda(fila, 2, s['lectura_ingresos'])
        fila += 1
        engine.escribir_celda(fila, 1, "Umbral Costos:", bold=True)
        engine.escribir_celda(fila, 2, s['lectura_costos'])
        fila += 2

        # Paso 5: Probabilidad conjunta
        fila = engine.escribir_subtitulo(fila, "A.5 Probabilidad Conjunta de Falla", 2)
        engine.escribir_celda(fila, 1, "P(Falla conjunta):", bold=True)
        cell = engine.escribir_celda(fila, 2, res_rm115['prob_falla_conjunta'])
        cell.number_format = '0.00%'
        fila += 2

        # Paso 6: Decisión
        fila = engine.escribir_subtitulo(fila, "A.6 Regla de Decisión", 3)
        dec = res_rm115['nivel_decision']
        color_dec = 'D4EDDA' if 'ROBUSTO' in dec['nivel'] else ('FFF3CD' if 'VIABLE' in dec['nivel'] else 'F8D7DA')
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        engine.escribir_celda(fila, 1, dec['nivel'], bold=True, size=12,
                               fill_color=color_dec, align_h='center')
        fila += 1
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        engine.escribir_celda(fila, 1, dec['recomendacion'], align_h='left')
        fila += 3
    else:
        engine.escribir_celda(fila, 1, "⚠️ No disponible: ejecute el análisis RM 115/2015 en la interfaz.", bold=True)
        fila += 3

    # ============================================================
    # SECCIÓN B: MONTE CARLO (siempre generada si hay datos)
    # ============================================================
    fila = engine.escribir_titulo_seccion(fila, "B. SIMULACIÓN MONTE CARLO", 10)
    fila += 1

    res_mc = None
    if 'riesgo_mc_resultado' in st.session_state and st.session_state['riesgo_mc_resultado']:
        res_mc = st.session_state['riesgo_mc_resultado']
    else:
        # Calcular con defaults si no hay resultado en session_state
        try:
            res_mc = simular_montecarlo(config, datos_puente, n_simulaciones=1000,
                                        cv_precios=0.20, cv_rendimientos=0.15,
                                        cv_costos_prod=0.10, cv_om=0.10, cv_inversion=0.10)
        except Exception:
            pass

    if res_mc:
        # Parámetros
        fila = engine.escribir_subtitulo(fila, "B.1 Parámetros de Simulación", 2)
        params_mc = [
            ("N° Simulaciones", res_mc['n_simulaciones']),
            ("VAN Esperado (Bs)", f"{res_mc['van_mean']:,.0f}"),
            ("Desv. Estándar (Bs)", f"{res_mc['van_std']:,.0f}"),
            ("Prob. Falla", f"{res_mc['prob_falla']*100:.1f}%"),
            ("VaR 10% (Bs)", f"{res_mc['var_10']:,.0f}"),
            ("TIR Esperada (%)", f"{res_mc['tir_mean']:.2f}" if not np.isnan(res_mc['tir_mean']) else "N/A"),
        ]
        for label, val in params_mc:
            engine.escribir_celda(fila, 1, label, bold=True, align_h='left')
            engine.escribir_celda(fila, 2, val, align_h='right')
            fila += 1
        fila += 1

        # Percentiles
        fila = engine.escribir_subtitulo(fila, "B.2 Percentiles del VAN", 2)
        percentiles = [5, 10, 25, 50, 75, 90, 95]
        for i, p in enumerate(percentiles):
            engine.escribir_celda(fila, 1, f"P{p}%", bold=True, align_h='left')
            engine.escribir_celda(fila, 2, res_mc['van_percentiles'][i], align_h='right',
                                   number_format='#,##0.00')
            fila += 1
        fila += 1

        # Histograma (datos para gráfico)
        fila = engine.escribir_subtitulo(fila, "B.3 Distribución del VAN (Datos para Histograma)", 2)
        vans = res_mc['vans']
        bins = np.histogram_bin_edges(vans, bins=30)
        hist, edges = np.histogram(vans, bins=bins)
        bin_centers = (edges[:-1] + edges[1:]) / 2
        df_hist = pd.DataFrame({'VAN': bin_centers, 'Frecuencia': hist})
        df_hist = df_hist[df_hist['Frecuencia'] > 0]

        engine.escribir_celda(fila, 1, "VAN (Bs)", bold=True, fill_color=engine.COLOR_HEADER)
        engine.escribir_celda(fila, 2, "Frecuencia", bold=True, fill_color=engine.COLOR_HEADER)
        fila += 1
        for _, row in df_hist.iterrows():
            engine.escribir_fila_datos(fila, [row['VAN'], row['Frecuencia']])
            fila += 1
        fila += 1

        # Conclusión
        if res_mc['prob_falla'] < 0.05:
            conclusion_mc = "✅ Riesgo muy bajo: probabilidad de pérdida < 5%. Proyecto robusto."
        elif res_mc['prob_falla'] < 0.20:
            conclusion_mc = f"⚠️ Riesgo moderado: probabilidad de pérdida del {res_mc['prob_falla']*100:.1f}%. Planificar contingencias."
        else:
            conclusion_mc = f"🔴 Riesgo elevado: probabilidad de pérdida del {res_mc['prob_falla']*100:.1f}%. Revisar supuestos."
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
        engine.escribir_celda(fila, 1, conclusion_mc, bold=True, size=11, align_h='left')
    else:
        engine.escribir_celda(fila, 1, "⚠️ No disponible: ejecute la simulación Monte Carlo en la interfaz.", bold=True)

    engine.ajustar_anchos()
# ============================================================
# GENERADOR WORKBOOK COMPLETO CON EXCEL REPORT ENGINE
# ============================================================
def generar_workbook_reportes(config, datos_puente, df_fin, ind_fin, df_eco, ind_eco) -> bytes:
    """Genera workbook completo usando ExcelReportEngine."""
    engine = ExcelReportEngine()
    anios = config.generar_anios()
    
    # --- HOJA PREP ---
    ws_prep = engine.crear_hoja("CONFIG")
    engine.activar_hoja("CONFIG")
    _generar_hoja_prep_engine(engine, config, datos_puente)
    engine.ajustar_anchos()
    
    # --- HOJA E_FIN ---
    ws_efin = engine.crear_hoja("E_FIN")
    engine.activar_hoja("E_FIN")
    _generar_hoja_efin_engine(engine, config, df_fin, ind_fin)
    engine.ajustar_anchos()
    
    # --- HOJA E_ECO ---
    ws_eeco = engine.crear_hoja("E_ECO")
    engine.activar_hoja("E_ECO")
    _generar_hoja_eeco_engine(engine, config, df_eco, ind_eco)
    engine.ajustar_anchos()
    
    # --- HOJA Indicadores ---
    ws_ind = engine.crear_hoja("Indicadores")
    engine.activar_hoja("Indicadores")
    _generar_hoja_indicadores_engine(engine, config, ind_fin, ind_eco, datos_puente)
    engine.ajustar_anchos()
    
    # --- HOJA Sensibilidad ---
    ws_sens = engine.crear_hoja("Sensibilidad")
    engine.activar_hoja("Sensibilidad")
    _generar_hoja_sensibilidad_engine(engine, config, datos_puente, df_fin, ind_fin, df_eco, ind_eco)
    engine.ajustar_anchos()

    # --- HOJA Sostenibilidad ---
    ws_sost = engine.crear_hoja("Sostenibilidad")
    engine.activar_hoja("Sostenibilidad")
    _generar_hoja_sostenibilidad_engine(engine, config, df_fin)
    engine.ajustar_anchos()
    
    # HOJA Riesgo
    engine.crear_hoja("Riesgo")
    engine.activar_hoja("Riesgo")
    _generar_hoja_riesgo_engine(engine, config, datos_puente, df_fin, ind_fin, df_eco, ind_eco)
    engine.ajustar_anchos()

    return engine.to_bytes()

# ============================================================
# FLUJOS HORIZONTALES  —  MAPEO RPC PRODUCCIÓN CORREGIDO
# ============================================================
def construir_matriz_financiera(config: ConfiguracionProyecto, datos: Dict,
                                fin_params: Optional[Dict] = None) -> pd.DataFrame:
    anios = config.generar_anios()
    n = len(anios)
    matriz = {}   # <-- ¡IMPORTANTE! Define el diccionario aquí

    # Parámetros financieros con valores por defecto si no se pasan
    if fin_params is None:
        fin_params = {
            'costo_financiero': 0.0,
            'depreciacion_pos': 0.0,
            'depreciacion_neg': 0.0,
            'amortizacion': 0.0,
            'prestamo': 0.0,
            'valor_salvamento': 0.0
        }

    # 1. INGRESOS POR VENTAS (ya escalados en leer_cultivos)
    # ------------------------------------------------------------
    ingresos_cultivos = datos['cultivos']['ingreso_incremental_por_cultivo']
    ingresos = [0.0] * n
    # Filas por cultivo
    for nombre_cult, monto_anual in ingresos_cultivos.items():
        fila_cult = [0.0] * n
        for i in range(n):
            if i < config.duracion_inversion:
                fila_cult[i] = 0.0
            elif i == config.duracion_inversion:
                fila_cult[i] = monto_anual * config.indice_impacto
            else:
                fila_cult[i] = monto_anual
        matriz[f'1.1 {nombre_cult}'] = fila_cult
        for i in range(n):
            ingresos[i] += fila_cult[i]
    # Total ingresos
    matriz['1. INGRESOS POR VENTAS'] = ingresos

    # 2. COSTOS DE INVERSIÓN — separado en Obras Civiles y Servicios
    # ------------------------------------------------------------
    inv_obras_rpc = datos['inversion'].get('inversion_obras_rpc',
                                           datos['inversion']['inversion_rpc'])
    inv_serv_rpc  = datos['inversion'].get('inversion_servicios_rpc',
                                           {k:0 for k in ['BT','BNT','MOC','MOS','MONU','MONR']})
    # 2.1 OBRAS CIVILES (Infraestructura)
    obras_filas = {
        '2.1.1 Bienes Transables (Inv)': 'BT',
        '2.1.2 Bienes no Transables (Inv)': 'BNT',
        '2.1.3 Mano de Obra Calificada (Inv)': 'MOC',
        '2.1.4 Mano de Obra Semicalificada (Inv)': 'MOS',
        '2.1.5 M.O. No Calificada Urbana (Inv)': 'MONU',
        '2.1.6 M.O. No Calificada Rural (Inv)': 'MONR',
    }
    subtotal_obras = [0.0] * n
    for nombre, key in obras_filas.items():
        fila = [0.0] * n
        monto = inv_obras_rpc.get(key, 0)
        if config.duracion_inversion > 0 and monto > 0:
            anual = monto / config.duracion_inversion
            for i in range(min(config.duracion_inversion, n)):
                fila[i] = anual
                subtotal_obras[i] += anual
        matriz[nombre] = fila
    matriz['2.1 SUBTOTAL OBRAS CIVILES'] = subtotal_obras

    # 2.2 SERVICIOS (Supervisión + ATI)
    serv_filas = {
        '2.2.1 Bienes Transables (InvServ)': 'BT',
        '2.2.2 Bienes no Transables (InvServ)': 'BNT',
        '2.2.3 Mano de Obra Calificada (InvServ)': 'MOC',
        '2.2.4 Mano de Obra Semicalificada (InvServ)': 'MOS',
        '2.2.5 M.O. No Calificada Urbana (InvServ)': 'MONU',
        '2.2.6 M.O. No Calificada Rural (InvServ)': 'MONR',
    }
    subtotal_serv = [0.0] * n
    for nombre, key in serv_filas.items():
        fila = [0.0] * n
        monto = inv_serv_rpc.get(key, 0)
        if config.duracion_inversion > 0 and monto > 0:
            anual = monto / config.duracion_inversion
            for i in range(min(config.duracion_inversion, n)):
                fila[i] = anual
                subtotal_serv[i] += anual
        matriz[nombre] = fila
    matriz['2.2 SUBTOTAL SERVICIOS'] = subtotal_serv

    subtotal_inv = [subtotal_obras[i] + subtotal_serv[i] for i in range(n)]
    matriz['2. TOTAL COSTOS DE INVERSIÓN'] = subtotal_inv

    # 3. COSTOS DE OPERACIÓN Y MANTENIMIENTO
    # ------------------------------------------------------------
    om_rpc = datos['inversion']['om_rpc']
    om_total = datos['inversion']['om_total']
    tiene_om = om_total > 0

    subtotal_om = [0.0] * n
    if tiene_om:
        om_filas = {
            '3.1.1 Bienes Transables (O&M)': 'BT',
            '3.1.2 Bienes no Transables (O&M)': 'BNT',
            '3.1.3 Mano de Obra Calificada (O&M)': 'MOC',
            '3.1.4 Mano de Obra Semicalificada (O&M)': 'MOS',
            '3.1.5 M.O. No Calificada Urbana (O&M)': 'MONU',
            '3.1.6 M.O. No Calificada Rural (O&M)': 'MONR',
        }
        for nombre, key in om_filas.items():
            fila = [0.0] * n
            monto = om_rpc.get(key, 0)
            for i in range(config.duracion_inversion, n):
                fila[i] = monto
                subtotal_om[i] += monto
            matriz[nombre] = fila
    matriz['3.1 SUBTOTAL O&M'] = subtotal_om
    matriz['3. TOTAL COSTOS DE OPERACIÓN'] = subtotal_om.copy()

    # 4. COSTOS DE PRODUCCIÓN (incrementales, ya escalados por superficie)
    # ------------------------------------------------------------
    prod_rpc = datos['cultivos']['produccion_rpc']
    prod_filas = {
        '4.1.1 Bienes Transables (Prod)': 'BT',
        '4.1.2 Bienes no Transables (Prod)': 'BNT',
        '4.1.3 Mano de Obra Calificada (Prod)': 'MOC',
        '4.1.4 Mano de Obra Semicalificada (Prod)': 'MOS',
        '4.1.5 M.O. No Calificada Urbana (Prod)': 'MONU',
        '4.1.6 M.O. No Calificada Rural (Prod)': 'MONR',
    }
    subtotal_prod = [0.0] * n
    for nombre, key in prod_filas.items():
        fila = [0.0] * n
        monto_anual = prod_rpc.get(key, 0)
        for i in range(n):
            if i < config.duracion_inversion:
                fila[i] = 0.0
            elif i == config.duracion_inversion:
                fila[i] = monto_anual * config.indice_impacto
            else:
                fila[i] = monto_anual
            subtotal_prod[i] += fila[i]
        matriz[nombre] = fila
    matriz['4.1 SUBTOTAL PRODUCCIÓN'] = subtotal_prod
    matriz['4. TOTAL COSTOS DE PRODUCCIÓN'] = subtotal_prod.copy()

    # 5. COMERCIALIZACIÓN (cero en este modelo)
    # ------------------------------------------------------------
    matriz['5. TOTAL COSTOS DE COMERCIALIZ.'] = [0.0] * n

    # 6. IMPACTO AMBIENTAL
    # ------------------------------------------------------------
    mit_total = datos['inversion']['mitigacion_total']
    subtotal_amb = [0.0] * n
    if mit_total > 0:
        subtotal_amb[0] = mit_total
    matriz['6. TOTAL COSTOS AMBIENTALES'] = subtotal_amb

    # 7. TOTAL COSTOS OPERATIVOS (sin financieros)
    # ------------------------------------------------------------
    total_costos = [0.0] * n
    for i in range(n):
        total_costos[i] = (subtotal_inv[i] + subtotal_om[i] +
                                      subtotal_prod[i] + subtotal_amb[i])
    matriz['7. TOTAL COSTOS'] = total_costos

    # 8. PARTIDAS FINANCIERAS Y FLUJO NETO
    # ------------------------------------------------------------
    costos_fin = [fin_params['costo_financiero']] * n
    matriz['Costos Financieros (Intereses)'] = costos_fin

    dep_pos = [fin_params['depreciacion_pos']] * n
    matriz['Depreciación (+)'] = dep_pos

    dep_neg = [fin_params['depreciacion_neg']] * n
    matriz['Depreciación (-)'] = dep_neg

    amort = [fin_params['amortizacion']] * n
    matriz['Amortización (+)'] = amort

    prestamo = [0.0] * n
    if fin_params['prestamo'] != 0:
        prestamo[0] = fin_params['prestamo']  # se espera negativo
    matriz['Préstamo (-)'] = prestamo

    salvamento = [0.0] * n
    if n > 0:
        salvamento[-1] = fin_params['valor_salvamento']
    matriz['Valor de Salvamento'] = salvamento

    # Flujo de Fondos Neto
    flujo_neto = [0.0] * n
    for i in range(n):
        flujo_neto[i] = (ingresos[i]
                         - total_costos[i]
                         - costos_fin[i]
                         + dep_pos[i]
                         - dep_neg[i]
                         + amort[i]
                         + prestamo[i]
                         + salvamento[i])
    matriz['8. FLUJO DE FONDOS NETO'] = flujo_neto

    # Construcción del DataFrame
    # ------------------------------------------------------------
    df = pd.DataFrame(matriz).T
    df.columns = anios
    df['Total'] = df.sum(axis=1)
    return df

def construir_matriz_economica(config: ConfiguracionProyecto, df_fin: pd.DataFrame) -> pd.DataFrame:
    anios = config.generar_anios()
    rpc = config.rpc
    
    # Copia base desde financiera
    df_eco = df_fin.copy()
    idx_existentes = set(df_eco.index)

    # La RM 115 Art. 20 exige corregir distorsiones en beneficios transables
    if '1. INGRESOS POR VENTAS' in idx_existentes:
        ingresos_mercado = df_fin.loc['1. INGRESOS POR VENTAS', anios].astype(float)
        pct_t = config.pct_produccion_transable
        rpc_div = rpc['divisa']
        
        # Fórmula: Ingreso_PC = Ingreso_PM * [(1-pct_t)*1 + pct_t*rpc_divisa]
        # Es decir: la parte no transable queda igual, la transable se corrige por RPC
        factor_ingreso_eco = (1 - pct_t) + (pct_t * rpc_div)
        df_eco.loc['1. INGRESOS POR VENTAS', anios] = ingresos_mercado * factor_ingreso_eco
        
        # También corregir filas individuales de cultivos (1.1 Nombre)
        cult_rows = [idx for idx in df_fin.index if idx.startswith('1.1 ')]
        for cult_idx in cult_rows:
            df_eco.loc[cult_idx, anios] = df_fin.loc[cult_idx, anios].astype(float) * factor_ingreso_eco
    
    factores_rpc = {
        '2.1.1 Bienes Transables (Inv)': rpc['divisa'],
        '2.1.2 Bienes no Transables (Inv)': 1.0,
        '2.1.3 Mano de Obra Calificada (Inv)': rpc['mo_calificada'],
        '2.1.4 Mano de Obra Semicalificada (Inv)': rpc['mo_semicalificada'],
        '2.1.5 M.O. No Calificada Urbana (Inv)': rpc['mo_no_calif_urbana'],
        '2.1.6 M.O. No Calificada Rural (Inv)': rpc['mo_no_calif_rural'],

        '2.2.1 Bienes Transables (InvServ)': rpc['divisa'],
        '2.2.2 Bienes no Transables (InvServ)': 1.0,
        '2.2.3 Mano de Obra Calificada (InvServ)': rpc['mo_calificada'],
        '2.2.4 Mano de Obra Semicalificada (InvServ)': rpc['mo_semicalificada'],
        '2.2.5 M.O. No Calificada Urbana (InvServ)': rpc['mo_no_calif_urbana'],
        '2.2.6 M.O. No Calificada Rural (InvServ)': rpc['mo_no_calif_rural'],
        
        '3.1.1 Bienes Transables (O&M)': rpc['divisa'],
        '3.1.2 Bienes no Transables (O&M)': 1.0,
        '3.1.3 Mano de Obra Calificada (O&M)': rpc['mo_calificada'],
        '3.1.4 Mano de Obra Semicalificada (O&M)': rpc['mo_semicalificada'],
        '3.1.5 M.O. No Calificada Urbana (O&M)': rpc['mo_no_calif_urbana'],
        '3.1.6 M.O. No Calificada Rural (O&M)': rpc['mo_no_calif_rural'],
        
        '4.1.1 Bienes Transables (Prod)': rpc['divisa'],
        '4.1.2 Bienes no Transables (Prod)': 1.0,
        '4.1.3 Mano de Obra Calificada (Prod)': rpc['mo_calificada'],
        '4.1.4 Mano de Obra Semicalificada (Prod)': rpc['mo_semicalificada'],
        '4.1.5 M.O. No Calificada Urbana (Prod)': rpc['mo_no_calif_urbana'],
        '4.1.6 M.O. No Calificada Rural (Prod)': rpc['mo_no_calif_rural'],
        
        '6. TOTAL COSTOS AMBIENTALES': 1.0,
    }
    
    for idx, factor in factores_rpc.items():
        if idx in idx_existentes:
            df_eco.loc[idx, anios] = df_fin.loc[idx, anios].astype(float) * factor

    obras_rows = [k for k in factores_rpc.keys() if '(Inv)' in k and '(InvServ)' not in k and k in idx_existentes]
    serv_rows  = [k for k in factores_rpc.keys() if '(InvServ)' in k and k in idx_existentes]
    om_rows    = [k for k in factores_rpc.keys() if '(O&M)' in k and k in idx_existentes]
    prod_rows  = [k for k in factores_rpc.keys() if '(Prod)' in k and k in idx_existentes]

    if obras_rows:
        df_eco.loc['2.1 SUBTOTAL OBRAS CIVILES', anios] = df_eco.loc[obras_rows, anios].sum()
    else:
        df_eco.loc['2.1 SUBTOTAL OBRAS CIVILES', anios] = 0.0

    if serv_rows:
        df_eco.loc['2.2 SUBTOTAL SERVICIOS', anios] = df_eco.loc[serv_rows, anios].sum()
    else:
        df_eco.loc['2.2 SUBTOTAL SERVICIOS', anios] = 0.0

    if om_rows:
        df_eco.loc['3.1 SUBTOTAL O&M', anios] = df_eco.loc[om_rows, anios].sum()
    else:
        df_eco.loc['3.1 SUBTOTAL O&M', anios] = 0.0

    if prod_rows:
        df_eco.loc['4.1 SUBTOTAL PRODUCCIÓN', anios] = df_eco.loc[prod_rows, anios].sum()
    else:
        df_eco.loc['4.1 SUBTOTAL PRODUCCIÓN', anios] = 0.0

    df_eco.loc['2. TOTAL COSTOS DE INVERSIÓN', anios] = (
        df_eco.loc['2.1 SUBTOTAL OBRAS CIVILES', anios] +
        df_eco.loc['2.2 SUBTOTAL SERVICIOS', anios]
    )
    df_eco.loc['3. TOTAL COSTOS DE OPERACIÓN', anios] = df_eco.loc['3.1 SUBTOTAL O&M', anios]
    df_eco.loc['4. TOTAL COSTOS DE PRODUCCIÓN', anios] = df_eco.loc['4.1 SUBTOTAL PRODUCCIÓN', anios]
  
    if '6. TOTAL COSTOS AMBIENTALES' in idx_existentes:
        amb_val = df_eco.loc['6. TOTAL COSTOS AMBIENTALES', anios]
    else:
        amb_val = 0.0
        df_eco.loc['6. TOTAL COSTOS AMBIENTALES', anios] = 0.0
    
    df_eco.loc['7. TOTAL COSTOS', anios] = (
        df_eco.loc['2. TOTAL COSTOS DE INVERSIÓN', anios] +
        df_eco.loc['3. TOTAL COSTOS DE OPERACIÓN', anios] +
        df_eco.loc['4. TOTAL COSTOS DE PRODUCCIÓN', anios] +
        amb_val
    )

    df_eco.loc['8. FLUJO DE FONDOS NETO', anios] = (
        df_eco.loc['1. INGRESOS POR VENTAS', anios] - 
        df_eco.loc['7. TOTAL COSTOS', anios]
    )
    
    df_eco['Total'] = df_eco[anios].sum(axis=1)
    return df_eco

# ==============sensibilidad================
def recalcular_con_variacion(
    config: ConfiguracionProyecto,
    datos_puente: Dict,
    var_inversion: float = 0.0,
    var_om: float = 0.0,
    var_produccion: float = 0.0,
    var_ingresos: float = 0.0,
    var_tasa_priv: float = 0.0,
    var_tasa_soc: float = 0.0
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict, Dict]:
    datos_mod = copy.deepcopy(datos_puente)
    config_mod = copy.deepcopy(config)

    # Ajustar inversión
    if var_inversion != 0:
        factor = 1 + var_inversion
        datos_mod['inversion']['inversion_total'] *= factor
        for k in datos_mod['inversion']['inversion_rpc']:
            datos_mod['inversion']['inversion_rpc'][k] *= factor
        if datos_mod['inversion']['mitigacion_total'] > 0:
            datos_mod['inversion']['mitigacion_total'] *= factor
            for k in datos_mod['inversion']['mitigacion_rpc']:
                datos_mod['inversion']['mitigacion_rpc'][k] *= factor

    # Ajustar O&M
    if var_om != 0:
        factor = 1 + var_om
        if datos_mod['inversion']['om_total'] > 0:
            datos_mod['inversion']['om_total'] *= factor
            for k in datos_mod['inversion']['om_rpc']:
                datos_mod['inversion']['om_rpc'][k] *= factor

    # Ajustar costos de producción
    if var_produccion != 0:
        factor = 1 + var_produccion
        for k in datos_mod['cultivos']['produccion_rpc']:
            if k != 'TOTAL':
                datos_mod['cultivos']['produccion_rpc'][k] *= factor
        datos_mod['cultivos']['produccion_rpc']['TOTAL'] = sum(
            v for k, v in datos_mod['cultivos']['produccion_rpc'].items() if k != 'TOTAL'
        )

    if var_ingresos != 0:
        factor = 1 + var_ingresos
        # Ajustar total agregado
        datos_mod['cultivos']['ingreso_incremental_anual'] *= factor
        # NUEVO: Ajustar desglose por cultivo para mantener consistencia
        if 'ingreso_incremental_por_cultivo' in datos_mod['cultivos']:
            for cult in datos_mod['cultivos']['ingreso_incremental_por_cultivo']:
                datos_mod['cultivos']['ingreso_incremental_por_cultivo'][cult] *= factor

    # Ajustar tasas de descuento (suma puntos porcentuales, no multiplicativo)
    config_mod.tasa_privada_descuento = max(0.001, config.tasa_privada_descuento + var_tasa_priv)
    config_mod.tasa_social_descuento = max(0.001, config.tasa_social_descuento + var_tasa_soc)

    # CORRECCIÓN: extraer parámetros financieros de config_mod 
    fin_params_mod = {
        'costo_financiero': config_mod.costo_financiero,
        'depreciacion_pos': config_mod.depreciacion_pos,
        'depreciacion_neg': config_mod.depreciacion_neg,
        'amortizacion': config_mod.amortizacion,
        'prestamo': config_mod.prestamo,
        'valor_salvamento': config_mod.valor_salvamento
    }
    # Reconstruir matrices usando los parámetros financieros actualizados
    df_fin_mod = construir_matriz_financiera(config_mod, datos_mod, fin_params_mod)

    # IMPORTANTE: Para económica, usar la matriz financiera ya modificada
    df_eco_mod = construir_matriz_economica(config_mod, df_fin_mod)

    anios = config_mod.generar_anios()
    ind_fin_mod = calcular_indicadores(df_fin_mod, config_mod.tasa_privada_descuento, anios)
    ind_eco_mod = calcular_indicadores(df_eco_mod, config_mod.tasa_social_descuento, anios)

    return df_fin_mod, df_eco_mod, ind_fin_mod, ind_eco_mod

# ============================================================
# INDICADORES
# ============================================================
def calcular_indicadores(df: pd.DataFrame, tasa: float, anios: List[int]) -> Dict:
    flujos = df.loc['8. FLUJO DE FONDOS NETO', anios].values.astype(float).tolist()
    ingresos = df.loc['1. INGRESOS POR VENTAS', anios].values.astype(float).tolist()
    costos = df.loc['7. TOTAL COSTOS', anios].values.astype(float).tolist()
    
    van = calcular_van(tasa, flujos)
    tir = calcular_tir(flujos)
    
    va_ing = calcular_van(tasa, ingresos)
    va_cost = calcular_van(tasa, costos)
    rbc = va_ing / va_cost if va_cost != 0 else 0
    
    n = len(anios)
    if tasa > 0 and n > 0:
        factor = (tasa * (1 + tasa)**n) / ((1 + tasa)**n - 1)
    else:
        factor = 0
    cae = van * factor
    
    return {
        'VAN': van,
        'TIR': tir,
        'Relacion_BC': rbc,
        'CAE': cae,
        'VA_Ingresos': va_ing,   # VAIS en económico
        'VA_Costos': va_cost,    # VACS en económico
        'flujos': flujos
    }

# ============================================================
# CREAR / LIMPIAR PROYECTO
# ============================================================
def crear_proyecto_nuevo(ruta: str = "proyecto_activo.xlsx"):
    dm = DataManager(ruta)
    
    ws_cult = wb.active
    ws_cult.title = "Cultivos"
    headers = ["Nombre","Codigo","Sup_SP_Ha","Sup_CP_Ha","Rend_SP","Rend_CP",
               "Perd_SP_%","Perd_CP_%","Precio_Bs_Ton","CostoTotal_SP","CostoTotal_CP",
               "BT_SP", "BNT_SP", "MONR_SP", "MONU_SP", "MOS_SP", "MOC_SP", "BT_CP", "BNT_CP", "MONR_CP", "MONU_CP", "MOS_CP", "MOC_CP",
               "Ingreso_SP","Ingreso_CP","Fecha_Agregado"]
    ws_cult.append(headers)
    
    ws_inv = wb.create_sheet("Inversion_Resumen")
    ws_inv.append(["RESUMEN DE INVERSIÓN POR RPC - RM 115/2015"])
    ws_inv.append([])
    ws_inv.append(["Categoría","BT","BNT","MOC","MOS","MONU","MONR","TOTAL"])
    categorias = [
        ["OBRAS CIVILES",0,0,0,0,0,0,0],
        ["Asistencia Técnica Integral",0,0,0,0,0,0,0],
        ["Supervisión de Obras",0,0,0,0,0,0,0],
        ["Operación y Mantenimiento",0,0,0,0,0,0,0],
        ["Mitigación Ambiental",0,0,0,0,0,0,0],
        ["TOTAL INVERSIÓN",0,0,0,0,0,0,0]
    ]
    for cat in categorias:
        ws_inv.append(cat)
    
    return dm.crear_proyecto_nuevo()
    
# ============================================================
# SOSTENIBILIDAD OPERATIVA — Fase 3
# ============================================================
def calcular_sostenibilidad_operativa(config: ConfiguracionProyecto, 
                                       df_fin: pd.DataFrame,
                                       anios: List[int]) -> Dict:
    """
    Evalúa si los ingresos cubren los costos operativos a precios de mercado.
    Art. 16.9 RM 115/2015 — Verificación de sostenibilidad operativa.
    """
    n = len(anios)
    
    # Ingresos anuales (precios de mercado)
    ingresos = df_fin.loc['1. INGRESOS POR VENTAS', anios].values.astype(float)
    
    # Costos operativos (O&M + Producción + Comercialización)
    # No incluye inversión ni ambiental
    om = df_fin.loc['3. TOTAL COSTOS DE OPERACIÓN', anios].values.astype(float) \
         if '3. TOTAL COSTOS DE OPERACIÓN' in df_fin.index else np.zeros(n)
    prod = df_fin.loc['4. TOTAL COSTOS DE PRODUCCIÓN', anios].values.astype(float) \
           if '4. TOTAL COSTOS DE PRODUCCIÓN' in df_fin.index else np.zeros(n)
    com = df_fin.loc['5. TOTAL COSTOS DE COMERCIALIZ.', anios].values.astype(float) \
          if '5. TOTAL COSTOS DE COMERCIALIZ.' in df_fin.index else np.zeros(n)
    
    costos_op = om + prod + com
    
    # Superávit / Déficit operativo
    superavit = ingresos - costos_op
    
    # Ratio de cobertura (ingresos / costos operativos)
    ratio_cobertura = np.divide(ingresos, costos_op, 
                                out=np.full_like(ingresos, np.nan), 
                                where=costos_op!=0)
    
    # Punto de equilibrio: primer año con superávit >= 0 después de la inversión
    anio_equilibrio = None
    anios_operacion = list(range(config.duracion_inversion, n))
    for i in anios_operacion:
        if superavit[i] >= 0:
            anio_equilibrio = anios[i]
            break
    
    # Años con déficit operativo consecutivos desde inicio de operación
    anos_deficit = 0
    for i in anios_operacion:
        if superavit[i] < 0:
            anos_deficit += 1
        else:
            break
    
    # Años con superávit consecutivos desde inicio de operación
    anos_superavit = 0
    for i in anios_operacion:
        if superavit[i] >= 0:
            anos_superavit += 1
        else:
            break
    
    # Promedios del período de operación
    ingresos_op = ingresos[config.duracion_inversion:]
    costos_op_mean = np.mean(costos_op[config.duracion_inversion:]) if len(costos_op) > config.duracion_inversion else 0
    ratio_promedio = np.nanmean(ratio_cobertura[config.duracion_inversion:]) if len(ratio_cobertura) > config.duracion_inversion else 0
    
    return {
        'anios': anios,
        'ingresos': ingresos.tolist(),
        'costos_operativos': costos_op.tolist(),
        'superavit_operativo': superavit.tolist(),
        'ratio_cobertura': ratio_cobertura.tolist(),
        'anio_equilibrio': anio_equilibrio,
        'anos_deficit_consecutivos': anos_deficit,
        'anos_superavit_consecutivos': anos_superavit,
        'ratio_cobertura_promedio': ratio_promedio,
        'costo_operativo_promedio': costos_op_mean,
        'sostenible': anio_equilibrio is not None,
    }
# ============================================================
# SIMULACIÓN MONTE CARLO PARA ANÁLISIS DE RIESGO
# ============================================================
def simular_montecarlo(
    config: ConfiguracionProyecto,
    datos_puente: Dict,
    n_simulaciones: int = 1000,
    cv_precios: float = 0.20,
    cv_rendimientos: float = 0.15,
    cv_costos_prod: float = 0.10,
    cv_om: float = 0.10,
    cv_inversion: float = 0.10,
    semilla: int = 42
) -> Dict:
    """
    Simula escenarios aleatorios de VANF (financiero) mediante Monte Carlo.
    Variables consideradas:
      - Precios de productos (distribución log-normal)
      - Rendimientos agrícolas (distribución normal truncada)
      - Costos de producción (normal)
      - Costos O&M (normal)
      - Inversión total (normal)
    Retorna diccionario con estadísticas y resultados.
    """
    np.random.seed(semilla)
    n = len(config.generar_anios())
    
    # 1. Extraer valores base de los datos
    # Ingreso anual incremental (promedio de cultivos)
    ingreso_base = datos_puente['cultivos']['ingreso_incremental_anual']
    
    # Desglose de producción RPC (para ajustar costos)
    prod_rpc_base = datos_puente['cultivos']['produccion_rpc'].copy()
    # O&M total anual
    om_base = datos_puente['inversion']['om_total']
    # Inversión total
    inv_base = datos_puente['inversion']['inversion_total']
    
    # 2. Extraer parámetros de config
    tasa_priv = config.tasa_privada_descuento
    
    # 3. Prepara arrays para almacenar resultados
    vans = np.zeros(n_simulaciones)
    tirs = np.zeros(n_simulaciones)
    flujos_sim = []
    
    # 4. Bucle de simulación
    for i in range(n_simulaciones):
        # --- Generar variables aleatorias ---
        # Precios (log-normal para evitar negativos)
        precio_factor = np.random.lognormal(mean=0, sigma=cv_precios)
        # Rendimientos (normal, truncados para no ser negativos)
        rend_factor = 1 + np.random.normal(0, cv_rendimientos)
        rend_factor = max(rend_factor, 0.1)  # mínimo 10% del base
        # Costos de producción (normal)
        prod_factor = 1 + np.random.normal(0, cv_costos_prod)
        prod_factor = max(prod_factor, 0.1)
        # O&M (normal)
        om_factor = 1 + np.random.normal(0, cv_om)
        om_factor = max(om_factor, 0.1)
        # Inversión (normal)
        inv_factor = 1 + np.random.normal(0, cv_inversion)
        inv_factor = max(inv_factor, 0.1)
        
        # --- Aplicar factores a los datos base ---
        # Clonar datos_puente para esta simulación
        datos_sim = copy.deepcopy(datos_puente)
        
        # Ajustar ingresos: factor combinado precio * rendimiento
        datos_sim['cultivos']['ingreso_incremental_anual'] = ingreso_base * precio_factor * rend_factor
        # Ajustar desglose por cultivo (si existe)
        if 'ingreso_incremental_por_cultivo' in datos_sim['cultivos']:
            for cult in datos_sim['cultivos']['ingreso_incremental_por_cultivo']:
                datos_sim['cultivos']['ingreso_incremental_por_cultivo'][cult] *= precio_factor * rend_factor
        
        # Ajustar costos de producción (RPC) por factor
        for k in datos_sim['cultivos']['produccion_rpc']:
            if k != 'TOTAL':
                datos_sim['cultivos']['produccion_rpc'][k] *= prod_factor
        datos_sim['cultivos']['produccion_rpc']['TOTAL'] = sum(
            v for k, v in datos_sim['cultivos']['produccion_rpc'].items() if k != 'TOTAL'
        )
        
        # Ajustar O&M
        if datos_sim['inversion']['om_total'] > 0:
            datos_sim['inversion']['om_total'] *= om_factor
            for k in datos_sim['inversion']['om_rpc']:
                datos_sim['inversion']['om_rpc'][k] *= om_factor
        
        # Ajustar Inversión
        datos_sim['inversion']['inversion_total'] *= inv_factor
        for k in datos_sim['inversion']['inversion_rpc']:
            datos_sim['inversion']['inversion_rpc'][k] *= inv_factor
        if datos_sim['inversion']['mitigacion_total'] > 0:
            datos_sim['inversion']['mitigacion_total'] *= inv_factor
            for k in datos_sim['inversion']['mitigacion_rpc']:
                datos_sim['inversion']['mitigacion_rpc'][k] *= inv_factor
        
        # --- Construir matriz financiera para este escenario ---
        fin_params = {
            'costo_financiero': config.costo_financiero,
            'depreciacion_pos': config.depreciacion_pos,
            'depreciacion_neg': config.depreciacion_neg,
            'amortizacion': config.amortizacion,
            'prestamo': config.prestamo,
            'valor_salvamento': config.valor_salvamento
        }
        df_fin_sim = construir_matriz_financiera(config, datos_sim, fin_params)
        anios = config.generar_anios()
        ind_fin = calcular_indicadores(df_fin_sim, tasa_priv, anios)
        
        vans[i] = ind_fin['VAN']
        tirs[i] = ind_fin['TIR'] if ind_fin['TIR'] is not None else -999
        flujos_sim.append(ind_fin['flujos'])  # guardar flujo para análisis posterior
    
    # --- Estadísticas ---
    van_mean = np.mean(vans)
    van_std = np.std(vans)
    van_percentiles = np.percentile(vans, [5, 10, 25, 50, 75, 90, 95])
    prob_falla = np.mean(vans < 0)
    
    # TIR media (solo simulaciones con TIR válida)
    tirs_valid = tirs[tirs > -999]
    tir_mean = np.mean(tirs_valid) if len(tirs_valid) > 0 else np.nan
    tir_std = np.std(tirs_valid) if len(tirs_valid) > 0 else np.nan
    
    # VaR al 10% (percentil 10 del VAN)
    var_10 = van_percentiles[1]  # percentil 10
    
    return {
        'vans': vans,
        'tirs': tirs,
        'flujos_sim': flujos_sim,
        'van_mean': van_mean,
        'van_std': van_std,
        'van_percentiles': van_percentiles,
        'prob_falla': prob_falla,
        'tir_mean': tir_mean,
        'tir_std': tir_std,
        'var_10': var_10,
        'n_simulaciones': n_simulaciones
    }

# ============================================================
# ANÁLISIS DE RIESGOS — Metodología probabilística RM 115/2015
# ============================================================

RIESGOS_DEFAULT = [
    {
        'nombre': 'Riesgo Climático (sequía / helada / granizo)',
        'probabilidad': 0.15,
        'impacto': 0.30,
        'afecta': 'ingresos',
        'fase': 'operacion',
    },
    {
        'nombre': 'Riesgo de Mercado (caída de precios agrícolas)',
        'probabilidad': 0.20,
        'impacto': 0.15,
        'afecta': 'ingresos',
        'fase': 'operacion',
    },
    {
        'nombre': 'Riesgo de Ejecución (retraso de obra)',
        'probabilidad': 0.10,
        'impacto': 0.10,
        'afecta': 'costos',
        'fase': 'inversion',
    },
    {
        'nombre': 'Riesgo Operativo (incremento de costos O&M)',
        'probabilidad': 0.15,
        'impacto': 0.20,
        'afecta': 'costos',
        'fase': 'operacion',
    },
    {
        'nombre': 'Riesgo Social (conflictos de uso de agua)',
        'probabilidad': 0.10,
        'impacto': 0.25,
        'afecta': 'ingresos',
        'fase': 'operacion',
    },
    {
        'nombre': 'Riesgo de Infraestructura (daño por eventos extremos)',
        'probabilidad': 0.08,
        'impacto': 0.35,
        'afecta': 'costos',
        'fase': 'todos',
    },
]


def clasificar_nivel_riesgo(nr: float) -> str:
    """Semaforización del Nivel de Riesgo (NR = P x I, escala 0-1)."""
    nr_escala25 = nr * 25
    if nr_escala25 <= 5:
        return 'Bajo'
    elif nr_escala25 <= 12:
        return 'Medio'
    elif nr_escala25 <= 19:
        return 'Alto'
    else:
        return 'Crítico'


def calcular_color_riesgo(clasificacion: str) -> str:
    """Retorna código hex para color de fondo según nivel."""
    return {
        'Bajo': 'D4EDDA',
        'Medio': 'FFF3CD',
        'Alto': 'F8D7DA',
        'Crítico': '721C24',
    }.get(clasificacion, 'FFFFFF')


def calcular_matriz_riesgos(riesgos: List[Dict]) -> List[Dict]:
    """PASO 1: Matriz de riesgos (Probabilidad × Impacto)."""
    matriz = []
    for r in riesgos:
        nr = r['probabilidad'] * r['impacto']
        matriz.append({
            'nombre': r['nombre'],
            'probabilidad': r['probabilidad'],
            'impacto': r['impacto'],
            'afecta': r['afecta'],
            'fase': r['fase'],
            'nivel_riesgo': nr,
            'clasificacion': clasificar_nivel_riesgo(nr),
        })
    return matriz


def calcular_perdida_esperada_anual(
    riesgos: List[Dict],
    ingresos: np.ndarray,
    costos: np.ndarray,
    n: int,
    duracion_inversion: int,
) -> Dict[str, np.ndarray]:
    """PASO 2: Pérdida esperada por año."""
    perdida_ingresos = np.zeros(n)
    perdida_costos = np.zeros(n)

    for anio_pos in range(n):
        es_inversion = anio_pos < duracion_inversion
        es_operacion = anio_pos >= duracion_inversion

        for r in riesgos:
            aplica_fase = (
                r['fase'] == 'todos'
                or (r['fase'] == 'inversion' and es_inversion)
                or (r['fase'] == 'operacion' and es_operacion)
            )
            if not aplica_fase:
                continue

            if r['afecta'] == 'ingresos':
                perdida_ingresos[anio_pos] += r['probabilidad'] * r['impacto'] * ingresos[anio_pos]
            elif r['afecta'] == 'costos':
                perdida_costos[anio_pos] += r['probabilidad'] * r['impacto'] * costos[anio_pos]

    return {
        'perdida_ingresos': perdida_ingresos,
        'perdida_costos': perdida_costos,
        'perdida_total': perdida_ingresos + perdida_costos,
    }


def calcular_perdida_impacto_pleno(
    riesgos: List[Dict],
    ingresos: np.ndarray,
    costos: np.ndarray,
    n: int,
    duracion_inversion: int,
) -> np.ndarray:
    """Pérdida con impacto pleno (probabilidad=1) para escenario adverso."""
    perdida = np.zeros(n)
    for anio_pos in range(n):
        es_inversion = anio_pos < duracion_inversion
        es_operacion = anio_pos >= duracion_inversion
        for r in riesgos:
            aplica_fase = (
                r['fase'] == 'todos'
                or (r['fase'] == 'inversion' and es_inversion)
                or (r['fase'] == 'operacion' and es_operacion)
            )
            if not aplica_fase:
                continue
            base = ingresos[anio_pos] if r['afecta'] == 'ingresos' else costos[anio_pos]
            perdida[anio_pos] += r['impacto'] * base
    return perdida


def calcular_van_esperado(
    tasa: float,
    flujo_base: np.ndarray,
    flujo_ajustado: np.ndarray,
    perdida_impacto_pleno: np.ndarray,
    prob_optimista: float = 0.20,
    prob_esperado: float = 0.55,
    prob_adverso: float = 0.25,
) -> Dict:
    """PASO 3: Valor esperado del VAN sobre 3 escenarios."""
    # Normalizar probabilidades
    total_prob = prob_optimista + prob_esperado + prob_adverso
    if abs(total_prob - 1.0) > 1e-6:
        prob_optimista /= total_prob
        prob_esperado /= total_prob
        prob_adverso /= total_prob

    van_optimista = calcular_van(tasa, flujo_base.tolist())
    van_esperado_flujo = calcular_van(tasa, flujo_ajustado.tolist())
    flujo_adverso = flujo_base - perdida_impacto_pleno
    van_adverso = calcular_van(tasa, flujo_adverso.tolist())

    e_van = (
        prob_optimista * van_optimista
        + prob_esperado * van_esperado_flujo
        + prob_adverso * van_adverso
    )

    return {
        'van_optimista': van_optimista,
        'van_esperado_escenario': van_esperado_flujo,
        'van_adverso': van_adverso,
        'flujo_adverso': flujo_adverso,
        'probabilidades': {
            'optimista': prob_optimista,
            'esperado': prob_esperado,
            'adverso': prob_adverso,
        },
        'van_esperado': e_van,
    }


def calcular_sensibilidad_umbral(van_base: float, va_ingresos: float, va_costos: float) -> Dict:
    """PASO 4: Sensibilidad de quiebre."""
    var_critica_ingresos = van_base / va_ingresos if va_ingresos != 0 else None
    var_critica_costos = van_base / va_costos if va_costos != 0 else None

    return {
        'variacion_critica_ingresos_pct': var_critica_ingresos,
        'variacion_critica_costos_pct': var_critica_costos,
        'lectura_ingresos': (
            f"Los ingresos pueden caer hasta {abs(var_critica_ingresos)*100:.1f}% "
            f"antes de que el VAN se vuelva cero"
        ) if var_critica_ingresos is not None else "No calculable",
        'lectura_costos': (
            f"Los costos pueden subir hasta {abs(var_critica_costos)*100:.1f}% "
            f"antes de que el VAN se vuelva negativo"
        ) if var_critica_costos is not None else "No calculable",
    }


def calcular_probabilidad_falla_conjunta(riesgos: List[Dict]) -> float:
    """PASO 5: Probabilidad de que ocurra al menos un riesgo (independientes)."""
    prob_ninguno = 1.0
    for r in riesgos:
        prob_ninguno *= (1 - r['probabilidad'])
    return 1 - prob_ninguno


def evaluar_regla_decision(e_van: float, prob_falla: float) -> Dict:
    """PASO 6: Regla de decisión con semáforo."""
    if e_van > 0 and prob_falla < 0.25:
        nivel = "PROYECTO ROBUSTO ANTE RIESGO"
        color = "🟢"
        recomendacion = "El proyecto mantiene viabilidad económica incluso bajo condiciones adversas. Continuar con la ejecución."
    elif e_van > 0 and prob_falla <= 0.50:
        nivel = "VIABLE — REQUIERE MEDIDAS DE MITIGACIÓN"
        color = "🟡"
        recomendacion = "El proyecto es viable en valor esperado, pero la probabilidad de falla conjunta es significativa. Incorporar medidas de mitigación en el diseño."
    else:
        nivel = "REVISAR DISEÑO — NO VIABLE EN ESTAS CONDICIONES"
        color = "🔴"
        recomendacion = "El VAN esperado es negativo o la probabilidad de falla es muy alta. Se requiere rediseñar supuestos, reducir costos o incrementar beneficios antes de aprobar."
    return {
        'nivel': nivel,
        'color': color,
        'recomendacion': recomendacion,
    }


def analizar_riesgo_rm115(
    config,
    df_fin: pd.DataFrame,
    ind_fin: Dict,
    riesgos: List[Dict] = None,
    prob_optimista: float = 0.20,
    prob_esperado: float = 0.55,
    prob_adverso: float = 0.25,
) -> Dict:
    """
    Función principal de análisis de riesgos RM 115/2015.
    Retorna diccionario completo con todos los pasos.
    """
    if riesgos is None:
        riesgos = RIESGOS_DEFAULT

    anios = config.generar_anios()
    n = len(anios)
    tasa = config.tasa_privada_descuento

    ingresos = df_fin.loc['1. INGRESOS POR VENTAS', anios].values.astype(float)
    costos = df_fin.loc['7. TOTAL COSTOS', anios].values.astype(float)
    flujo_base = df_fin.loc['8. FLUJO DE FONDOS NETO', anios].values.astype(float)

    # PASO 1
    matriz_riesgos = calcular_matriz_riesgos(riesgos)

    # PASO 2
    perdidas = calcular_perdida_esperada_anual(
        riesgos, ingresos, costos, n, config.duracion_inversion
    )
    flujo_ajustado = flujo_base - perdidas['perdida_total']
    van_ajustado_riesgo = calcular_van(tasa, flujo_ajustado.tolist())

    # PASO 3
    perdida_plena = calcular_perdida_impacto_pleno(
        riesgos, ingresos, costos, n, config.duracion_inversion
    )
    resultado_van = calcular_van_esperado(
        tasa, flujo_base, flujo_ajustado, perdida_plena,
        prob_optimista, prob_esperado, prob_adverso
    )

    # PASO 4
    va_ingresos = calcular_van(tasa, ingresos.tolist())
    va_costos = calcular_van(tasa, costos.tolist())
    sensibilidad = calcular_sensibilidad_umbral(ind_fin['VAN'], va_ingresos, va_costos)

    # PASO 5
    prob_falla_conjunta = calcular_probabilidad_falla_conjunta(riesgos)

    # PASO 6
    decision = evaluar_regla_decision(resultado_van['van_esperado'], prob_falla_conjunta)

    # Escenarios para compatibilidad con interfaz
    escenarios = {
        'Optimista (sin riesgos)': {
            'van': resultado_van['van_optimista'],
            'flujo': flujo_base.tolist(),
        },
        'Esperado (pérdida ponderada por probabilidad)': {
            'van': resultado_van['van_esperado_escenario'],
            'flujo': flujo_ajustado.tolist(),
        },
        'Adverso (impacto pleno de riesgos)': {
            'van': resultado_van['van_adverso'],
            'flujo': resultado_van['flujo_adverso'].tolist(),
        },
    }

    return {
        'escenarios': escenarios,
        'van_base': resultado_van['van_optimista'],
        'van_peor': resultado_van['van_adverso'],
        'van_mejor': resultado_van['van_optimista'],
        'van_esperado': resultado_van['van_esperado'],
        'prob_falla_estimada': prob_falla_conjunta,
        'var_10': resultado_van['van_adverso'],
        'num_escenarios_falla': sum(1 for v in escenarios.values() if v['van'] < 0),
        'anios': anios,
        'matriz_riesgos': matriz_riesgos,
        'perdida_esperada_anual': {
            'ingresos': perdidas['perdida_ingresos'].tolist(),
            'costos': perdidas['perdida_costos'].tolist(),
            'total': perdidas['perdida_total'].tolist(),
        },
        'flujo_ajustado_riesgo': flujo_ajustado.tolist(),
        'van_ajustado_riesgo': van_ajustado_riesgo,
        'probabilidades_escenarios': resultado_van['probabilidades'],
        'sensibilidad_umbral': sensibilidad,
        'prob_falla_conjunta': prob_falla_conjunta,
        'nivel_decision': decision,
        'riesgos_config': riesgos,
        'prob_opt_config': prob_optimista,
        'prob_esp_config': prob_esperado,
        'prob_adv_config': prob_adverso,
    }

# ============================================================
# INTERFAZ PRINCIPAL
# ============================================================
def main():
    st.title("💧 Evaluación Económica y Financiera - RM 115/2015")
    st.markdown("**Metodología:** Costo-Beneficio y Costo-Eficiencia")
    
    # ============================================================
    # DETECTAR PROYECTO ACTIVO (NUEVO - Fase 4)
    # ============================================================
    pm = ProjectManager()
    rutas = pm.get_rutas_activas()
    
    if rutas is None:
        st.error("❌ No hay un proyecto activo.")
        st.info("Vaya a **app.py** (orquestador), cree un nuevo proyecto o seleccione uno existente desde el panel lateral.")
        st.stop()
    
    # Rutas dinámicas del proyecto activo
    ruta_puente = rutas["excel"]
    ruta_config = rutas["config"]
    nombre_proyecto = rutas["nombre"]
    
    st.info(f"📁 Proyecto activo: **{nombre_proyecto}** ")
    
    # ============================================================
    # CARGAR CONFIGURACIÓN DESDE PROYECTO ACTIVO
    # ============================================================
    # ============================================================
    # SIDEBAR DE CONFIGURACIÓN
    # ============================================================
    # ============================================================
    # CONSTRUIR CONFIGURACIÓN Y GUARDAR EN PROYECTO ACTIVO
    # ============================================================
    # ============================================================
    # CARGAR CONFIGURACIÓN DESDE PROYECTO ACTIVO (única fuente de verdad)
    # ============================================================
    config_existente = None
    try:
        if os.path.exists(ruta_config):
            config_existente = ConfiguracionProyecto.cargar(ruta_config)
    except Exception:
        pass

    if config_existente is None:
        st.error("❌ No se encontró configuración del proyecto activo.")
        st.info("Vaya a **Configuración** primero para establecer los parámetros del proyecto.")
        st.stop()

    config = config_existente  # Usar directamente, no reconstruir

    fin_params = {
        'costo_financiero': config.costo_financiero,
        'depreciacion_pos': config.depreciacion_pos,
        'depreciacion_neg': config.depreciacion_neg,
        'amortizacion': config.amortizacion,
        'prestamo': config.prestamo,
        'valor_salvamento': config.valor_salvamento
    }

    # ============================================================
    # LECTURA DEL PUENTE CON DATA_MANAGER (ruta dinámica)
    # ============================================================
    anios = config.generar_anios()
    # CORRECCIÓN: Usar ProjectManager para obtener ruta dinámica
    pm_eval = ProjectManager()
    rutas_eval = pm_eval.get_rutas_activas()
    if rutas_eval:
        ruta_puente = rutas_eval["excel"]
    else:
        ruta_puente = "proyecto_activo.xlsx"
    dm = DataManager(ruta_puente)  # ← Usa ruta del proyecto activo

    if not os.path.exists(ruta_puente):
        st.warning(f"⚠️ No se encontró el puente: `{ruta_puente}`")
        st.info("Primero debe completar los módulos de Costos e Inversión.")
        st.stop()
         
    datos_cult = dm.leer_cultivos()
    datos_inv = dm.leer_inversion_resumen()

    # Permitir continuar si al menos hay configuración válida
    tiene_datos = False
    if datos_cult.get('ok') and datos_cult.get('num_cultivos', 0) > 0:
        tiene_datos = True
    if datos_inv.get('ok') and datos_inv.get('inversion_total', 0) > 0:
        tiene_datos = True    
    if not tiene_datos:
        st.error("❌ No se encontraron datos de cultivos o inversión en el puente.")
        st.info("💡 Vaya a los módulos **Costos** e **Inversión** primero para poblar los datos.")
        if datos_cult.get('errores'):
            for err in datos_cult['errores']:
                st.caption(f"• Cultivos: {err}")
        if datos_inv.get('errores'):
            for err in datos_inv['errores']:
                st.caption(f"• Inversión: {err}")
        st.stop() 
 
    datos_puente = {
        'cultivos': datos_cult,
        'inversion': datos_inv
    }
    
    if datos_inv['inversion_total'] == 0 and datos_cult['ingreso_incremental_anual'] == 0:
        st.info("ℹ️ El proyecto activo no tiene inversión ni ingresos registrados.")
    
    # ============================================================
    # MATRICES FINANCIERA Y ECONÓMICA
    # ============================================================
    df_fin = construir_matriz_financiera(config, datos_puente, fin_params)
    df_fin_sin_fin = construir_matriz_financiera(config, datos_puente, None)
    df_eco = construir_matriz_economica(config, df_fin_sin_fin)
    
    ind_fin = calcular_indicadores(df_fin, config.tasa_privada_descuento, anios)
    ind_eco = calcular_indicadores(df_eco, config.tasa_social_descuento, anios)
    
    # ============================================================
    # TABS
    # ============================================================
    tab_fin, tab_eco, tab_sost, tab_ind, tab_sens, tab_riesgo, tab_reportes = st.tabs([
        "💵 Financiera", 
        "🌍 Económica", 
        "🛡️ Sostenibilidad",
        "📊 Indicadores y Costo-Eficiencia",
        "📈 Sensibilidad",
        "⚠️ Riesgo",
        "📑 Reportes"
    ])
    
    # =================== TAB FINANCIERA ===================
    with tab_fin:
        st.header("EVALUACIÓN FINANCIERA (Precios de Mercado)")
        st.caption(f"Tasa privada de descuento: {config.tasa_privada_descuento*100}% | Años: {anios[0]} - {anios[-1]}")
        
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("VANF", f"Bs {ind_fin['VAN']:,.0f}")
        c2.metric("TIRF", f"{ind_fin['TIR']*100:.2f}%" if ind_fin['TIR'] else "N/A")
        c3.metric("Relación B/C", f"{ind_fin['Relacion_BC']:.2f}")
        c4.metric("CAEF", f"Bs {ind_fin['CAE']:,.0f}")
        
        st.markdown("---")
        st.subheader("Flujo de Fondos Horizontal - E_FIN")
        
        def formatear_df(df):
            return df.style.format("{:,.0f}").map(
                lambda x: "color: red" if isinstance(x, (int, float)) and x < 0 else None
            )        

        st.dataframe(formatear_df(df_fin), use_container_width=True)
    
    # =================== TAB ECONÓMICA ===================
    with tab_eco:
        st.header("EVALUACIÓN ECONÓMICA (Precios de Cuenta)")
        st.caption(f"Tasa social de descuento: {config.tasa_social_descuento*100}% | RPC aplicados a costos")
        
        # NUEVO: Mostrar corrección de ingresos
        factor_ing = (1 - config.pct_produccion_transable) + (config.pct_produccion_transable * config.rpc['divisa'])
        col_info1, col_info2 = st.columns(2)
        with col_info1:
            st.info(f"🌾 **Producción transable:** {config.pct_produccion_transable*100:.0f}%")
        with col_info2:
            st.info(f"💱 **Factor corrección ingresos:** {factor_ing:.4f} "
                    f"(={(1-config.pct_produccion_transable):.2f}×1 + {config.pct_produccion_transable:.2f}×{config.rpc['divisa']})")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("VANE", f"Bs {ind_eco['VAN']:,.0f}")
        c2.metric("TIRE", f"{ind_eco['TIR']*100:.2f}%" if ind_eco['TIR'] else "N/A")
        c3.metric("Relación B/C Econ.", f"{ind_eco['Relacion_BC']:.2f}")
        c4.metric("CAEE", f"Bs {ind_eco['CAE']:,.0f}")
        
        with st.expander("Ver factores RPC aplicados"):
            rpc_df = pd.DataFrame({
                'Recurso': ['Divisa/Transables', 'MO Calificada', 'MO Semicalificada', 
                           'MO No Calificada Urbana', 'MO No Calificada Rural', 'Bienes No Transables'],
                'RPC': [config.rpc['divisa'], config.rpc['mo_calificada'], config.rpc['mo_semicalificada'],
                       config.rpc['mo_no_calif_urbana'], config.rpc['mo_no_calif_rural'], 1.0],
                'Significado': ['Costo oportunidad divisas', 'Costo social alta calificación',
                              'Costo social media calif.', 'Costo social mano obra urbana',
                              'Costo social mano obra rural', 'Sin corrección']
            })
            st.dataframe(rpc_df, use_container_width=True)
        
        st.markdown("---")
        st.subheader("Flujo de Fondos Horizontal - E_ECO")
        st.dataframe(formatear_df(df_eco), use_container_width=True)

    # =================== TAB SOSTENIBILIDAD ===================
    with tab_sost:
        st.header("🛡️ Sostenibilidad Operativa")
        st.caption("Art. 16.9 RM 115/2015 — Verificación de que ingresos cubran costos de operación y mantenimiento")
        
        ind_sost = calcular_sostenibilidad_operativa(config, df_fin, anios)
        
        # KPIs principales
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Ratio Cobertura Prom.", f"{ind_sost['ratio_cobertura_promedio']:.2f}x",
                 help="Ingresos / Costos operativos promedio del período")
        c2.metric("Año Equilibrio", 
                 str(ind_sost['anio_equilibrio']) if ind_sost['anio_equilibrio'] else "No alcanza",
                 help="Primer año donde ingresos ≥ costos operativos")
        c3.metric("Años Déficit Inicial", str(ind_sost['anos_deficit_consecutivos']),
                 help="Años consecutivos con pérdida operativa desde inicio")
        c4.metric("¿Es Sostenible?", "✅ Sí" if ind_sost['sostenible'] else "❌ No",
                 help="Alcanza equilibrio operativo durante el período")
        
        st.markdown("---")
        
        # Tabla detallada
        df_sost = pd.DataFrame({
            'Año': anios,
            'Ingresos (Bs)': ind_sost['ingresos'],
            'Costos Operativos (Bs)': ind_sost['costos_operativos'],
            'Superávit / Déficit (Bs)': ind_sost['superavit_operativo'],
            'Ratio Cobertura': ind_sost['ratio_cobertura'],
        })
        
        def color_superavit(val):
            if isinstance(val, (int, float)):
                if val > 0:
                    return 'background-color: #d4edda; color: #155724'  # verde
                elif val < 0:
                    return 'background-color: #f8d7da; color: #721c24'  # rojo
            return ''
        
        st.dataframe(
            df_sost.style
            .format({
                'Ingresos (Bs)': '{:,.0f}',
                'Costos Operativos (Bs)': '{:,.0f}',
                'Superávit / Déficit (Bs)': '{:,.0f}',
                'Ratio Cobertura': '{:.2f}x'
            })
            .map(color_superavit, subset=['Superávit / Déficit (Bs)']),
            use_container_width=True,
            hide_index=True
        )
        
        # Recomendación
        st.markdown("---")
        st.subheader("🎯 Diagnóstico de Sostenibilidad")
        if ind_sost['sostenible']:
            if ind_sost['anos_deficit_consecutivos'] == 0:
                st.success("### ✅ SOSTENIBILIDAD PLENA DESDE EL INICIO")
                st.write("El proyecto genera superávit operativo desde el primer año de operación. "
                         "Los ingresos cubren holgadamente los costos de O&M y producción.")
            else:
                st.warning("### ⚠️ SOSTENIBILIDAD DIFERIDA")
                st.write(f"El proyecto alcanza el punto de equilibrio operativo en el año **{ind_sost['anio_equilibrio']}**. "
                         f"Requiere **{ind_sost['anos_deficit_consecutivos']} años** de apoyo inicial para cubrir déficits operativos.")
        else:
            st.error("### ❌ INSOSTENIBLE OPERATIVAMENTE")
            st.write("Los ingresos nunca alcanzan a cubrir los costos operativos durante el período de diseño. "
                     "Se requiere revisar precios, rendimientos o costos de producción.")
    
    # =================== TAB INDICADORES ===================
    with tab_ind:
        st.header("Indicadores y Costo-Eficiencia")
        
        comp = pd.DataFrame({
            'Indicador': ['VAN', 'TIR (%)', 'Relación B/C', 'CAE (Bs)'],
            'Financiero': [
                f"{ind_fin['VAN']:,.0f}",
                f"{ind_fin['TIR']*100:.2f}" if ind_fin['TIR'] else "N/A",
                f"{ind_fin['Relacion_BC']:.2f}",
                f"{ind_fin['CAE']:,.0f}"
            ],
            'Económico': [
                f"{ind_eco['VAN']:,.0f}",
                f"{ind_eco['TIR']*100:.2f}" if ind_eco['TIR'] else "N/A",
                f"{ind_eco['Relacion_BC']:.2f}",
                f"{ind_eco['CAE']:,.0f}"
            ]
        })
        st.subheader("Comparativo Financiero vs Económico")
        st.table(comp)
        
        st.subheader("Indicadores de Costo-Eficiencia")
        inv_total = datos_inv['inversion_total']
        area_inc = config.area_incremental if config.area_incremental > 0 else 1
        familias = config.total_familias if config.total_familias > 0 else 1
        poblacion = config.poblacion_base if config.poblacion_base > 0 else 1
        
        efic = pd.DataFrame({
            'Indicador': [
                'Inversión por Hectárea',
                'Inversión por Familia',
                'CAEF por Hectárea',
                'CAEE por Hectárea',
                'CAEF por Población',
                'CAEE por Población',
                'Costo Inversión / Beneficiario'
            ],
            'Valor': [
                f"Bs {inv_total/area_inc:,.0f}",
                f"Bs {inv_total/familias:,.0f}",
                f"Bs {ind_fin['CAE']/area_inc:,.0f}",
                f"Bs {ind_eco['CAE']/area_inc:,.0f}",
                f"Bs {ind_fin['CAE']/poblacion:,.0f}",
                f"Bs {ind_eco['CAE']/poblacion:,.0f}",
                f"Bs {inv_total/poblacion:,.0f}"
            ],
            'Unidad': ['Bs/Ha', 'Bs/Fam', 'Bs/Ha/año', 'Bs/Ha/año', 'Bs/hab/año', 'Bs/hab/año', 'Bs/hab']
        })
        st.dataframe(efic, use_container_width=True)
        
        st.markdown("---")
        st.subheader("🎯 Recomendación de Decisión")
        vanf_pos = ind_fin['VAN'] > 0
        vane_pos = ind_eco['VAN'] > 0
        tirf_ok = ind_fin['TIR'] and ind_fin['TIR'] > config.tasa_privada_descuento
        tire_ok = ind_eco['TIR'] and ind_eco['TIR'] > config.tasa_social_descuento
        
        if vane_pos and vanf_pos and tire_ok and tirf_ok:
            st.success("### ✅ EJECUTAR EL PROYECTO")
            st.write("Rentable social y financieramente. Aporta bienestar y es sostenible.")
            st.balloons()
        elif vane_pos and tire_ok:
            st.warning("### ⚠️ EJECUTAR CON PRECAUCIÓN")
            st.write(f"VANE positivo (Bs {ind_eco['VAN']:,.0f}) pero VANF = Bs {ind_fin['VAN']:,.0f}.")
        else:
            st.error("### ❌ NO EJECUTAR")
            if not vane_pos:
                st.write(f"VANE negativo (Bs {ind_eco['VAN']:,.0f}).")
            if not tire_ok:
                st.write(f"TIRE ({ind_eco['TIR']*100:.2f}%) inferior a tasa social.")

    # =================== TAB SENSIBILIDAD ===================
    with tab_sens:
        st.header("📈 Análisis de Sensibilidad")
        st.markdown("Ajuste los parámetros clave para evaluar el impacto en los indicadores financieros y económicos.")

        col1, col2 = st.columns([1, 2])
        with col1:
            st.subheader("Variaciones porcentuales")
            var_inv = st.slider("Inversión", -50, 50, 0, 5, format="%d%%") / 100
            var_om = st.slider("Costos O&M", -50, 50, 0, 5, format="%d%%") / 100
            var_prod = st.slider("Costos de Producción", -50, 50, 0, 5, format="%d%%") / 100
            var_ing = st.slider("Ingresos por Ventas", -50, 50, 0, 5, format="%d%%") / 100
            var_tpriv_pp = st.slider("Tasa Privada (puntos %)", -5.0, 5.0, 0.0, 0.5, format="%.2f") / 100
            var_tsoc_pp = st.slider("Tasa Social (puntos %)", -5.0, 5.0, 0.0, 0.5, format="%.2f") / 100

            recalcular = st.button("🔄 Recalcular", use_container_width=True)

        with col2:
            if recalcular:
                df_fin_mod, df_eco_mod, ind_fin_mod, ind_eco_mod = recalcular_con_variacion(
                    config, datos_puente, var_inv, var_om, var_prod, var_ing,
                    var_tpriv_pp, var_tsoc_pp
                )
            else:
                df_fin_mod, df_eco_mod = df_fin, df_eco
                ind_fin_mod, ind_eco_mod = ind_fin, ind_eco

            # Guardar en session_state para el reporte Excel - Por confirmar este bloque
            st.session_state['sensibilidad_personalizada'] = {
                'var_inv': var_inv, 'var_om': var_om, 'var_prod': var_prod,
                'var_ing': var_ing, 'var_tpriv': var_tpriv_pp, 'var_tsoc': var_tsoc_pp,
                'vanf_base': ind_fin['VAN'], 'vanf_mod': ind_fin_mod['VAN'],
                'tirf_base': ind_fin['TIR']*100 if ind_fin['TIR'] else 0,
                'tirf_mod': ind_fin_mod['TIR']*100 if ind_fin_mod['TIR'] else 0,
                'bcf_base': ind_fin['Relacion_BC'], 'bcf_mod': ind_fin_mod['Relacion_BC'],
                'caef_base': ind_fin['CAE'], 'caef_mod': ind_fin_mod['CAE'],
                'vane_base': ind_eco['VAN'], 'vane_mod': ind_eco_mod['VAN'],
                'tire_base': ind_eco['TIR']*100 if ind_eco['TIR'] else 0,
                'tire_mod': ind_eco_mod['TIR']*100 if ind_eco_mod['TIR'] else 0,
                'bce_base': ind_eco['Relacion_BC'], 'bce_mod': ind_eco_mod['Relacion_BC'],
                'caee_base': ind_eco['CAE'], 'caee_mod': ind_eco_mod['CAE'],
            }

            # Tabla comparativa
            data_comp = {
                'Indicador': [
                    'VAN Financiero', 'TIR Financiero (%)', 'B/C Financiero', 'CAE Financiero',
                    'VAN Económico', 'TIR Económico (%)', 'B/C Económico', 'CAE Económico'
                ],
                'Escenario Base': [
                    f"{ind_fin['VAN']:,.0f}",
                    f"{ind_fin['TIR']*100:.2f}" if ind_fin['TIR'] else "N/A",
                    f"{ind_fin['Relacion_BC']:.2f}",
                    f"{ind_fin['CAE']:,.0f}",
                    f"{ind_eco['VAN']:,.0f}",
                    f"{ind_eco['TIR']*100:.2f}" if ind_eco['TIR'] else "N/A",
                    f"{ind_eco['Relacion_BC']:.2f}",
                    f"{ind_eco['CAE']:,.0f}",
                ],
                'Escenario Modificado': [
                    f"{ind_fin_mod['VAN']:,.0f}",
                    f"{ind_fin_mod['TIR']*100:.2f}" if ind_fin_mod['TIR'] else "N/A",
                    f"{ind_fin_mod['Relacion_BC']:.2f}",
                    f"{ind_fin_mod['CAE']:,.0f}",
                    f"{ind_eco_mod['VAN']:,.0f}",
                    f"{ind_eco_mod['TIR']*100:.2f}" if ind_eco_mod['TIR'] else "N/A",
                    f"{ind_eco_mod['Relacion_BC']:.2f}",
                    f"{ind_eco_mod['CAE']:,.0f}",
                ]
            }
            df_comp = pd.DataFrame(data_comp)
            st.subheader("Comparación de Indicadores")
            st.dataframe(df_comp, use_container_width=True, hide_index=True)

            # Comparación de VAN (tabla resumen)
            van_resumen = pd.DataFrame({
                'Escenario': ['Financiero Base', 'Financiero Modificado', 
                              'Económico Base', 'Económico Modificado'],
                'VAN (Bs)': [ind_fin['VAN'], ind_fin_mod['VAN'], 
                             ind_eco['VAN'], ind_eco_mod['VAN']]
            })
            st.dataframe(van_resumen.style.format({'VAN (Bs)': '{:,.0f}'}), 
                        use_container_width=True, hide_index=True)

            # Indicadores de costo-eficiencia modificados
            st.subheader("Indicadores de Costo-Eficiencia Modificados")
            area_inc = config.area_incremental if config.area_incremental > 0 else 1
            familias = config.total_familias if config.total_familias > 0 else 1
            poblacion = config.poblacion_base if config.poblacion_base > 0 else 1
            inv_total_mod = datos_puente['inversion']['inversion_total'] * (1 + var_inv)

            ce_mod = pd.DataFrame({
                'Indicador': [
                    'Inversión por Hectárea',
                    'Inversión por Familia',
                    'CAEF por Hectárea',
                    'CAEE por Hectárea',
                    'CAEF por Población',
                    'CAEE por Población'
                ],
                'Valor Modificado': [
                    f"Bs {inv_total_mod/area_inc:,.0f}",
                    f"Bs {inv_total_mod/familias:,.0f}",
                    f"Bs {ind_fin_mod['CAE']/area_inc:,.0f}",
                    f"Bs {ind_eco_mod['CAE']/area_inc:,.0f}",
                    f"Bs {ind_fin_mod['CAE']/poblacion:,.0f}",
                    f"Bs {ind_eco_mod['CAE']/poblacion:,.0f}"
                ]
            })
            st.dataframe(ce_mod, use_container_width=True, hide_index=True)

    # =================== TAB RIESGO (DUAL: MONTECARLO + PROBABILÍSTICO) ===================
    with tab_riesgo:
        st.header("⚠️ Análisis de Riesgos")
        st.caption("Seleccione la metodología y configure los parámetros. Los resultados se vinculan automáticamente al reporte Excel.")

        # --- SELECTOR DE MÉTODO ---
        if 'metodo_riesgo_activo' not in st.session_state:
            st.session_state['metodo_riesgo_activo'] = "Probabilístico RM 115/2015"

        metodo_riesgo = st.radio(
            "🧮 Metodología de evaluación",
            options=["Probabilístico RM 115/2015", "Simulación Monte Carlo"],
            index=0 if st.session_state['metodo_riesgo_activo'] == "Probabilístico RM 115/2015" else 1,
            horizontal=True,
            help="RM 115: método trazable de 6 pasos con escenarios definidos. Monte Carlo: simulación estocástica con distribuciones de probabilidad."
        )
        st.session_state['metodo_riesgo_activo'] = metodo_riesgo
        st.divider()

        # ============================================================
        # MÉTODO A: PROBABILÍSTICO RM 115/2015
        # ============================================================
        if metodo_riesgo == "Probabilístico RM 115/2015":
            st.subheader("📋 Metodología Probabilística RM 115/2015")

            # Inicializar estado
            if 'riesgo_rm115_params' not in st.session_state:
                st.session_state['riesgo_rm115_params'] = {
                    'riesgos': copy.deepcopy(RIESGOS_DEFAULT),
                    'prob_opt': 0.20,
                    'prob_esp': 0.55,
                    'prob_adv': 0.25,
                }
            if 'riesgo_rm115_resultado' not in st.session_state:
                st.session_state['riesgo_rm115_resultado'] = None

            rp = st.session_state['riesgo_rm115_params']

            # --- CONFIGURACIÓN DE ESCENARIOS ---
            st.markdown("**🎚️ Probabilidades de los 3 escenarios**")
            p_opt = st.slider("Optimista — ningún riesgo se materializa", 0.0, 1.0, rp['prob_opt'], 0.01)
            p_esp = st.slider("Esperado — pérdida ponderada por probabilidad", 0.0, 1.0, rp['prob_esp'], 0.01)
            p_adv = st.slider("Adverso — todos los riesgos con impacto pleno", 0.0, 1.0, rp['prob_adv'], 0.01)

            total_p = p_opt + p_esp + p_adv
            if abs(total_p - 1.0) > 0.001:
                st.warning(f"⚠️ La suma de probabilidades es {total_p*100:.0f}%. Se normalizará automáticamente al ejecutar.")
            else:
                st.success(f"✅ Suma de probabilidades = 100%")

            # --- CATÁLOGO DE RIESGOS EDITABLE ---
            st.markdown("**📋 Catálogo de riesgos identificados**")
            df_riesgos_edit = pd.DataFrame([
                {
                    'Riesgo': r['nombre'],
                    'Probabilidad': r['probabilidad'],
                    'Impacto': r['impacto'],
                    'Afecta': r['afecta'],
                    'Fase': r['fase'],
                }
                for r in rp['riesgos']
            ])

            df_edited = st.data_editor(
                df_riesgos_edit,
                column_config={
                    'Probabilidad': st.column_config.NumberColumn(min_value=0.0, max_value=1.0, step=0.01, format="%.2f"),
                    'Impacto': st.column_config.NumberColumn(min_value=0.0, max_value=1.0, step=0.01, format="%.2f"),
                    'Afecta': st.column_config.SelectboxColumn(options=['ingresos', 'costos']),
                    'Fase': st.column_config.SelectboxColumn(options=['inversion', 'operacion', 'todos']),
                },
                num_rows="dynamic",
                use_container_width=True,
                key="riesgos_editor_rm115"
            )

            if st.button("🔄 Ejecutar Análisis RM 115", type="primary", use_container_width=True):
                riesgos_sync = []
                for _, row in df_edited.iterrows():
                    riesgos_sync.append({
                        'nombre': row['Riesgo'],
                        'probabilidad': float(row['Probabilidad']),
                        'impacto': float(row['Impacto']),
                        'afecta': row['Afecta'],
                        'fase': row['Fase'],
                    })
                rp['riesgos'] = riesgos_sync
                rp['prob_opt'] = p_opt
                rp['prob_esp'] = p_esp
                rp['prob_adv'] = p_adv

                with st.spinner("Calculando matriz de riesgos y escenarios..."):
                    res = analizar_riesgo_rm115(
                        config, df_fin, ind_fin,
                        riesgos=riesgos_sync,
                        prob_optimista=p_opt,
                        prob_esperado=p_esp,
                        prob_adverso=p_adv,
                    )
                    st.session_state['riesgo_rm115_resultado'] = res
                st.success("✅ Análisis RM 115 completado")

            # --- RESULTADOS ---
            if st.session_state['riesgo_rm115_resultado'] is not None:
                st.divider()
                st.markdown("### 📊 Resultados del Análisis")
                res = st.session_state['riesgo_rm115_resultado']

                # PASO 1: Matriz de riesgos
                st.markdown("**Paso 1 — Matriz de Riesgos (Probabilidad × Impacto)**")
                df_matriz = pd.DataFrame(res['matriz_riesgos'])
                df_matriz['Nivel'] = df_matriz['nivel_riesgo'].apply(lambda x: f"{x:.3f}")
                df_matriz['Probabilidad'] = df_matriz['probabilidad'].apply(lambda x: f"{x:.0%}")
                df_matriz['Impacto'] = df_matriz['impacto'].apply(lambda x: f"{x:.0%}")
                df_display = df_matriz[['nombre', 'Probabilidad', 'Impacto', 'Nivel', 'clasificacion', 'afecta', 'fase']].copy()
                df_display.columns = ['Riesgo', 'Prob.', 'Impacto', 'Nivel', 'Clasificación', 'Afecta a', 'Fase']

                def color_clasificacion(val):
                    color = calcular_color_riesgo(val)
                    if val == 'Crítico':
                        return f'background-color: {color}; color: white; font-weight: bold'
                    return f'background-color: {color}; color: black'

                st.dataframe(
                    df_display.style.map(color_clasificacion, subset=['Clasificación']),
                    use_container_width=True, hide_index=True
                )

                # PASO 2: Pérdida esperada
                st.markdown("**Paso 2 — Pérdida Esperada Anual**")
                df_perdida = pd.DataFrame({
                    'Año': anios,
                    'Pérdida Ingresos (Bs)': res['perdida_esperada_anual']['ingresos'],
                    'Pérdida Costos (Bs)': res['perdida_esperada_anual']['costos'],
                    'Pérdida Total (Bs)': res['perdida_esperada_anual']['total'],
                })
                st.dataframe(df_perdida.style.format('{:,.0f}'), use_container_width=True, hide_index=True)

                # PASO 3: Escenarios
                '''
                st.markdown("**Paso 3 — Escenarios de VAN**")
                c1, c2, c3 = st.columns(3)
                c1.metric("VAN Optimista", f"Bs {res['van_base']:,.0f}")
                c2.metric("VAN Esperado", f"Bs {res['van_esperado']:,.0f}")
                c3.metric("VAN Adverso", f"Bs {res['van_peor']:,.0f}")

                esc_data = pd.DataFrame({
                    'Escenario': ['Optimista', 'Esperado', 'Adverso'],
                    'VAN': [res['van_base'],
                            res['escenarios']['Esperado (pérdida ponderada por probabilidad)']['van'],
                            res['van_peor']],
                    'Probabilidad': [res['probabilidades_escenarios']['optimista'],
                                     res['probabilidades_escenarios']['esperado'],
                                     res['probabilidades_escenarios']['adverso']]
                })
                fig, ax = plt.subplots(figsize=(7, 3.5))
                colors = ['#28a745', '#ffc107', '#dc3545']
                bars = ax.bar(esc_data['Escenario'], esc_data['VAN'], color=colors, edgecolor='black')
                ax.axhline(y=0, color='black', linestyle='-', linewidth=0.8)
                ax.set_ylabel("VAN (Bs)")
                for bar, prob in zip(bars, esc_data['Probabilidad']):
                    h = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., h,
                           f'{h:,.0f}\n({prob*100:.0f}%)',
                           ha='center', va='bottom' if h >= 0 else 'top', fontsize=9)
                st.pyplot(fig)
                plt.close()

                '''# PASO 4: Sensibilidad
                st.markdown("**Paso 4 — Sensibilidad de Quiebre**")
                s = res['sensibilidad_umbral']
                cq1, cq2 = st.columns(2)
                cq1.metric("Caída máxima de Ingresos", f"{abs(s['variacion_critica_ingresos_pct'])*100:.1f}%" if s['variacion_critica_ingresos_pct'] else "N/A")
                cq1.caption(s['lectura_ingresos'])
                cq2.metric("Incremento máximo de Costos", f"{abs(s['variacion_critica_costos_pct'])*100:.1f}%" if s['variacion_critica_costos_pct'] else "N/A")
                cq2.caption(s['lectura_costos'])

                # PASO 5: Probabilidad conjunta
                st.markdown("**Paso 5 — Probabilidad Conjunta de Falla**")
                st.metric("P(Falla conjunta)", f"{res['prob_falla_conjunta']*100:.1f}%",
                         help="Probabilidad de que ocurra al menos uno de los riesgos identificados (asumiendo independencia)")

                # PASO 6: Decisión
                st.markdown("**Paso 6 — Regla de Decisión**")
                dec = res['nivel_decision']
                if "ROBUSTO" in dec['nivel']:
                    st.success(f"### {dec['color']} {dec['nivel']}")
                elif "VIABLE" in dec['nivel']:
                    st.warning(f"### {dec['color']} {dec['nivel']}")
                else:
                    st.error(f"### {dec['color']} {dec['nivel']}")
                st.write(dec['recomendacion'])

                with st.expander("Ver flujos de fondos detallados por escenario"):
                    df_flujos_comp = pd.DataFrame({
                        'Año': anios,
                        'Optimista': res['escenarios']['Optimista (sin riesgos)']['flujo'],
                        'Esperado': res['escenarios']['Esperado (pérdida ponderada por probabilidad)']['flujo'],
                        'Adverso': res['escenarios']['Adverso (impacto pleno de riesgos)']['flujo'],
                    })
                    st.dataframe(df_flujos_comp.style.format('{:,.0f}').map(
                        lambda x: 'color: red' if isinstance(x, (int, float)) and x < 0 else 'color: black'
                    ), use_container_width=True, hide_index=True)

        # ============================================================
        # MÉTODO B: MONTE CARLO
        # ============================================================
        else:
            st.subheader("🎲 Simulación Monte Carlo")

            if 'riesgo_mc_params' not in st.session_state:
                st.session_state['riesgo_mc_params'] = {
                    'n_sim': 1000, 'cv_precios': 0.20, 'cv_rend': 0.15,
                    'cv_costos': 0.10, 'cv_om': 0.10, 'cv_inv': 0.10,
                }
            if 'riesgo_mc_resultado' not in st.session_state:
                st.session_state['riesgo_mc_resultado'] = None

            mc = st.session_state['riesgo_mc_params']

            n_sim = st.number_input("N° de simulaciones", 100, 10000, mc['n_sim'], 100, key="mc_nsim")
            cv_precios = st.slider("Coef. Variación — Precios", 0.05, 0.50, mc['cv_precios'], 0.01, key="mc_precios")
            cv_rend = st.slider("Coef. Variación — Rendimientos", 0.05, 0.50, mc['cv_rend'], 0.01, key="mc_rend")
            cv_costos = st.slider("Coef. Variación — Costos Producción", 0.05, 0.50, mc['cv_costos'], 0.01, key="mc_costos")
            cv_om = st.slider("Coef. Variación — O&M", 0.05, 0.50, mc['cv_om'], 0.01, key="mc_om")
            cv_inv = st.slider("Coef. Variación — Inversión", 0.05, 0.50, mc['cv_inv'], 0.01, key="mc_inv")

            if st.button("🔄 Ejecutar Simulación Monte Carlo", type="primary", use_container_width=True):
                mc['n_sim'] = n_sim; mc['cv_precios'] = cv_precios; mc['cv_rend'] = cv_rend
                mc['cv_costos'] = cv_costos; mc['cv_om'] = cv_om; mc['cv_inv'] = cv_inv

                with st.spinner(f"Simulando {n_sim} escenarios..."):
                    res_mc = simular_montecarlo(
                        config, datos_puente, n_simulaciones=n_sim,
                        cv_precios=cv_precios, cv_rendimientos=cv_rend,
                        cv_costos_prod=cv_costos, cv_om=cv_om, cv_inversion=cv_inv
                    )
                    st.session_state['riesgo_mc_resultado'] = res_mc
                st.success("✅ Simulación completada")

            if st.session_state['riesgo_mc_resultado'] is not None:
                st.divider()
                st.markdown("### 📊 Resultados de la Simulación")
                res = st.session_state['riesgo_mc_resultado']

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("VAN Esperado", f"Bs {res['van_mean']:,.0f}")
                c2.metric("Prob. Falla", f"{res['prob_falla']*100:.1f}%")
                c3.metric("VaR 10%", f"Bs {res['var_10']:,.0f}")
                c4.metric("TIR Esperada", f"{res['tir_mean']:.2f}%" if not np.isnan(res['tir_mean']) else "N/A")

                st.markdown("**Distribución del VAN**")
                stats_df = pd.DataFrame({
                    'Estadístico': ['Media', 'Mediana (P50)', 'Desv. Estándar', 'Mínimo (P5)', 'Máximo (P95)', 'VaR 10%'],
                    'Valor (Bs)': [res['van_mean'], res['van_percentiles'][3], res['van_std'],
                                   res['van_percentiles'][0], res['van_percentiles'][-1], res['var_10']]
                })
                st.dataframe(stats_df.style.format({'Valor (Bs)': '{:,.0f}'}), use_container_width=True, hide_index=True)

                percentiles_df = pd.DataFrame({
                    'Percentil': ['5%', '10%', '25%', '50%', '75%', '90%', '95%'],
                    'VAN (Bs)': [f"{p:,.0f}" for p in res['van_percentiles']]
                })
                st.dataframe(percentiles_df, hide_index=True, use_container_width=True)

                st.markdown("**Interpretación**")
                if res['prob_falla'] < 0.05:
                    st.success("✅ **Riesgo muy bajo**: Probabilidad de pérdida < 5%. Proyecto robusto.")
                elif res['prob_falla'] < 0.20:
                    st.warning(f"⚠️ **Riesgo moderado**: Probabilidad de pérdida del {res['prob_falla']*100:.1f}%. Planificar contingencias.")
                else:
                    st.error(f"🔴 **Riesgo elevado**: Probabilidad de pérdida del {res['prob_falla']*100:.1f}%. Revisar supuestos.")

                st.download_button(
                    label="📥 Descargar resultados (CSV)",
                    data=pd.DataFrame({'VAN': res['vans'], 'TIR': res['tirs']}).to_csv(index=False),
                    file_name="montecarlo_resultados.csv",
                    mime="text/csv"
                )
    # ============================================================
    # TAB REPORTES 
    # ============================================================
    with tab_reportes:
        st.header("📑 Reportes de Evaluación")

        # --- Producción Agrícola ---
        st.markdown("---")
        st.subheader("Producción Agrícola")

        df_vn = generar_df_valor_neto_produccion(datos_cult)
        df_rpc = generar_df_costos_rpc_incremental(datos_cult)
        
        st.markdown("**VALOR NETO DE PRODUCCIÓN (SP vs CP vs Incremental)**")
        st.dataframe(
            df_vn.style.format(lambda x: f"{x:,.2f}" if isinstance(x, (int, float)) else x),
            use_container_width=True,
            height=400
        )
        
        st.markdown("**DESGLOSE RPC DE COSTOS INCREMENTALES**")
        st.dataframe(
            df_rpc.style.format(lambda x: f"{x:,.2f}" if isinstance(x, (int, float)) else x),
            use_container_width=True,
            height=300
        )
        
        st.markdown("---")
        m1, m2, m3, m4 = st.columns(4)
        total_incr_ing = df_vn['Δ Ingreso (Bs)'].iloc[:-1].sum()
        total_incr_costo = df_vn['Δ Costo (Bs)'].iloc[:-1].sum()
        total_incr_vneto = df_vn['Δ V.Neto (Bs)'].iloc[:-1].sum()
        sup_total_cp = df_vn['Sup. CP (Ha)'].iloc[:-1].sum()
        
        m1.metric("Ingreso Incremental Total", f"Bs {total_incr_ing:,.0f}")
        m2.metric("Costo Incremental Total", f"Bs {total_incr_costo:,.0f}")
        m3.metric("Valor Neto Incremental", f"Bs {total_incr_vneto:,.0f}")
        m4.metric("Superficie Total CP", f"{sup_total_cp:,.2f} Ha")

        # Descarga
        st.markdown("---")
        st.subheader("Descarga de Reportes Completos")
        excel_bytes = generar_workbook_reportes(config, datos_puente, df_fin, ind_fin, df_eco, ind_eco)
        st.download_button(
            label="📥 Descargar Reporte Completo (.xlsx)",
            data=excel_bytes,
            file_name=f"Reportes_{config.nombre.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.caption("El archivo contiene las hojas PREP, E_FIN, E_ECO, Indicadores y Sensibilidad.")

if __name__ == "__main__":
    main()
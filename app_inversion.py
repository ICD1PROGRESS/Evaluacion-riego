import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List, Optional, Tuple, Set
from dataclasses import dataclass, field
import re
import io
import unicodedata

# ============================================================
# IMPORTAR CAPA CENTRALIZADA (Fase 2)
# ============================================================
from core.config import ConfiguracionProyecto
from core.data_manager import DataManager
from core.project_manager import ProjectManager
from core.database import ProyectoDB

pm = ProjectManager()
rutas = pm.get_rutas_activas()

if rutas is None:
    st.error("No hay proyecto activo")
    st.stop()

RUTA_EXCEL = rutas["excel"]
RUTA_CONFIG = rutas["config"]
RUTA_DB = os.path.join(os.path.dirname(RUTA_EXCEL), "proyecto.db")
# ============================================================
# CONSTANTES Y CONFIGURACIÓN
# ============================================================
TIPOS_RPC = {
    'bt': {'nombre': 'Bienes Transables', 'color': '#4472C4'},
    'bnt': {'nombre': 'Bienes No Transables', 'color': '#70AD47'},
    'moc': {'nombre': 'MO Calificada', 'color': '#FFC000'},
    'mos': {'nombre': 'MO Semicalificada', 'color': '#ED7D31'},
    'monu': {'nombre': 'MO No Calificada Urbana', 'color': '#5B9BD5'},
    'monr': {'nombre': 'MO No Calificada Rural', 'color': '#A5A5A5'},
    'N': {'nombre': 'No Clasificado', 'color': '#FF0000'}
}

CATEGORIAS = {
    'obras': {'nombre': 'Obras Civiles', 'icono': '🏗️', 'detalle': 'completo'},
    'ati': {'nombre': 'Asistencia Técnica Integral', 'icono': '👥', 'detalle': 'global'},
    'supervision': {'nombre': 'Supervisión de Obras', 'icono': '📋', 'detalle': 'global'},
    'om': {'nombre': 'Operación y Mantenimiento', 'icono': '🔧', 'detalle': 'global'},
    'ambiental': {'nombre': 'Mitigación Ambiental', 'icono': '🌱', 'detalle': 'global'}
}

# Mapeo: letra de sección (col 1) -> tipo_hoja interno (líneas de detalle)
SECCION_A_TIPO = {
    'A': 'materiales',
    'B': 'mano_obra',
    'C': 'equipo_maquinaria',
}

LETRA_A_COMPONENTE = {
    'F': 'cargas_sociales',
    'H': 'herramientas',
    'L': 'gastos_generales',
    'M': 'utilidad',
    'P': 'it',
    'O': 'iva',
    'Q': 'precio_unitario_final',
    # Los siguientes son subtotales informativos (no se persisten, pero pueden
    # ser útiles para validación): D, E, G, I, J, N
    'D': '_subtotal_materiales',
    'E': '_subtotal_mo_directa',
    'G': '_subtotal_mo_total',
    'I': '_subtotal_equipo_total',
    'J': '_subtotal_general',
    'N': '_parcial_sin_it',
}

COMPONENTES_PERSISTIBLES = (
    'cargas_sociales', 'herramientas', 'gastos_generales',
    'utilidad', 'it', 'iva', 'precio_unitario_final'
)

RPC_VALIDOS = ['bt', 'bnt', 'moc', 'mos', 'monu', 'monr']
RPC_MO = ['moc', 'mos', 'monu', 'monr']
RPC_NO_MO = ['bt', 'bnt']
# ============================================================
# MODELO DE DATOS
# ============================================================
@dataclass
class ItemObra:
    """Ítem detallado de obra civil (desde desglose_INVERSION.xlsx)"""
    id: int
    codigo_item: str
    descripcion: str
    unidad: str
    cantidad: float
    precio_unitario: float           # Precio del insumo directo (A/B/C)
    parcial_directo: float           # Parcial del insumo directo (= valor original Excel)
    tipo_hoja: str                   # 'materiales' | 'mano_obra' | 'equipo_maquinaria'
    subcategoria: str                # Nombre del APU al que pertenece
    tipo_rpc: str = 'N'
    fecha_clasificacion: str = ""
    precio_unitario_final: float = 0.0   # Q del APU (validación)
    parcial_real: float = 0.0            # parcial_directo * factor_apu
    factor_apu: float = 1.0
    indirectos_asignados: float = 0.0

    @property
    def parcial(self) -> float:
        """Valor económico total del ítem (directo + cuota proporcional de indirectos)."""
        if self.parcial_real > 0:
            return self.parcial_real
        return self.parcial_directo

    @property
    def precio_unitario_publicado(self) -> float:
        """Atajo al Q del APU."""
        return self.precio_unitario_final

    def recalcular_parcial_real(self) -> None:
        """Recalcula parcial_real a partir del factor APU."""
        if self.factor_apu > 0:
            self.parcial_real = self.parcial_directo * self.factor_apu
        else:
            self.parcial_real = self.parcial_directo
@dataclass
class ItemFOR5:
    """Ítem para el reporte Formato 5 - Presupuesto Desglosado"""
    nivel: int                       # 0=capítulo, 1=subcapítulo, 2=ítem, 3=subtotal, 4=total
    numero: str
    descripcion: str
    unidad: str = ""
    cantidad: float = 0.0
    precio_unitario: float = 0.0     # PU total (suma RPC por unidad)
    total: float = 0.0               # Cantidad × PU total
    bt: float = 0.0
    bnt: float = 0.0
    moc: float = 0.0
    mos: float = 0.0
    monu: float = 0.0
    monr: float = 0.0

@dataclass
class ServicioGlobal:
    """Servicios clasificados globalmente (ATI, Supervisión, OM)"""
    categoria: str
    descripcion: str
    monto_bt: float = 0.0
    monto_bnt: float = 0.0
    monto_moc: float = 0.0
    monto_mos: float = 0.0
    monto_monu: float = 0.0
    monto_monr: float = 0.0
    detalle_adjunto: str = ""

    @property
    def total(self) -> float:
        return (self.monto_bt + self.monto_bnt + self.monto_moc +
                self.monto_mos + self.monto_monu + self.monto_monr)

# ============================================================
# UTILIDADES
# ============================================================
def _to_float(x) -> float:
    """Normaliza números en formato español ('3,00' / '1.500,00') a float."""
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        return float(x)
    s = str(x).strip()
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _normalizar_descripcion(desc: str) -> str:
    """
    Normaliza descripciones para detectar insumos repetidos.
    - lowercase
    - sin acentos
    - colapsa espacios
    - quita puntuación trivial
    """
    if not isinstance(desc, str):
        return ""
    s = desc.strip().lower()
    # Quitar acentos
    s = ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )
    # Quitar puntuación trivial
    s = re.sub(r'[\.,;:\(\)/\\"]+', ' ', s)
    # Colapsar espacios
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# ============================================================
# GESTORES DE DATOS
# ============================================================
class GestorObrasCiviles:
    def __init__(self, ruta_desglose: str = None):
        self.ruta_desglose = ruta_desglose
        self.df_items = pd.DataFrame()
        self.subcategorias = []
        self.componentes_por_apu: Dict[str, Dict[str, float]] = {}
        self.precio_unitario_por_apu: Dict[str, float] = {}
        self.unidad_por_apu: Dict[str, str] = {}
        self.cantidad_por_apu: Dict[str, float] = {}
        self.asignacion_indirectos: Dict[str, Dict[str, str]] = {}

    def cargar_desde_excel(self) -> bool:
        """Carga el desglose desde una base de datos plana de APUs (Prescom o similar)."""
        if not os.path.exists(self.ruta_desglose):
            st.warning(f"No se encontró {self.ruta_desglose}")
            return False

        try:
            # Una sola hoja: la primera del archivo
            df = pd.read_excel(self.ruta_desglose, sheet_name=0, header=None)
        except Exception as e:
            st.error(f"Error abriendo el archivo APU: {e}")
            return False

        items = []
        componentes_por_apu: Dict[str, Dict[str, float]] = {}
        apu_actual = None       # Nombre del APU en proceso (subcategoria)
        seccion_actual = None   # 'materiales' | 'mano_obra' | 'equipo_maquinaria' | None
        q_por_apu: Dict[str, float] = {}   # cache de Q por APU para los ítems

        def _es_numero_item(x) -> bool:
            """Detecta si x corresponde a un número de ítem (entero positivo)."""
            if pd.isna(x):
                return False
            if isinstance(x, bool):
                return False
            if isinstance(x, (int, float)):
                return True
            if isinstance(x, str):
                s = x.strip()
                return s.isdigit()
            return False

        for _, row in df.iterrows():
            col0 = row[0] if len(row) > 0 else None
            col1 = row[1] if len(row) > 1 else None
            col2 = row[2] if len(row) > 2 else None
            col3 = row[3] if len(row) > 3 else None
            col4 = row[4] if len(row) > 4 else None
            col5 = row[5] if len(row) > 5 else None
            col6 = row[6] if len(row) > 6 else None

            # 1) Inicio de un nuevo APU con manejo de duplicados
            es_item = False
            nombre_apu = None
            for val in [col0, col1, col2, col3, col4, col5, col6]:
                if isinstance(val, str) and val.strip().lower().startswith('item:'):
                    nombre_apu = val.strip()[5:].strip()
                    es_item = True
                    break

            if es_item:
                seccion_actual = None
                # Generar clave única
                clave_apu = nombre_apu
                contador = 1
                while clave_apu in componentes_por_apu:
                    contador += 1
                    clave_apu = f"{nombre_apu} ({contador})"
                
                if clave_apu not in componentes_por_apu:
                    componentes_por_apu[clave_apu] = {c: 0.0 for c in COMPONENTES_PERSISTIBLES}
                
                apu_actual = clave_apu  # usar clave única en adelante
                
                # Guardar nombre original si lo deseas (opcional)
                # self.nombre_original_por_clave[clave_apu] = nombre_apu
                
                # Capturar cantidad y unidad
                cantidad_apu = 1.0
                unidad_apu = 'glb'
                if isinstance(col4, str):
                    cantidad_apu, unidad_apu = self._parsear_cantidad_unidad(col4.strip())
                elif isinstance(col5, str):
                    cantidad_apu, unidad_apu = self._parsear_cantidad_unidad(col5.strip())
                self.unidad_por_apu[apu_actual] = unidad_apu
                self.cantidad_por_apu[apu_actual] = cantidad_apu
                continue

            col0_es_numero = _es_numero_item(col0)

            # 2) Cabecera de sección: col 0 NO numérico, col 1 = letra A/B/C
            #    Acepta col0 NaN o cualquier marca de no-insumo (incluido '>')
            if (not col0_es_numero
                    and isinstance(col1, str)
                    and col1.strip() in SECCION_A_TIPO):
                if isinstance(col2, str):
                    seccion_actual = SECCION_A_TIPO[col1.strip()]
                continue

            # 3) Fila de componente indirecto: col 0 NO numérico (acepta '>', NaN, etc.),
            #    col 1 = letra F/H/L/M/P/O/Q. El valor está en col 6.
            if (not col0_es_numero
                    and isinstance(col1, str)
                    and col1.strip() in LETRA_A_COMPONENTE
                    and apu_actual is not None):
                letra = col1.strip()
                nombre_comp = LETRA_A_COMPONENTE[letra]
                if nombre_comp.startswith('_'):
                    # subtotal informativo: no persistir
                    continue
                valor = _to_float(col6)
                if valor != 0.0 or letra == 'Q':
                    # Para Q guardamos aunque sea 0 (es el precio final publicado)
                    componentes_por_apu[apu_actual][nombre_comp] = valor
                    if letra == 'Q':
                        q_por_apu[apu_actual] = valor
                continue

            # 4) Fila de insumo: col 0 numérico, col 1 == '-' (marcador de línea)
            if col0_es_numero and col1 == '-':
                # Si llegamos aquí sin APU o sin sección, la fila se ignora
                if not (apu_actual and seccion_actual):
                    continue
                try:
                    codigo = (
                        str(int(col0)) if not isinstance(col0, str)
                        else col0.strip()
                    )
                    desc = str(col2).strip() if pd.notna(col2) else ''
                    unidad = str(col3).strip() if pd.notna(col3) else 'glb'
                    cantidad = _to_float(col4)
                    unitario = _to_float(col5)
                    parcial = _to_float(col6)
                    q_apu = q_por_apu.get(apu_actual, 0.0)

                    if cantidad > 0 or unitario > 0:
                        items.append({
                            'codigo_item': codigo,
                            'descripcion': desc,
                            'descripcion_norm': _normalizar_descripcion(desc),
                            'unidad': unidad,
                            'cantidad': cantidad,
                            'precio_unitario': unitario,
                            'parcial_directo': parcial,
                            'tipo_hoja': seccion_actual,
                            'subcategoria': apu_actual,
                            'tipo_rpc': 'N',
                            'fecha_clasificacion': '',
                            'precio_unitario_final': q_apu,
                            'parcial_real': parcial,   # se recalcula abajo con factor APU
                            'factor_apu': 1.0,         # markup por indirectos del APU
                            'indirectos_asignados': 0.0,
                        })
                except (ValueError, IndexError):
                    continue

        self.df_items = pd.DataFrame(items)

        # Guardar componentes indirectos
        self.componentes_por_apu = componentes_por_apu
        self.precio_unitario_por_apu = {
            apu: comps.get('precio_unitario_final', 0.0)
            for apu, comps in componentes_por_apu.items()
        }

        # === DISTRIBUCIÓN PROPORCIONAL DE INDIRECTOS POR APU ===

        for apu, comps in componentes_por_apu.items():
            mask_apu = self.df_items['subcategoria'] == apu
            if not mask_apu.any():
                continue
            total_directo = self.df_items.loc[mask_apu, 'parcial_directo'].sum()
            q_apu = comps.get('precio_unitario_final', 0.0)
            if total_directo > 0 and q_apu > 0:
                factor = q_apu / total_directo
            else:
                factor = 1.0
            self.df_items.loc[mask_apu, 'parcial_real'] = (
                self.df_items.loc[mask_apu, 'parcial_directo'] * factor
            )
            self.df_items.loc[mask_apu, 'factor_apu'] = factor
            self.df_items.loc[mask_apu, 'indirectos_asignados'] = (
                self.df_items.loc[mask_apu, 'parcial_directo'] * (factor - 1)
            )
            self.df_items.loc[mask_apu, 'parcial'] = self.df_items.loc[mask_apu, 'parcial_real']

        if len(self.df_items) == 0:
            self.subcategorias = []
            st.error(
                "No se extrajeron ítems del archivo. "
                "Verifique que el archivo contenga bloques de APU con "
                "cabeceras de sección A (MATERIALES), B (MANO DE OBRA) y/o "
                "C (EQUIPO, MAQUINARIA Y HERRAMIENTAS)."
            )
            return False

        # Después: preservar orden de primera aparición (orden constructivo)
        seen = set()
        orden_apu = []
        for apu_key in self.df_items['subcategoria'].unique():
            orden_apu.append(apu_key)
        self.subcategorias = orden_apu

        return True
    
    def _parsear_cantidad_unidad(self, texto: str) -> Tuple[float, str]:
        """
        Extrae cantidad y unidad de celdas tipo:
          '1,00 glb' -> (1.0, 'glb')
        """
        s = texto.strip()
        if s.lower().startswith('unidad:'):
            return 1.0, s[7:].strip()
        # Patrón: número (con puntos de miles y/o coma decimal) + espacio + unidad
        match = re.match(r'^([\d\.,]+)\s+(.+)$', s)
        if match:
            num_str, unidad = match.groups()
            return _to_float(num_str), unidad.strip()
        # Fallback: si no coincide, asumir que todo es unidad
        return 1.0, s

    def cargar_desde_archivo(self, ruta: str) -> bool:
        """Carga un archivo de desglose desde cualquier ubicación."""
        if not os.path.exists(ruta):
            return False
        self.ruta_desglose = ruta
        return self.cargar_desde_excel()

    # --- NUEVO: helpers para la UI de clasificación inteligente ---
    def vista_insumos_unicos(self, subcat: Optional[str] = None) -> pd.DataFrame:

        if self.df_items is None or self.df_items.empty:
            return pd.DataFrame()

        df = self.df_items
        if subcat and subcat != 'TODAS':
            df = df[df['subcategoria'] == subcat]

        if df.empty:
            return pd.DataFrame()

        # Agrupar por descripción normalizada
        agg = df.groupby('descripcion_norm').agg(
            n_items=('codigo_item', 'count'),
            cantidad_total=('cantidad', 'sum'),
            parcial_directo_total=('parcial_directo', 'sum'),
            parcial_real_total=('parcial_real', 'sum'),
            apus_afectados=('subcategoria', lambda x: sorted(set(x))),
            tipos_rpc=('tipo_rpc', lambda x: list(x)),
        ).reset_index()

        # Descripción canónica = la más larga del grupo (suele ser la completa)
        def _canonica(grupo):
            descs = grupo['descripcion'].tolist()
            return max(descs, key=len) if descs else ''

        canon_map = (
            df.groupby('descripcion_norm')['descripcion']
            .apply(lambda s: max(s.tolist(), key=len))
        )
        unidad_map = (
            df.groupby('descripcion_norm')['unidad']
            .apply(lambda s: s.mode().iloc[0] if not s.mode().empty else 'glb')
        )
        tipo_hoja_map = (
            df.groupby('descripcion_norm')['tipo_hoja']
            .apply(lambda s: s.mode().iloc[0] if not s.mode().empty else 'materiales')
        )

        agg['descripcion_canon'] = agg['descripcion_norm'].map(canon_map)
        agg['unidad'] = agg['descripcion_norm'].map(unidad_map)
        agg['tipo_hoja'] = agg['descripcion_norm'].map(tipo_hoja_map)

        # RPC dominante
        def _rpc_mayoritario(grupo):
            rpcs = grupo['tipo_rpc'].tolist()
            if not rpcs:
                return 'N'
            # Si todos coinciden -> ese; si no, 'mixto'
            unicos = set(rpcs)
            if len(unicos) == 1:
                return next(iter(unicos))
            # Mayoritario
            from collections import Counter
            c = Counter(rpcs)
            return c.most_common(1)[0][0]

        rpc_map = df.groupby('descripcion_norm').apply(_rpc_mayoritario)
        agg['tipo_rpc'] = agg['descripcion_norm'].map(rpc_map)

        def _consistente(grupo):
            return len(set(grupo['tipo_rpc'].tolist())) == 1

        cons_map = df.groupby('descripcion_norm').apply(_consistente)
        agg['clasificacion_consistente'] = agg['descripcion_norm'].map(cons_map)

        # Ordenar por impacto económico
        agg = agg.sort_values('parcial_real_total', ascending=False).reset_index(drop=True)
        return agg

    def aplicar_clasificacion_por_grupo(
        self,
        descripcion_norm: str,
        tipo_rpc: str,
        sobrescribir: bool = True,
    ) -> Tuple[int, int]:
        if self.df_items is None or self.df_items.empty:
            return (0, 0)
        mask = self.df_items['descripcion_norm'] == descripcion_norm
        n_objetivo = int(mask.sum())
        if n_objetivo == 0:
            return (0, 0)

        if sobrescribir:
            self.df_items.loc[mask, 'tipo_rpc'] = tipo_rpc
            self.df_items.loc[mask, 'fecha_clasificacion'] = (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " (grupo)"
            )
            return (n_objetivo, 0)
        else:
            mask_n = mask & (self.df_items['tipo_rpc'] == 'N')
            n_actualizar = int(mask_n.sum())
            self.df_items.loc[mask_n, 'tipo_rpc'] = tipo_rpc
            self.df_items.loc[mask_n, 'fecha_clasificacion'] = (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " (grupo)"
            )
            return (n_actualizar, n_objetivo - n_actualizar)

    def actualizar_clasificacion(self, idx: int, tipo_rpc: str) -> bool:
        """Actualiza el tipo RPC de un ítem específico con alerta previa para N→BT"""
        if 0 <= idx < len(self.df_items):
            if self.df_items.at[idx, 'tipo_rpc'] == 'N' and tipo_rpc == 'bt':
                st.warning(
                    f"⚠️ **Alerta:** Ítem '{str(self.df_items.at[idx, 'descripcion'])[:30]}...' "
                    f"se clasifica como BT (previamente sin clasificar)"
                )
            self.df_items.at[idx, 'tipo_rpc'] = tipo_rpc
            self.df_items.at[idx, 'fecha_clasificacion'] = (
                datetime.now().strftime("%Y-%m-%d %H:%M")
            )
            return True
        return False

    def _mo_mayoritaria_apu(self, apu: str) -> str:
        """Determina la categoría MO mayoritaria en el APU (para usar cuando no hay MO directa)."""
        mask = (self.df_items['subcategoria'] == apu) & (self.df_items['tipo_hoja'] == 'mano_obra')
        if not mask.any():
            return 'monu'
        df_mo = self.df_items[mask]
        totales_mo = df_mo.groupby('tipo_rpc')['parcial_directo'].sum()
        totales_mo = totales_mo[totales_mo.index.isin(RPC_MO)]
        if totales_mo.empty:
            return 'monu'
        return str(totales_mo.idxmax())

    def calcular_totales_rpc(self) -> Dict[str, float]:
        """Suma los totales por RPC de todos los APUs."""
        if not self.subcategorias:
            return {rpc: 0.0 for rpc in RPC_VALIDOS}
        totales = {rpc: 0.0 for rpc in RPC_VALIDOS}
        
        for apu in self.subcategorias:
            t_unit = self.calcular_totales_rpc_por_apu(apu)
            cant = self.cantidad_por_apu.get(apu, 1.0)
            for rpc in RPC_VALIDOS:
                totales[rpc] += t_unit[rpc] * cant
        return totales
        
    def calcular_totales_rpc_por_apu(
        self,
        apu: str,
        override: Optional[Dict[str, str]] = None,
    ) -> Dict[str, float]:
        """
        Calcula los totales por RPC para un APU específico, incluyendo directos e indirectos.
        El resultado es el monto total para la cantidad de obra del APU.
        """
        if self.df_items is None or self.df_items.empty:
            return {rpc: 0.0 for rpc in RPC_VALIDOS}

        mask = self.df_items['subcategoria'] == apu
        if not mask.any():
            return {rpc: 0.0 for rpc in RPC_VALIDOS}

        df_apu = self.df_items[mask]
        comps = self.componentes_por_apu.get(apu, {})
        override = override or self.asignacion_indirectos.get(apu, {})

        # 1. Directos por RPC
        directos = {rpc: 0.0 for rpc in RPC_VALIDOS}
        for rpc in RPC_VALIDOS:
            directos[rpc] = df_apu[df_apu['tipo_rpc'] == rpc]['parcial_directo'].sum()

        # 2. Indirectos unitarios
        F = comps.get('cargas_sociales', 0.0)
        H = comps.get('herramientas', 0.0)
        L = comps.get('gastos_generales', 0.0)
        M = comps.get('utilidad', 0.0)
        P = comps.get('it', 0.0)
        O = comps.get('iva', 0.0)
        Q = comps.get('precio_unitario_final', 0.0)

        # 3. Asignación de F a MO (proporcional a MO directa)
        mo_directo = {rpc: directos[rpc] for rpc in RPC_MO}
        total_mo = sum(mo_directo.values())
        f_asignado = {rpc: 0.0 for rpc in RPC_MO}
        if total_mo > 0:
            for rpc in RPC_MO:
                f_asignado[rpc] = F * (mo_directo[rpc] / total_mo)
        else:
            # Si no hay MO directa, asignar a la categoría MO por defecto (monu) o la indicada en override
            rpc_default = override.get('cargas_sociales', 'monu')
            if rpc_default in RPC_MO:
                f_asignado[rpc_default] = F

        # 4. Asignación de H, L, M, P, O → BNT (con posibilidad de override por componente)
        # Por simplicidad, asignamos todos a BNT, pero se puede mejorar para permitir override individual
        indirectos_bnt = H + L + M + P + O

        # 5. Total unitario por RPC
        totales = {rpc: directos[rpc] for rpc in RPC_VALIDOS}
        for rpc in RPC_MO:
            totales[rpc] += f_asignado[rpc]
        totales['bnt'] += indirectos_bnt

        # 6. Ajuste por redondeo para que la suma sea exactamente Q
        suma = sum(totales.values())
        if abs(suma - Q) > 0.001 and suma > 0:
            factor = Q / suma
            for rpc in totales:
                totales[rpc] *= factor

        return totales

    def generar_dataframe_for5(self, estructura: Optional[List[Dict]] = None) -> pd.DataFrame:

        filas_for5: List[ItemFOR5] = []

        if estructura is None:
            filas_for5.append(ItemFOR5(
                nivel=0, numero="", descripcion="ITEMS DE OBRAS CIVILES"
            ))

            for i, apu in enumerate(self.subcategorias, 1):
                totales_apu_unit = self.calcular_totales_rpc_por_apu(apu)
                q_apu = sum(totales_apu_unit.values())
                if q_apu <= 0:
                    continue
                unidad_apu = self.unidad_por_apu.get(apu, 'glb')
                cantidad_apu = self.cantidad_por_apu.get(apu, 1.0)

                filas_for5.append(ItemFOR5(
                    nivel=2,
                    numero=f"I-{i}",
                    descripcion=apu.upper(),
                    unidad=unidad_apu,
                    cantidad=cantidad_apu,
                    precio_unitario=q_apu,                     # ← unitario
                    total=q_apu * cantidad_apu,                # ← total
                    bt=totales_apu_unit.get('bt', 0.0) * cantidad_apu,
                    bnt=totales_apu_unit.get('bnt', 0.0) * cantidad_apu,
                    moc=totales_apu_unit.get('moc', 0.0) * cantidad_apu,
                    mos=totales_apu_unit.get('mos', 0.0) * cantidad_apu,
                    monu=totales_apu_unit.get('monu', 0.0) * cantidad_apu,
                    monr=totales_apu_unit.get('monr', 0.0) * cantidad_apu,
                ))

            # Subtotal
            subtotal = ItemFOR5(nivel=3, numero="", descripcion="SUBTOTAL")
            for f in filas_for5:
                if f.nivel == 2:
                    subtotal.bt += f.bt
                    subtotal.bnt += f.bnt
                    subtotal.moc += f.moc
                    subtotal.mos += f.mos
                    subtotal.monu += f.monu
                    subtotal.monr += f.monr
            subtotal.total = (subtotal.bt + subtotal.bnt + subtotal.moc +
                              subtotal.mos + subtotal.monu + subtotal.monr)
            filas_for5.append(subtotal)

            # Total general
            total_gen = ItemFOR5(
                nivel=4, numero="", descripcion="TOTAL (Bs)",
                bt=subtotal.bt, bnt=subtotal.bnt, moc=subtotal.moc,
                mos=subtotal.mos, monu=subtotal.monu, monr=subtotal.monr,
                total=subtotal.total
            )
            filas_for5.append(total_gen)
        else:
            # Modo con estructura jerárquica definida por el usuario (futuro)
            pass

        data = []
        for f in filas_for5:
            data.append({
                'N°': f.numero,
                'DESCRIPCIÓN': f.descripcion,
                'UNIDAD': f.unidad if f.nivel == 2 else '',
                'CANTIDAD': f.cantidad if f.nivel == 2 else '',
                'PRECIO UNITARIO': f.precio_unitario if f.nivel == 2 else '',
                'TOTAL (Bs)': f.total if f.nivel in [2, 3, 4] else '',
                'BIENES TRANS.': f.bt,
                'MATERIAL LOCAL': f.bnt,
                'MDO CALIF.': f.moc,
                'MDO SEMI-CALIF.': f.mos,
                'MDO NO CALIF. URB.': f.monu,
                'MDO NO CALIF. RUR.': f.monr,
            })

        return pd.DataFrame(data)

    def obtener_items_por_clasificar(self) -> pd.DataFrame:
        """Retorna ítems que aún no tienen clasificación RPC"""
        return self.df_items[self.df_items['tipo_rpc'] == 'N']

    def obtener_porcentaje_clasificado(self) -> float:
        """Porcentaje de ítems ya clasificados"""
        if len(self.df_items) == 0:
            return 0.0
        clasificados = len(self.df_items[self.df_items['tipo_rpc'] != 'N'])
        return (clasificados / len(self.df_items)) * 100

    # --- NUEVO: validación de captura ---
    def validar_captura(self, tolerancia: float = 0.01) -> Dict:
        """Valida que la suma de los totales por RPC coincida con el presupuesto publicado."""
        if self.df_items is None or self.df_items.empty:
            return {'ok': False, 'detalle': []}

        publicado_total = 0.0
        calculado_total = 0.0
        detalle = []
        for apu in self.subcategorias:
            q_apu = self.componentes_por_apu.get(apu, {}).get('precio_unitario_final', 0.0)
            cant = self.cantidad_por_apu.get(apu, 1.0)
            publicado_apu = q_apu * cant
            publicado_total += publicado_apu

            t_unit = self.calcular_totales_rpc_por_apu(apu)
            calculado_apu = sum(t_unit.values()) * self.cantidad_por_apu.get(apu, 1.0)
            calculado_total += calculado_apu

            detalle.append({
                'apu': apu,
                'publicado': publicado_apu,
                'capturado': calculado_apu,
                'diferencia': publicado_apu - calculado_apu,
                'pct_dif': (1 - calculado_apu / publicado_apu) * 100 if publicado_apu else 0.0,
            })

        diferencia = publicado_total - calculado_total
        pct = (diferencia / publicado_total) * 100 if publicado_total else 0.0
        return {
            'ok': abs(pct) <= tolerancia * 100,
            'publicado': publicado_total,
            'capturado': calculado_total,
            'diferencia': diferencia,
            'pct_dif': pct,
            'totales_por_rpc': self.calcular_totales_rpc(),
            'detalle': detalle,
        }

    # --- PERSISTENCIA EN BASE DE DATOS ---
    def cargar_desde_db(self, db: ProyectoDB) -> bool:
        df = db.listar_obras()
        if df.empty:
            return False
        self.df_items = df.copy()
        
        # Reconstruir subcategorías en orden de aparición
        seen = set()
        self.subcategorias = []
        for apu_key in self.df_items['subcategoria'].unique():
            if apu_key not in seen:
                seen.add(apu_key)
                self.subcategorias.append(apu_key)
        
        # Reconstruir componentes APU desde DB
        df_apu = db.listar_apu_componentes()
        self.componentes_por_apu = {}
        self.precio_unitario_por_apu = {}
        self.unidad_por_apu = {}
        self.cantidad_por_apu = {}
        self.asignacion_indirectos = {}
        
        if not df_apu.empty:
            import json
            for _, row in df_apu.iterrows():
                apu = row['apu']
                self.componentes_por_apu[apu] = {
                    'cargas_sociales': float(row.get('cargas_sociales', 0) or 0),
                    'herramientas': float(row.get('herramientas', 0) or 0),
                    'gastos_generales': float(row.get('gastos_generales', 0) or 0),
                    'utilidad': float(row.get('utilidad', 0) or 0),
                    'it': float(row.get('it', 0) or 0),
                    'iva': float(row.get('iva', 0) or 0),
                    'precio_unitario_final': float(row.get('precio_unitario_final', 0) or 0),
                }
                self.precio_unitario_por_apu[apu] = float(row.get('precio_unitario_final', 0) or 0)
                self.cantidad_por_apu[apu] = float(row.get('cantidad', 1.0) or 1.0)
                
                # Recuperar asignaciones de indirectos si existen
                asig = row.get('asignacion_json', '{}')
                try:
                    self.asignacion_indirectos[apu] = json.loads(asig) if asig else {}
                except:
                    self.asignacion_indirectos[apu] = {}
        
        # Reconstruir unidad por APU desde items (primera ocurrencia)
        for apu in self.subcategorias:
            df_apu_items = self.df_items[self.df_items['subcategoria'] == apu]
            if not df_apu_items.empty:
                self.unidad_por_apu[apu] = str(df_apu_items.iloc[0].get('unidad', 'glb'))
                if apu not in self.cantidad_por_apu:
                    self.cantidad_por_apu[apu] = 1.0
        
        return True

    def guardar_en_db(self, db: ProyectoDB) -> bool:
        """Guarda obras en la base de datos."""
        if self.df_items is None or self.df_items.empty:
            return False
        
        # 1. Guardar ítems
        df_guardar = self.df_items.copy()
        cols_requeridas = ['codigo_item', 'descripcion', 'descripcion_norm', 'unidad',
                           'cantidad', 'precio_unitario', 'parcial_directo', 'tipo_hoja',
                           'subcategoria', 'tipo_rpc', 'fecha_clasificacion',
                           'precio_unitario_final', 'parcial_real', 'factor_apu',
                           'indirectos_asignados', 'parcial']
        for col in cols_requeridas:
            if col not in df_guardar.columns:
                df_guardar[col] = 0.0 if col in ['cantidad','precio_unitario','parcial_directo',
                                                'precio_unitario_final','parcial_real','factor_apu',
                                                'indirectos_asignados','parcial'] else ''
        if 'id' in df_guardar.columns:
            df_guardar = df_guardar.drop(columns=['id'])
        
        ok = db.guardar_obras(df_guardar[cols_requeridas])
        if not ok:
            return False
        
        # 2. Guardar componentes + cantidad + asignaciones de indirectos
        if self.componentes_por_apu:
            import json
            rows = []
            for apu, comps in self.componentes_por_apu.items():
                row = {
                    'apu': apu,
                    'cantidad': self.cantidad_por_apu.get(apu, 1.0),
                    'asignacion_json': json.dumps(self.asignacion_indirectos.get(apu, {}))
                }
                row.update(comps)
                rows.append(row)
            df_comp = pd.DataFrame(rows)
            #db.guardar_apu_componentes(df_comp)
            if not db.guardar_apu_componentes(df_comp):
                return False   # <-- agregar esta línea para propagar el error  

        return True

class GestorServicios:
    """Gestiona ATI, Supervisión, OM y otros servicios (montos globales)"""

    def __init__(self, ruta_excel: str = "proyecto_activo.xlsx"):
        self.ruta_excel = ruta_excel
        self.servicios = {}
        # Inicializar estructura base siempre
        for cat in ['ati', 'supervision', 'om', 'ambiental']:
            self.servicios[cat] = ServicioGlobal(
                categoria=cat, descripcion=CATEGORIAS[cat]['nombre']
            )
        self.cargar()

    def cargar(self):
        """Carga servicios guardados previamente desde Excel (fallback)."""
        if not os.path.exists(self.ruta_excel):
            return
        try:
            wb = load_workbook(self.ruta_excel, data_only=True)
            if "Servicios_RPC" in wb.sheetnames:
                df = pd.read_excel(self.ruta_excel, sheet_name="Servicios_RPC")
                for _, row in df.iterrows():
                    cat = row['categoria']
                    if cat in self.servicios:
                        self.servicios[cat] = ServicioGlobal(
                            categoria=cat,
                            descripcion=row['descripcion'],
                            monto_bt=float(row.get('bt', 0)),
                            monto_bnt=float(row.get('bnt', 0)),
                            monto_moc=float(row.get('moc', 0)),
                            monto_mos=float(row.get('mos', 0)),
                            monto_monu=float(row.get('monu', 0)),
                            monto_monr=float(row.get('monr', 0)),
                            detalle_adjunto=str(row.get('detalle_adjunto', ''))
                        )
        except Exception as e:
            pass  # Mantener estructura base inicializada

    def cargar_desde_db(self, db: ProyectoDB):
        """Carga servicios desde la base de datos (sobrescribe Excel si hay datos)."""
        df = db.listar_servicios()
        if df.empty:
            return
        for _, row in df.iterrows():
            cat = row['categoria']
            if cat in self.servicios:
                self.servicios[cat] = ServicioGlobal(
                    categoria=cat,
                    descripcion=row['descripcion'],
                    monto_bt=float(row.get('bt', 0)),
                    monto_bnt=float(row.get('bnt', 0)),
                    monto_moc=float(row.get('moc', 0)),
                    monto_mos=float(row.get('mos', 0)),
                    monto_monu=float(row.get('monu', 0)),
                    monto_monr=float(row.get('monr', 0)),
                    detalle_adjunto=str(row.get('detalle_adjunto', ''))
                )

    def guardar_en_db(self, db: ProyectoDB) -> bool:
        """Guarda servicios en la base de datos."""
        rows = []
        for cat, serv in self.servicios.items():
            rows.append({
                'categoria': cat,
                'descripcion': serv.descripcion,
                'bt': serv.monto_bt,
                'bnt': serv.monto_bnt,
                'moc': serv.monto_moc,
                'mos': serv.monto_mos,
                'monu': serv.monto_monu,
                'monr': serv.monto_monr,
                'detalle_adjunto': serv.detalle_adjunto
            })
        df = pd.DataFrame(rows)
        return db.guardar_servicios(df)

    def actualizar_servicio(self, categoria: str, **kwargs):
        """Actualiza montos de un servicio"""
        if categoria in self.servicios:
            for key, value in kwargs.items():
                if hasattr(self.servicios[categoria], key):
                    setattr(self.servicios[categoria], key, value)

    def calcular_totales_rpc(self) -> Dict[str, float]:
        """Suma totales por RPC de todos los servicios"""
        totales = {rpc: 0.0 for rpc in TIPOS_RPC.keys() if rpc != 'N'}
        for serv in self.servicios.values():
            totales['bt'] += serv.monto_bt
            totales['bnt'] += serv.monto_bnt
            totales['moc'] += serv.monto_moc
            totales['mos'] += serv.monto_mos
            totales['monu'] += serv.monto_monu
            totales['monr'] += serv.monto_monr
        return totales

class ConsolidadorInversion:
    def __init__(self, obras: GestorObrasCiviles, servicios: GestorServicios):
        self.obras = obras
        self.servicios = servicios

    def obtener_totales_consolidados(self) -> Dict[str, float]:
        """Suma Obras Civiles + Servicios por RPC"""
        tot_obras = self.obras.calcular_totales_rpc()
        tot_serv = self.servicios.calcular_totales_rpc()
        consolidado = {}
        for rpc in tot_obras.keys():
            consolidado[rpc] = tot_obras.get(rpc, 0) + tot_serv.get(rpc, 0)
        return consolidado

# ============================================================
# INTERFAZ DE USUARIO CON SESSION_STATE
# ============================================================
def inicializar_session_state():
    """Inicializa los gestores en session_state, detectando cambios de proyecto.
    Carga desde proyecto.db primero; si no hay datos, deja estructura vacía."""
    ruta_actual = RUTA_EXCEL
    db = ProyectoDB(RUTA_DB)

    if 'proyecto_cargado_ruta' not in st.session_state:
        st.session_state.proyecto_cargado_ruta = ruta_actual

    if st.session_state.proyecto_cargado_ruta != ruta_actual:
        keys_to_clear = [
            'obras', 'servicios', 'datos_guardados',
            'proyecto_cargado_ruta', 'editor_obras',
            'filtro_subcat', 'filtro_tipo',
            'vista_unica_editor', 'mostrar_repetidos',
            'sobrescribir_grupo', 'editor_componentes_indirectos',
            'estructura_for5',
        ]
        for key in keys_to_clear:
            if key in st.session_state:
                del st.session_state[key]
        st.session_state.proyecto_cargado_ruta = ruta_actual
    if 'obras' not in st.session_state:
        obras = GestorObrasCiviles()
        if os.path.exists(RUTA_DB):
            try:
                if obras.cargar_desde_db(db):
                    pass  # Datos cargados desde DB
            except Exception as e:
                st.warning(f"No se pudieron cargar obras desde DB: {e}")
        st.session_state.obras = obras
    if 'servicios' not in st.session_state:
        servicios = GestorServicios(ruta_excel=RUTA_EXCEL)
        if os.path.exists(RUTA_DB):
            try:
                servicios.cargar_desde_db(db)
            except Exception as e:
                st.warning(f"No se pudieron cargar servicios desde DB: {e}")
        st.session_state.servicios = servicios
    if 'datos_guardados' not in st.session_state:
        st.session_state.datos_guardados = False

# ============================================================
# UI: NUEVA CLASIFICACIÓN INTELIGENTE
# ============================================================
def _render_clasificacion_inteligente(self, subcat_seleccionada: str, tipo_hoja: List[str]):

    df_filtrado = self.obras.df_items.copy()
    if subcat_seleccionada != 'TODAS':
        df_filtrado = df_filtrado[df_filtrado['subcategoria'] == subcat_seleccionada]
    if tipo_hoja:
        df_filtrado = df_filtrado[df_filtrado['tipo_hoja'].isin(tipo_hoja)]

    if df_filtrado.empty:
        st.info("No hay ítems con los filtros seleccionados.")
        return

    # Calcular totales en el scope filtrado
    total_directo = df_filtrado['parcial_directo'].sum()
    total_real = df_filtrado['parcial_real'].sum()
    perdidos = total_real - total_directo
    pct_perdido = (perdidos / total_real * 100) if total_real else 0.0

    c1, c2, c3 = st.columns(3)
    c1.metric("Total DIRECTO (A+B+C+F+H)", f"{total_directo:,.2f} Bs")
    c2.metric("Total REAL (con Q publicado)", f"{total_real:,.2f} Bs")
    delta_color = "off" if pct_perdido < 1 else "inverse"
    c3.metric("Indirectos capturados", f"{perdidos:,.2f} Bs",
              delta=f"{pct_perdido:.1f}% del total",
              delta_color=delta_color)

    # Validación contra Q publicado
    if subcat_seleccionada != 'TODAS':
        q_pub = self.obras.precio_unitario_por_apu.get(subcat_seleccionada, 0.0)
        real_filtrado = df_filtrado['parcial_real'].sum()
        if q_pub > 0:
            ratio = real_filtrado / q_pub if q_pub else 0
            st.caption(
                f"🧪 Validación APU: parcial_real={real_filtrado:,.2f} / "
                f"Q publicado={q_pub:,.2f} → ratio={ratio:.4f}"
            )

    st.divider()
    st.subheader("🧠 Clasificación inteligente por insumo único")

    df_unicos = self.obras.vista_insumos_unicos(subcat_seleccionada)
    if df_unicos.empty:
        st.info("No hay insumos únicos para mostrar.")
        return

    # Filtrar la vista única también por tipo_hoja
    if tipo_hoja:
        df_unicos = df_unicos[df_unicos['tipo_hoja'].isin(tipo_hoja)]

    st.caption(
        f"{len(df_unicos)} insumos únicos (de {len(df_filtrado)} ítems totales en el scope). "
        f"Cada fila representa un insumo; clasificar UNA fila propaga la asignación "
        f"a todos los ítems con la misma descripción."
    )

    # Configuración de la tabla editable
    opciones_rpc = list(TIPOS_RPC.keys())

    columnas_mostrar = [
        'descripcion_canon', 'unidad', 'tipo_hoja', 'n_items',
        'cantidad_total', 'parcial_directo_total', 'parcial_real_total',
        'tipo_rpc', 'clasificacion_consistente',
    ]

    df_editor = df_unicos[columnas_mostrar].copy()
    edited = st.data_editor(
        df_editor,
        column_config={
            "descripcion_canon": st.column_config.TextColumn(
                "Insumo", disabled=True, width="large"
            ),
            "unidad": st.column_config.TextColumn("Und.", disabled=True),
            "tipo_hoja": st.column_config.TextColumn("Sección", disabled=True),
            "n_items": st.column_config.NumberColumn(
                "# Ítems", disabled=True,
                help="Cantidad de líneas con esta descripción en el scope"
            ),
            "cantidad_total": st.column_config.NumberColumn(
                "Cant. total", disabled=True, format="%.2f"
            ),
            "parcial_directo_total": st.column_config.NumberColumn(
                "Parcial directo (Bs)", disabled=True, format="%.2f"
            ),
            "parcial_real_total": st.column_config.NumberColumn(
                "Parcial REAL (Bs)", disabled=True, format="%.2f",
                help="Incluye gastos generales, utilidad, IT, etc. (Q publicado)"
            ),
            "tipo_rpc": st.column_config.SelectboxColumn(
                "Clasificación RPC",
                options=opciones_rpc,
                help="Asignar al GRUPO: se propaga a todos los ítems repetidos"
            ),
            "clasificacion_consistente": st.column_config.CheckboxColumn(
                "Consistente", disabled=True,
                help="True si todos los ítems del grupo tienen el mismo RPC"
            ),
        },
        use_container_width=True,
        height=420,
        key="vista_unica_editor",
    )

    # Toggle para sobrescribir N
    col_opc1, col_opc2 = st.columns([1, 3])
    with col_opc1:
        sobrescribir = st.checkbox(
            "Sobrescribir clasificación existente",
            value=False,
            key="sobrescribir_grupo",
            help="Si está apagado, solo cambia los ítems en 'N' (respeta asignaciones manuales previas)"
        )
    with col_opc2:
        st.info(
            "💡 **Tip:** Apaga el switch para no perder clasificaciones manuales ya hechas. "
            "Enciéndelo solo si quieres reasignar masivamente (ej. corregir un grupo)."
        )

    col_btn1, col_btn2 = st.columns([1, 1])
    with col_btn1:
        aplicar_grupo = st.button(
            "💾 Aplicar clasificación a grupos", type="primary",
            key="btn_aplicar_grupo"
        )
    with col_btn2:
        mostrar_detalle = st.checkbox(
            "Mostrar repetidos expandidos",
            value=False,
            key="mostrar_repetidos"
        )

    if aplicar_grupo:
        n_items_total = 0
        n_grupos = 0
        alertas_bt = []
        for idx in edited.index:
            valor_original = df_unicos.loc[idx, 'tipo_rpc']
            valor_nuevo = edited.loc[idx, 'tipo_rpc']
            if valor_original != valor_nuevo:
                desc_norm = df_unicos.loc[idx, 'descripcion_norm']
                n_act, _ = self.obras.aplicar_clasificacion_por_grupo(
                    desc_norm, valor_nuevo, sobrescribir=sobrescribir
                )
                n_items_total += n_act
                n_grupos += 1
                if valor_nuevo == 'bt':
                    alertas_bt.append(
                        f"{df_unicos.loc[idx, 'descripcion_canon'][:40]}... "
                        f"({n_act} ítems)"
                    )

        if n_grupos > 0:
            st.success(
                f"✅ Clasificación aplicada: {n_grupos} grupo(s), "
                f"{n_items_total} ítem(s) actualizado(s)."
            )
            if alertas_bt and sobrescribir:
                with st.expander(
                    f"⚠️ {len(alertas_bt)} grupo(s) asignados a BT"
                ):
                    for a in alertas_bt[:20]:
                        st.text(f"• {a}")
                    if len(alertas_bt) > 20:
                        st.text(f"... y {len(alertas_bt) - 20} más")
            st.rerun()
        else:
            st.info("ℹ️ No se detectaron cambios en la tabla.")

    # Auto-completar N → BT como red de seguridad
    st.divider()
    col_auto1, col_auto2 = st.columns([1, 3])
    with col_auto1:
        auto_bt = st.button(
            "🤖 Auto-clasificar N → BT",
            type="secondary",
            key="btn_auto_bt"
        )
    with col_auto2:
        st.caption(
            "Marca como BT todos los ítems aún en 'N'. Útil como red de seguridad, "
            "pero revisa la lógica: muchos ítems 'N' deberían ser MO, no BT."
        )

    if auto_bt:
        mask_n = self.obras.df_items['tipo_rpc'] == 'N'
        n = int(mask_n.sum())
        if n == 0:
            st.info("No hay ítems en 'N'.")
        else:
            self.obras.df_items.loc[mask_n, 'tipo_rpc'] = 'bt'
            self.obras.df_items.loc[mask_n, 'fecha_clasificacion'] = (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " (auto-N→BT)"
            )
            st.warning(
                f"⚠️ {n} ítems auto-clasificados como BT. "
                f"Revisa si alguno debería ser MO u otra categoría."
            )
            st.rerun()

    # Vista expandida de repetidos (opcional)
    if mostrar_detalle:
        st.divider()
        st.subheader("📋 Detalle de ítems (vista expandida)")
        st.caption(
            "Cada fila es un ítem individual. Clasifica aquí si necesitas granularidad "
            "fina dentro de un grupo (ej. un caso especial)."
        )
        cols_detalle = [
            'codigo_item', 'descripcion', 'unidad', 'cantidad',
            'precio_unitario', 'parcial_directo', 'precio_unitario_final',
            'parcial_real', 'tipo_rpc', 'subcategoria'
        ]
        df_detalle = df_filtrado[cols_detalle].copy()
        edited_detalle = st.data_editor(
            df_detalle,
            column_config={
                "precio_unitario": st.column_config.NumberColumn(
                    "P.Unit. insumo", format="%.2f", disabled=True
                ),
                "parcial_directo": st.column_config.NumberColumn(
                    "Parcial directo", format="%.2f", disabled=True
                ),
                "precio_unitario_final": st.column_config.NumberColumn(
                    "Q (publicado)", format="%.2f", disabled=True
                ),
                "parcial_real": st.column_config.NumberColumn(
                    "Parcial REAL (cant×Q)", format="%.2f", disabled=True
                ),
                "tipo_rpc": st.column_config.SelectboxColumn(
                    "RPC", options=opciones_rpc
                ),
            },
            use_container_width=True,
            height=350,
            key="editor_detalle",
        )
        if st.button("💾 Aplicar clasificación detalle", key="btn_detalle"):
            cambios = 0
            for idx in edited_detalle.index:
                if idx in self.obras.df_items.index:
                    v_orig = self.obras.df_items.loc[idx, 'tipo_rpc']
                    v_new = edited_detalle.loc[idx, 'tipo_rpc']
                    if v_orig != v_new:
                        self.obras.df_items.loc[idx, 'tipo_rpc'] = v_new
                        self.obras.df_items.loc[idx, 'fecha_clasificacion'] = (
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " (detalle)"
                        )
                        cambios += 1
            if cambios:
                st.success(f"✅ {cambios} ítems actualizados a nivel de detalle.")
                st.rerun()
            else:
                st.info("Sin cambios.")

def _inyectar_render_inteligente():
    UIInversionCompleta._render_clasificacion_inteligente = _render_clasificacion_inteligente

# ============================================================
# CLASE UI PRINCIPAL (módulo de inversión)
# ============================================================
class UIInversionCompleta:
    """Interfaz unificada para Obras (detalle) + Servicios (global)"""

    def __init__(self):
        self.obras = st.session_state.obras
        self.servicios = st.session_state.servicios
        self.consolidador = ConsolidadorInversion(self.obras, self.servicios)
    def render(self):
        st.title("🏗️ Módulo de Inversión - Clasificación RPC")
        st.markdown("""
        **Metodología:**
        - **Obras Civiles**: Clasificación por Analisis de Precios Unitarios
        - **Servicios (ATI, Supervisión, OM)**: Clasificación global por montos agregados
        """)

        col_title, col_reload = st.columns([5, 1])
        with col_reload:
            if st.button("🔄 Actualizar", help="Limpia caché y recarga datos del proyecto activo",
                        use_container_width=True):
                for key in ['obras', 'servicios', 'datos_guardados', 'editor_obras',
                            'editor_detalle', 'filtro_subcat', 'filtro_tipo',
                            'vista_unica_editor', 'mostrar_repetidos',
                            'sobrescribir_grupo', 'editor_componentes_indirectos',
                            'estructura_for5']:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()

        pct = self.obras.obtener_porcentaje_clasificado()
        st.progress(pct/100, text=f"Progreso clasificación Obras: {pct:.1f}%")

        # Validación mejorada (alerta si > 1%)
        if not self.obras.df_items.empty and self.obras.precio_unitario_por_apu:
            v = self.obras.validar_captura(tolerancia=0.01)
            if v['publicado'] > 0:
                if not v['ok']:
                    st.error(
                        f"🚨 **Validación APU:** Diferencia significativa. "
                        f"Capturado={v['capturado']:,.2f} Bs vs "
                        f"Publicado={v['publicado']:,.2f} Bs "
                        f"(dif: {v['pct_dif']:+.2f}%)."
                    )
                    with st.expander("Ver detalle por APU"):
                        st.dataframe(
                            pd.DataFrame(v['detalle']).style.format({
                                'capturado': '{:,.2f}',
                                'publicado': '{:,.2f}',
                                'diferencia': '{:,.2f}',
                                'pct_dif': '{:.2f}%'
                            }),
                            use_container_width=True
                        )
                else:
                    st.success(
                        f"✅ Validación APU OK: diferencia {v['pct_dif']:+.2f}% (≤ 1% tolerancia)"
                    )

        tabs = st.tabs([
            "📋 Clasificación Obras",
            "🔧 Componentes Indirectos",
            "📄 Reporte FOR5",
            "👥 Servicios (ATI/Sup/OM)",
            "📊 Resumen Consolidado",
            "💾 Guardar Proyecto"
        ])

        with tabs[0]:
            self._render_clasificacion_obras()
        with tabs[1]:
            self._render_componentes_indirectos()
        with tabs[2]:
            self._render_reporte_for5()
        with tabs[3]:
            self._render_servicios_global()
        with tabs[4]:
            self._render_resumen_consolidado()
        with tabs[5]:
            self._render_guardar()

    def _render_clasificacion_obras(self):
        st.header("Clasificación de Obras Civiles")

        col_file, col_info = st.columns([2, 1])
        with col_file:
            archivo_desglose = st.file_uploader(
                "📁 Cargar archivo Analisis de Precios Unitarios (Excel)",
                type=['xlsx', 'xls'],
                help="Seleccione el archivo Excel con bloques de APU con formato Prescom o similar"
            )

        with col_info:
            if self.obras.df_items.empty:
                st.info("💡 Cargue un archivo para comenzar la clasificación")
            else:
                st.success(f"✅ {len(self.obras.df_items)} ítems cargados")
                n_apus = len(self.obras.subcategorias)
                st.caption(f"{n_apus} APUs · {len(self.obras.componentes_por_apu)} con Q capturado")

        if archivo_desglose is not None:
            temp_ruta = os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "temp_desglose.xlsx"
            )
            with open(temp_ruta, "wb") as f:
                f.write(archivo_desglose.getvalue())

            if self.obras.ruta_desglose != temp_ruta or self.obras.df_items.empty:
                with st.spinner("Cargando desglose..."):
                    ok = self.obras.cargar_desde_archivo(temp_ruta)
                    if ok:
                        st.session_state.obras = self.obras
                        st.success(f"✅ Archivo cargado: {archivo_desglose.name}")
                        st.rerun()
                    else:
                        st.error("❌ No se pudo cargar el archivo. Verifique el formato.")
                return

        if self.obras.df_items.empty:
            st.warning("⚠️ No hay datos cargados. Por favor, cargue un archivo de desglose de inversión.")
            return

        # Filtros
        col1, col2 = st.columns([1, 3])
        with col1:
            _conteo_por_apu = (
                self.obras.df_items['subcategoria'].value_counts().to_dict()
            )
            _subcat_labels = {
                apu: f"{i+1:02d}. {apu} "
                     f"({_conteo_por_apu[apu]} ítem{'s' if _conteo_por_apu[apu] != 1 else ''})"
                for i, apu in enumerate(self.obras.subcategorias)
            }
            subcat_seleccionada = st.selectbox(
                "Filtrar por APU",
                options=['TODAS'] + self.obras.subcategorias,
                format_func=lambda x: (
                    '📋 Todas las APUs' if x == 'TODAS' else _subcat_labels.get(x, x)
                ),
                key="filtro_subcat"
            )

        with col2:
            tipo_hoja = st.multiselect(
                "Tipo de insumo",
                ['materiales', 'mano_obra', 'equipo_maquinaria'],
                default=['materiales', 'mano_obra', 'equipo_maquinaria'],
                key="filtro_tipo"
            )

        st.divider()

        # LLAMADA A LA UI INTELIGENTE
        self._render_clasificacion_inteligente(subcat_seleccionada, tipo_hoja)

        # Resumen rápido por RPC (incluye indirectos capturados)
        if not self.obras.df_items.empty:
            st.divider()
            cols = st.columns(7)
            for i, (rpc, info) in enumerate(TIPOS_RPC.items()):
                if rpc == 'N':
                    continue
                col_suma = (
                    'parcial_real' if 'parcial_real' in self.obras.df_items.columns
                    and self.obras.df_items['parcial_real'].sum() > 0
                    else 'parcial'
                )
                total_rpc = self.obras.df_items[
                    self.obras.df_items['tipo_rpc'] == rpc
                ][col_suma].sum()
                cantidad_rpc = len(self.obras.df_items[self.obras.df_items['tipo_rpc'] == rpc])
                with cols[i]:
                    st.metric(
                        label=f"{rpc.upper()}",
                        value=f"{total_rpc:,.0f} Bs",
                        delta=f"{cantidad_rpc} ítems"
                    )
            st.caption("💡 Los montos mostrados arriba son **valores unitarios** (por 1 unidad de obra). "
                       "Los totales del presupuesto se calculan en el Resumen Consolidado y el FOR5.")
                                   
            pendientes_n = len(self.obras.df_items[self.obras.df_items['tipo_rpc'] == 'N'])
            if pendientes_n > 0:
                st.warning(
                    f"⚠️ Quedan {pendientes_n} ítems sin clasificar (N). "
                    f"Use 'Auto-clasificar N → BT' solo como red de seguridad."
                )

    def _render_servicios_global(self):
        st.header("Clasificación Global de Servicios")
        st.markdown("Ingrese los montos totales por categoría RPC (puede adjuntar detalle si lo desea)")

        for cat_key, cat_info in CATEGORIAS.items():
            if cat_key == 'obras':
                continue

            with st.expander(f"{cat_info['icono']} {cat_info['nombre']}", expanded=True):
                serv = self.servicios.servicios.get(
                    cat_key, ServicioGlobal(cat_key, cat_info['nombre'])
                )

                col1, col2, col3 = st.columns([2, 3, 2])

                with col1:
                    st.markdown("**Descripción:**")
                    desc = st.text_area(
                        f"desc_{cat_key}",
                        value=serv.descripcion,
                        label_visibility="collapsed",
                        height=100
                    )

                with col2:
                    st.markdown("**Desglose RPC:**")
                    cols_rpc = st.columns(3)
                    montos = {}

                    with cols_rpc[0]:
                        montos['bt'] = st.number_input("BT (Bs)", value=serv.monto_bt, key=f"{cat_key}_bt")
                        montos['bnt'] = st.number_input("BNT (Bs)", value=serv.monto_bnt, key=f"{cat_key}_bnt")

                    with cols_rpc[1]:
                        montos['moc'] = st.number_input("MOC (Bs)", value=serv.monto_moc, key=f"{cat_key}_moc")
                        montos['mos'] = st.number_input("MOS (Bs)", value=serv.monto_mos, key=f"{cat_key}_mos")

                    with cols_rpc[2]:
                        montos['monu'] = st.number_input("MONU (Bs)", value=serv.monto_monu, key=f"{cat_key}_monu")
                        montos['monr'] = st.number_input("MONR (Bs)", value=serv.monto_monr, key=f"{cat_key}_monr")

                with col3:
                    st.markdown("**Total y Detalle:**")
                    total_cat = sum(montos.values())
                    st.metric("Total Categoría", f"{total_cat:,.2f} Bs")

                    detalle = st.text_area(
                        "Detalle adjunto (opcional)",
                        value=serv.detalle_adjunto,
                        placeholder="Puede pegar aquí el detalle de ítems o referencia a archivo...",
                        key=f"det_{cat_key}",
                        height=80
                    )

                    if st.button(f"💾 Guardar {cat_info['nombre']}", key=f"save_{cat_key}"):
                        self.servicios.actualizar_servicio(
                            cat_key,
                            descripcion=desc,
                            monto_bt=montos['bt'],
                            monto_bnt=montos['bnt'],
                            monto_moc=montos['moc'],
                            monto_mos=montos['mos'],
                            monto_monu=montos['monu'],
                            monto_monr=montos['monr'],
                            detalle_adjunto=detalle
                        )
                        st.session_state.servicios = self.servicios
                        st.success(f"✅ {cat_info['nombre']} guardado en memoria.")

    def _render_resumen_consolidado(self):
        st.header("Resumen Consolidado de Inversión")

        tot_obras = self.obras.calcular_totales_rpc()
        tot_serv = self.servicios.calcular_totales_rpc()
        consolidado = self.consolidador.obtener_totales_consolidados()

        col_debug1, col_debug2 = st.columns(2)
        with col_debug1:
            st.caption(f"Obras: {len(self.obras.df_items)} ítems cargados")
        with col_debug2:
            st.caption(f"Servicios: {len([s for s in self.servicios.servicios.values() if s.total > 0])} categorías con datos")

        data_comp = []
        for rpc in ['bt', 'bnt', 'moc', 'mos', 'monu', 'monr']:
            data_comp.append({
                'RPC': f"{rpc.upper()} - {TIPOS_RPC[rpc]['nombre']}",
                'Obras Civiles': tot_obras.get(rpc, 0),
                'Servicios': tot_serv.get(rpc, 0),
                'TOTAL': consolidado.get(rpc, 0),
                '% del Total': 0.0
            })

        df_comp = pd.DataFrame(data_comp)
        total_general = df_comp['TOTAL'].sum()

        if total_general > 0:
            df_comp['% del Total'] = (df_comp['TOTAL'] / total_general * 100).round(2)

        st.dataframe(
            df_comp.style.format({
                'Obras Civiles': '{:,.2f}',
                'Servicios': '{:,.2f}',
                'TOTAL': '{:,.2f}',
                '% del Total': '{:.1f}%'
            }),
            use_container_width=True
        )
        '''
        st.subheader("Distribución de la Inversión")
        chart_data = df_comp.set_index('RPC')[['Obras Civiles', 'Servicios']]
        st.bar_chart(chart_data)
        '''
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Obras", f"{sum(tot_obras.values()):,.0f} Bs")
        col2.metric("Total Servicios", f"{sum(tot_serv.values()):,.0f} Bs")
        col3.metric("INVERSIÓN TOTAL", f"{total_general:,.0f} Bs")
        col4.metric("Ítems Obras Clasificados", f"{self.obras.obtener_porcentaje_clasificado():.1f}%")

        # Detalle de componentes indirectos
        if self.obras.componentes_por_apu:
            with st.expander("🧩 Componentes indirectos capturados por APU"):
                rows = []
                for apu, comps in self.obras.componentes_por_apu.items():
                    rows.append({
                        'APU': apu,
                        'Cargas Soc. (F)': comps.get('cargas_sociales', 0),
                        'Herramientas (H)': comps.get('herramientas', 0),
                        'Gast. Grales. (L)': comps.get('gastos_generales', 0),
                        'Utilidad (M)': comps.get('utilidad', 0),
                        'IT (P)': comps.get('it', 0),
                        'IVA (O)': comps.get('iva', 0),
                        'TOTAL Q': comps.get('precio_unitario_final', 0),
                    })
                df_indirectos = pd.DataFrame(rows)
                st.dataframe(
                    df_indirectos.style.format({
                        col: '{:,.2f}' for col in df_indirectos.columns if col != 'APU'
                    }),
                    use_container_width=True
                )

    def _render_guardar(self):
        st.header("Guardar en Proyecto Activo")

        st.info("Esta acción actualizará el archivo proyecto_activo.xlsx con:")
        st.markdown("""
        - Hoja **Obras_Detalle**: Ítems clasificados de obras civiles
        - Hoja **Inversion_Resumen**: Totales consolidados por RPC
        - Hoja **APU_Componentes**: Componentes indirectos capturados (NUEVO)
        """)

        col_save, col_clear = st.columns([3, 1])
        with col_save:
            if st.button("💾 Guardar Todo en Proyecto", type="primary", use_container_width=True):
                ok = self._guardar_en_data_manager()
                if ok:
                    st.session_state.datos_guardados = True
                    st.success("✅ Proyecto guardado exitosamente en proyecto_activo.xlsx")
                else:
                    st.error("❌ Error al guardar")

        with col_clear:
            st.warning("⚠️ Esta acción eliminará la información del proyecto actual.")
            if st.button("🗑️ Limpiar Proyecto (Nuevo)", type="secondary"):
                self._limpiar_proyecto()

    def _guardar_en_data_manager(self) -> bool:
        """Guarda en DB (persistencia) y en Excel (puente para evaluación)."""
        try:
            db = ProyectoDB(RUTA_DB)

            # 1. PERSISTIR EN SQLITE (proyecto.db)
            if not self.obras.df_items.empty:
                if not self.obras.guardar_en_db(db):
                    st.error("❌ Error guardando obras en proyecto.db")
                    return False

            if not self.servicios.guardar_en_db(db):
                st.error("❌ Error guardando servicios en proyecto.db")
            '''
            if self.obras.componentes_por_apu:
                rows = []
                for apu, comps in self.obras.componentes_por_apu.items():
                    row = {'apu': apu}
                    row.update(comps)
                    rows.append(row)
                df_comp = pd.DataFrame(rows)
                if not db.guardar_apu_componentes(df_comp):
                    st.warning("⚠️ Error guardando APU_Componentes en DB")
            '''
            # 2. PUENTE EXCEL (para que app_evaluacion_mon.py pueda leerlo)
            dm = DataManager(self.servicios.ruta_excel)

            if not self.obras.df_items.empty:
                ok, errores = dm.guardar_hoja("Obras_Detalle", self.obras.df_items, validar=False)
                if not ok:
                    st.error(f"Error guardando Obras_Detalle en Excel: {errores}")
                    return False

            if self.obras.componentes_por_apu:
                rows = []
                for apu, comps in self.obras.componentes_por_apu.items():
                    rows.append({'apu': apu, **comps})
                df_comp = pd.DataFrame(rows)
                ok, errores = dm.guardar_hoja("APU_Componentes", df_comp, validar=False)
                if not ok:
                    st.error(f"Error guardando APU_Componentes en Excel: {errores}")

            self._actualizar_resumen_inversion_data_manager(dm)
            return True
        except Exception as e:
            st.error(f"Error guardando: {e}")
            return False

    def _actualizar_resumen_inversion_data_manager(self, dm: DataManager):
        """Actualiza hoja de resumen consolidado usando DataManager."""
        consolidado = self.consolidador.obtener_totales_consolidados()

        data = []
        tot_obras = self.obras.calcular_totales_rpc()
        data.append({
            'Categoría': 'OBRAS CIVILES',
            'BT': tot_obras.get('bt', 0),
            'BNT': tot_obras.get('bnt', 0),
            'MOC': tot_obras.get('moc', 0),
            'MOS': tot_obras.get('mos', 0),
            'MONU': tot_obras.get('monu', 0),
            'MONR': tot_obras.get('monr', 0),
            'TOTAL': sum(tot_obras.values())
        })

        for cat, serv in self.servicios.servicios.items():
            data.append({
                'Categoría': serv.descripcion,
                'BT': serv.monto_bt,
                'BNT': serv.monto_bnt,
                'MOC': serv.monto_moc,
                'MOS': serv.monto_mos,
                'MONU': serv.monto_monu,
                'MONR': serv.monto_monr,
                'TOTAL': serv.total
            })

        data.append({
            'Categoría': 'TOTAL INVERSIÓN',
            'BT': consolidado.get('bt', 0),
            'BNT': consolidado.get('bnt', 0),
            'MOC': consolidado.get('moc', 0),
            'MOS': consolidado.get('mos', 0),
            'MONU': consolidado.get('monu', 0),
            'MONR': consolidado.get('monr', 0),
            'TOTAL': sum(consolidado.values())
        })

        df_resumen = pd.DataFrame(data)
        ok, errores = dm.guardar_hoja("Inversion_Resumen", df_resumen, validar=False)
        if not ok:
            st.error(f"Error guardando Inversion_Resumen: {errores}")

    def _limpiar_proyecto(self):
        try:
            # Limpiar base de datos
            db = ProyectoDB(RUTA_DB)
            db.limpiar_datos_proyecto()
            # Limpiar Excel puente
            dm = DataManager(self.servicios.ruta_excel)
            dm.crear_proyecto_nuevo()
            st.session_state.obras = GestorObrasCiviles()
            st.session_state.servicios = GestorServicios(ruta_excel=RUTA_EXCEL)
            st.session_state.datos_guardados = False
            st.success("✅ Proyecto limpiado en DB y Excel.")
            st.rerun()
        except Exception as e:
            st.error(f"❌ Error al limpiar proyecto: {e}")
            
    def _render_componentes_indirectos(self):
        """Pestaña de configuración de overrides para componentes indirectos."""
        st.header("🔧 Componentes Indirectos por APU")
        st.markdown("""
        **Asignación por defecto** (se aplica si no hay override):
        - **F (Cargas Sociales)** → MO mayoritaria del APU, prorrateada si hay varias
        - **H (Herramientas), L (Gastos Grales.), M (Utilidad), P (IT), O (IVA)** → BNT

        **Override** (opcional): si tu APU tiene una composición particular
        (ej. 100% MOC en MO, o quieres que el IVA vaya a BT), edita la celda
        `→RPC` correspondiente. Los cambios se aplican al guardar.
        """)

        if not self.obras.componentes_por_apu:
            st.info("No hay componentes indirectos cargados. Carga primero un archivo de APU.")
            return

        # ─────────────────────────────────────────────────────────
        # 1) Tabla editable de overrides
        # ─────────────────────────────────────────────────────────
        rows = []
        for apu, comps in self.obras.componentes_por_apu.items():
            mo_default = self.obras._mo_mayoritaria_apu(apu)
            ov = self.obras.asignacion_indirectos.get(apu, {})

            rows.append({
                'APU': apu,
                'F (Cargas Soc.)':  comps.get('cargas_sociales', 0),
                'F→RPC':            ov.get('cargas_sociales',   mo_default),
                'H (Herramientas)': comps.get('herramientas',    0),
                'H→RPC':            ov.get('herramientas',       'bnt'),
                'L (Gast. Grales.)':comps.get('gastos_generales', 0),
                'L→RPC':            ov.get('gastos_generales',   'bnt'),
                'M (Utilidad)':     comps.get('utilidad',        0),
                'M→RPC':            ov.get('utilidad',           'bnt'),
                'P (IT)':           comps.get('it',              0),
                'P→RPC':            ov.get('it',                 'bnt'),
                'O (IVA)':          comps.get('iva',             0),
                'O→RPC':            ov.get('iva',                'bnt'),
            })

        df_comp = pd.DataFrame(rows)

        edited = st.data_editor(
            df_comp,
            column_config={
                'APU': st.column_config.TextColumn("APU", disabled=True, width="large"),
                'F (Cargas Soc.)':  st.column_config.NumberColumn("F (Bs)", disabled=True, format="%.2f"),
                # F solo puede ir a categorías MO (validación de negocio)
                'F→RPC':            st.column_config.SelectboxColumn("F →",  options=RPC_MO),
                'H (Herramientas)': st.column_config.NumberColumn("H (Bs)", disabled=True, format="%.2f"),
                'H→RPC':            st.column_config.SelectboxColumn("H →",  options=RPC_VALIDOS),
                'L (Gast. Grales.)':st.column_config.NumberColumn("L (Bs)", disabled=True, format="%.2f"),
                'L→RPC':            st.column_config.SelectboxColumn("L →",  options=RPC_VALIDOS),
                'M (Utilidad)':     st.column_config.NumberColumn("M (Bs)", disabled=True, format="%.2f"),
                'M→RPC':            st.column_config.SelectboxColumn("M →",  options=RPC_VALIDOS),
                'P (IT)':           st.column_config.NumberColumn("P (Bs)", disabled=True, format="%.2f"),
                'P→RPC':            st.column_config.SelectboxColumn("P →",  options=RPC_VALIDOS),
                'O (IVA)':          st.column_config.NumberColumn("O (Bs)", disabled=True, format="%.2f"),
                'O→RPC':            st.column_config.SelectboxColumn("O →",  options=RPC_VALIDOS),
            },
            use_container_width=True,
            height=420,
            key="editor_componentes_indirectos_v2",
        )

        # ─────────────────────────────────────────────────────────
        # 2) Preview en vivo: distribución por RPC del APU seleccionado
        # ─────────────────────────────────────────────────────────
        with st.expander("👁️ Preview en vivo de la distribución por RPC", expanded=False):
            st.caption(
                "Selecciona un APU para ver cómo queda su distribución con los "
                "overrides actuales (sin guardar todavía)."
            )
            apu_preview = st.selectbox(
                "APU a previsualizar",
                options=self.obras.subcategorias,
                key="preview_apu_indirectos",
            )
            if apu_preview:
                # Construir overrides temporales a partir del editor
                row = edited[edited['APU'] == apu_preview].iloc[0]
                asign_temp = {
                    'cargas_sociales':  row['F→RPC'],
                    'herramientas':     row['H→RPC'],
                    'gastos_generales': row['L→RPC'],
                    'utilidad':         row['M→RPC'],
                    'it':               row['P→RPC'],
                    'iva':              row['O→RPC'],
                }
                #t = self.obras.calcular_totales_rpc_por_apu(apu_preview, override=asign_temp)

                t_unit = self.obras.calcular_totales_rpc_por_apu(apu_preview, override=asign_temp)
                cant = self.obras.cantidad_por_apu.get(apu_preview, 1.0)
                t = {rpc: t_unit[rpc] * cant for rpc in t_unit}

                # Render en columnas con los colores de TIPOS_RPC
                cols = st.columns(len(RPC_VALIDOS))
                total_q = sum(t.values())
                for col, rpc in zip(cols, RPC_VALIDOS):
                    with col:
                        pct = (t[rpc] / total_q * 100) if total_q > 0 else 0
                        #nombre = self.obras.TIPOS_RPC[rpc]['nombre'] if hasattr(self.obras, 'TIPOS_RPC') else rpc
                        nombre = TIPOS_RPC[rpc]['nombre']
                        st.metric(
                            label=f"{rpc.upper()} - {nombre[:18]}",
                            value=f"{t[rpc]:,.0f}",
                            delta=f"{pct:.1f}%",
                        )

                # Verificación de cierre
                q_apu = self.obras.componentes_por_apu[apu_preview].get('precio_unitario_final', 0)
                cant_apu = self.obras.cantidad_por_apu.get(apu_preview, 1.0)
                cierre = sum(t.values())
                diff = cierre - (q_apu * cant_apu)
                if abs(diff) < 0.01:
                    st.success(f"✅ Cierra exacto a Q×cant: {cierre:,.2f} = {q_apu:,.2f} × {cant_apu:,.2f}")
                else:
                    st.error(
                        f"❌ No cierra: calculado={cierre:,.2f} vs esperado={q_apu*cant_apu:,.2f} "
                        f"(diff={diff:+.2f})"
                    )

        # ─────────────────────────────────────────────────────────
        # 3) Botón guardar
        # ─────────────────────────────────────────────────────────
        st.divider()
        col_save, col_reset = st.columns([1, 1])
        with col_save:
            if st.button("💾 Guardar asignaciones de indirectos", type="primary",
                         key="btn_guardar_indirectos_v2"):
                for _, row in edited.iterrows():
                    apu = row['APU']
                    self.obras.asignacion_indirectos[apu] = {
                        'cargas_sociales':  row['F→RPC'],
                        'herramientas':     row['H→RPC'],
                        'gastos_generales': row['L→RPC'],
                        'utilidad':         row['M→RPC'],
                        'it':               row['P→RPC'],
                        'iva':              row['O→RPC'],
                    }
                st.success("✅ Asignaciones guardadas. Los totales por RPC se recalcularán automáticamente.")
                st.rerun()

        with col_reset:
            if st.button("🔄 Restaurar defaults", type="secondary",
                         key="btn_reset_indirectos_v2"):
                self.obras.asignacion_indirectos = {}
                st.info("ℹ️ Overrides eliminados. Se usarán los defaults (F→MO mayoria, H/L/M/P/O→BNT).")
                st.rerun()

    def _render_reporte_for5(self):
        st.header("📄 Reporte Formato 5 - Presupuesto Desglosado")
        st.markdown("Genera el archivo Excel con formato oficial.")

        if self.obras.df_items.empty:
            st.warning("No hay datos cargados. Cargue un archivo de APU primero.")
            return

        col1, col2 = st.columns([2, 1])
        with col1:
            if st.button("🔄 Generar vista previa FOR5", type="primary"):
                st.session_state.df_for5 = self.obras.generar_dataframe_for5()

        if 'df_for5' not in st.session_state or st.session_state.df_for5.empty:
            return

        df_original = st.session_state.df_for5

        num_cols = ['CANTIDAD', 'PRECIO UNITARIO', 'TOTAL',
                    'BIENES TRANS.', 'MATERIAL LOCAL', 'MDO CALIF.',
                    'MDO SEMI-CALIF.', 'MDO NO CALIF. URB.', 'MDO NO CALIF. RUR.']
        
        df_vista = df_original.copy()
        for col in num_cols:
            if col in df_vista.columns:
                df_vista[col] = df_vista[col].apply(
                    lambda x: f"{x:,.2f}" if isinstance(x, (int, float)) and x != '' else x
                )

        st.dataframe(df_vista, use_container_width=True, height=500)

        with col2:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_original.to_excel(writer, index=False, sheet_name='FOR5')
                ws = writer.sheets['FOR5']

                # Ajustar anchos
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 45
                ws.column_dimensions['C'].width = 10
                for col_letter in ['D','E','F','G','H','I','J','K','L','M']:
                    ws.column_dimensions[col_letter].width = 10

                # Estilos
                header_font = Font(bold=True, size=10, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    desc = row[1].value
                    if desc == "SUBTOTAL":
                        for cell in row:
                            cell.fill = subtotal_fill
                            cell.font = Font(bold=True)
                    elif desc == "TOTAL $US":
                        for cell in row:
                            cell.fill = total_fill
                            cell.font = Font(bold=True)
                    elif desc and isinstance(desc, str) and desc.isupper() and desc != "OBRAS CIVILES":
                        for cell in row:
                            cell.font = Font(bold=True, size=10)

                # Formato numérico en columnas de datos
                from openpyxl.styles import numbers
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for idx, cell in enumerate(row):
                        if idx >= 5 and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'

            st.download_button(
                label="📥 Descargar FOR5 (.xlsx)",
                data=output.getvalue(),
                file_name="Formato5_Presupuesto_Desglosado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
# Inyectar el método de clasificación inteligente en la clase
_inyectar_render_inteligente()

# ============================================================
# FUNCIÓN DE INTEGRACIÓN CON APP_V2.PY
# ============================================================
def obtener_datos_inversion_completa(periodo_anios: int = 10) -> Dict:
    if 'obras' in st.session_state and 'servicios' in st.session_state:
        obras = st.session_state.obras
        servicios = st.session_state.servicios
    else:
        obras = GestorObrasCiviles()
        servicios = GestorServicios()

    consolidador = ConsolidadorInversion(obras, servicios)
    flujos = {
        'bt': 0.0, 'bnt': 0.0, 'moc': 0.0, 'mos': 0.0,
        'monu': 0.0, 'monr': 0.0,
    }

    return {
        'inversion': type('CostosDesglose', (), {
            'bienes_transables': flujos['bt'],
            'bienes_no_transables': flujos['bnt'],
            'mo_calificada': flujos['moc'],
            'mo_semicalificada': flujos['mos'],
            'mo_no_calif_urbana': flujos['monu'],
            'mo_no_calif_rural': flujos['monr']
        })(),
        'totales_por_categoria': {
            'obras': obras.calcular_totales_rpc(),
            'servicios': servicios.calcular_totales_rpc()
        },
        'consolidado': consolidador.obtener_totales_consolidados()
    }

# ============================================================
# EJECUCIÓN
# ============================================================
def main():
    st.set_page_config(
        page_title="Módulo Inversión - Proyectos de Riego",
        page_icon="🏗️",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    inicializar_session_state()
    ui = UIInversionCompleta()
    ui.render()

if __name__ == "__main__":
    main()

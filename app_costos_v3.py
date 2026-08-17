# app_costos_v3.py (versión corregida)
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import streamlit as st
import pandas as pd
#import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Optional, Tuple

from core.config import ConfiguracionProyecto
from core.global_db import GlobalDB
from core.database import ProyectoDB
from core.project_manager import ProjectManager
from core.data_manager import DataManager
from dataclasses import dataclass
@dataclass
class ResultadoConcepto:
    """Resultado de cálculo para un concepto individual."""
    concepto: str
    unidad: str
    categoria: str
    clase_rpc: str
    cantidad_sp: float
    cantidad_cp: float
    precio_efectivo: float
    total_sp: float
    total_cp: float
    bt_sp: float
    bnt_sp: float
    monr_sp: float
    monu_sp: float
    mos_sp: float
    moc_sp: float
    bt_cp: float
    bnt_cp: float
    monr_cp: float
    monu_cp: float
    mos_cp: float
    moc_cp: float
    observaciones: str = ""

@dataclass
class ResultadoAdicional:
    """Resultado de cálculo para un costo adicional."""
    categoria: str
    descripcion: str
    tipo: str
    valor_orig: float
    base: str
    monto_sp: float
    monto_cp: float
    bt_sp: float
    bnt_sp: float
    monr_sp: float
    monu_sp: float
    mos_sp: float
    moc_sp: float
    bt_cp: float
    bnt_cp: float
    monr_cp: float
    monu_cp: float
    mos_cp: float
    moc_cp: float

@dataclass
class ResultadoCalculo:
    """Resultado completo del cálculo de costos."""
    # Por hectarea
    conceptos_sp: float
    conceptos_cp: float
    adicionales_sp: float
    adicionales_cp: float
    total_sp: float
    total_cp: float

    # Desglose RPC por hectarea
    bt_sp: float
    bnt_sp: float
    monr_sp: float
    monu_sp: float
    mos_sp: float
    moc_sp: float
    bt_cp: float
    bnt_cp: float
    monr_cp: float
    monu_cp: float
    mos_cp: float
    moc_cp: float

    # Por superficie (proyecto)
    sup_sp: float
    sup_cp: float
    total_proy_sp: float
    total_proy_cp: float
    ingreso_sp: float
    ingreso_cp: float
    utilidad_sp: float
    utilidad_cp: float
    relacion_bc_sp: float
    relacion_bc_cp: float

    # Detalles
    detalle_conceptos: List[ResultadoConcepto]
    detalle_adicionales: List[ResultadoAdicional]

# ============================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================
pm = ProjectManager()
rutas = pm.get_rutas_activas()
if rutas is None:
    st.error("No hay proyecto activo")
    st.stop()

RUTA_CONFIG = rutas["config"]
RUTA_PROYECTO_DB = pm.get_ruta_db()
RUTA_GLOBAL_DB = pm.get_ruta_global_db()
RUTA_EXCEL_PUENTE = rutas["excel"]

# ============================================================
# CALCULADORA DE COSTOS
# ============================================================
class CalculadoraCostos:
    """
    Calculadora profesional de costos de produccion agricola.
    Compatible con Resolucion Ministerial N 115/2015 (Bolivia).
    """
    def __init__(self, global_db, config):
        self.global_db = global_db
        self.config = config

    def _ceros(self) -> Dict[str, float]:
        return {'bt': 0.0, 'bnt': 0.0, 'monr': 0.0, 'monu': 0.0, 'mos': 0.0, 'moc': 0.0, 'total': 0.0}

    def _precio_efectivo(self, row: pd.Series) -> float:
        """
        Determina el precio efectivo aplicando PRECIO_OVERRIDE si existe.
        """
        precio_ref = float(row.get('PRECIO_UNITARIO', 0) or 0)
        precio_override = float(row.get('PRECIO_OVERRIDE', 0) or 0)
        return precio_override if precio_override > 0 else precio_ref

    def _cantidad_cp(self, row: pd.Series, cantidad_sp: float) -> float:
        """
        Determina la cantidad CON PROYECTO con la siguiente prioridad:
        1. CANTIDAD_CP explicita del usuario (incluyendo 0)
        2. Factor automatico segun observaciones/categoria (config)
        3. Igual a SP (sin cambios)
        """
        # Prioridad 1: Valor explicito del usuario (puede ser 0)
        if 'CANTIDAD_CP' in row:
            val_cp = row['CANTIDAD_CP']
            # Puede venir como NaN, None, o numero
            if pd.notna(val_cp):
                return float(val_cp)

        # Prioridad 2: Logica automatica segun Resolucion 115
        obs = str(row.get('OBSERVACIONES', '')).lower()
        categoria = str(row.get('CATEGORIA', '')).lower().strip()

        if any(palabra in obs for palabra in ['riego', 'ano', 'mejorado']):
            return cantidad_sp * self.config.factor_incremento_riego
        elif categoria == 'mano de obra':
            return cantidad_sp * self.config.factor_incremento_mo

        # Prioridad 3: Sin cambios
        return cantidad_sp

    def calcular_desde_df(self, df: pd.DataFrame, es_con: bool = False) -> Dict[str, float]:
        #Metodo LEGADO. Calcula totales RPC a partir de un DataFrame.

        if df.empty:
            return self._ceros()

        totales = self._ceros()
        totales.pop('total')

        for _, row in df.iterrows():
            val = row.get('CANTIDAD', 1.0)
            cantidad_sp = float(val) if pd.notna(val) else 1.0
            precio = self._precio_efectivo(row)
            clasif = str(row.get('CLASE_RPC', 'bnt')).lower().strip()

            if es_con:
                cantidad = self._cantidad_cp(row, cantidad_sp)
            else:
                cantidad = cantidad_sp

            total = cantidad * precio
            if clasif in totales:
                totales[clasif] += total
            else:
                totales['bnt'] += total

        if self.config.aplicar_gastos_administrativos:
            factor = 1.0 + self.config.porcentaje_gastos_administrativos
            for k in totales:
                totales[k] *= factor

        totales['total'] = sum(totales.values())
        return totales

    def calcular_conceptos_detallado(self, df: pd.DataFrame) -> Tuple[List[ResultadoConcepto], Dict, Dict]:
        """
        Calcula el detalle completo de conceptos para SP y CP.
        """
        if df.empty:
            return [], self._ceros(), self._ceros()

        resultados = []
        totales_sp = self._ceros()
        totales_sp.pop('total')
        totales_cp = self._ceros()
        totales_cp.pop('total')

        for _, row in df.iterrows():
            concepto = str(row.get('CONCEPTO', ''))
            unidad = str(row.get('UNIDAD', ''))
            categoria = str(row.get('CATEGORIA', '')).upper().strip()
            clasif = str(row.get('CLASE_RPC', 'bnt')).lower().strip()
            obs = str(row.get('OBSERVACIONES', ''))

            val = row.get('CANTIDAD', 1.0)
            cantidad_sp = float(val) if pd.notna(val) else 1.0
            precio = self._precio_efectivo(row)
            cantidad_cp = self._cantidad_cp(row, cantidad_sp)

            total_sp = cantidad_sp * precio
            total_cp = cantidad_cp * precio

            # Desglose RPC SP
            bt_sp = total_sp if clasif == 'bt' else 0.0
            bnt_sp = total_sp if clasif == 'bnt' else 0.0
            monr_sp = total_sp if clasif == 'monr' else 0.0
            monu_sp = total_sp if clasif == 'monu' else 0.0
            mos_sp = total_sp if clasif == 'mos' else 0.0
            moc_sp = total_sp if clasif == 'moc' else 0.0

            # Desglose RPC CP
            bt_cp = total_cp if clasif == 'bt' else 0.0
            bnt_cp = total_cp if clasif == 'bnt' else 0.0
            monr_cp = total_cp if clasif == 'monr' else 0.0
            monu_cp = total_cp if clasif == 'monu' else 0.0
            mos_cp = total_cp if clasif == 'mos' else 0.0
            moc_cp = total_cp if clasif == 'moc' else 0.0

            # Acumular totales
            totales_sp['bt'] += bt_sp
            totales_sp['bnt'] += bnt_sp
            totales_sp['monr'] += monr_sp
            totales_sp['monu'] += monu_sp
            totales_sp['mos'] += mos_sp
            totales_sp['moc'] += moc_sp

            totales_cp['bt'] += bt_cp
            totales_cp['bnt'] += bnt_cp
            totales_cp['monr'] += monr_cp
            totales_cp['monu'] += monu_cp
            totales_cp['mos'] += mos_cp
            totales_cp['moc'] += moc_cp

            resultados.append(ResultadoConcepto(
                concepto=concepto, unidad=unidad, categoria=categoria,
                clase_rpc=clasif, cantidad_sp=cantidad_sp, cantidad_cp=cantidad_cp,
                precio_efectivo=precio, total_sp=total_sp, total_cp=total_cp,
                bt_sp=bt_sp, bnt_sp=bnt_sp, monr_sp=monr_sp, monu_sp=monu_sp,
                mos_sp=mos_sp, moc_sp=moc_sp, bt_cp=bt_cp, bnt_cp=bnt_cp,
                monr_cp=monr_cp, monu_cp=monu_cp, mos_cp=mos_cp, moc_cp=moc_cp,
                observaciones=obs
            ))

        totales_sp['total'] = sum(totales_sp.values())
        totales_cp['total'] = sum(totales_cp.values())

        return resultados, totales_sp, totales_cp

    def calcular_costos_adicionales_v2(
        self, 
        subtotal_directos_sp: float, 
        subtotal_directos_cp: float,
        proporciones_rpc_sp: Dict[str, float],
        proporciones_rpc_cp: Dict[str, float],
        df_adicionales: pd.DataFrame
    ) -> Tuple[List[ResultadoAdicional], Dict, Dict]:
        """
        Calcula costos adicionales con prorrateo RPC profesional.

        Args:
            subtotal_directos_sp: Suma de conceptos SP
            subtotal_directos_cp: Suma de conceptos CP
            proporciones_rpc_sp: Dict {rpc: proporcion} para SP
            proporciones_rpc_cp: Dict {rpc: proporcion} para CP
            df_adicionales: DataFrame con TIPO, VALOR, BASE_CALCULO, CATEGORIA, DESCRIPCION

        Retorna:
            - Lista de ResultadoAdicional
            - Totales SP por RPC (incluyendo adicionales)
            - Totales CP por RPC (incluyendo adicionales)
        """
        if df_adicionales.empty:
            return [], self._ceros(), self._ceros()

        resultados = []
        adic_sp = self._ceros()
        adic_sp.pop('total')
        adic_cp = self._ceros()
        adic_cp.pop('total')

        # Para base 'total', necesitamos acumular los montos previos
        acum_sp = 0.0
        acum_cp = 0.0

        for _, row in df_adicionales.iterrows():
            tipo = str(row.get('TIPO', 'fijo')).lower().strip()
            valor = float(row.get('VALOR', 0) or 0)
            base = str(row.get('BASE_CALCULO', 'ninguna')).lower().strip()
            categoria = str(row.get('CATEGORIA', ''))
            desc = str(row.get('DESCRIPCION', ''))

            # Calcular montos SP y CP
            if tipo == 'porcentaje':
                if base == 'directos':
                    monto_sp = subtotal_directos_sp * (valor / 100.0)
                    monto_cp = subtotal_directos_cp * (valor / 100.0)
                elif base == 'total':
                    monto_sp = (subtotal_directos_sp + acum_sp) * (valor / 100.0)
                    monto_cp = (subtotal_directos_cp + acum_cp) * (valor / 100.0)
                else:
                    monto_sp = 0.0
                    monto_cp = 0.0
            else:  # fijo
                monto_sp = valor
                monto_cp = valor

            # Prorratear por RPC segun proporciones de costos directos
            def prorratear(monto, props):
                return {
                    'bt': monto * props.get('bt', 0),
                    'bnt': monto * props.get('bnt', 0),
                    'monr': monto * props.get('monr', 0),
                    'monu': monto * props.get('monu', 0),
                    'mos': monto * props.get('mos', 0),
                    'moc': monto * props.get('moc', 0)
                }

            rpc_sp = prorratear(monto_sp, proporciones_rpc_sp)
            rpc_cp = prorratear(monto_cp, proporciones_rpc_cp)

            # Acumular para base 'total'
            acum_sp += monto_sp
            acum_cp += monto_cp

            # Sumar a totales de adicionales
            for k in adic_sp:
                adic_sp[k] += rpc_sp[k]
                adic_cp[k] += rpc_cp[k]

            resultados.append(ResultadoAdicional(
                categoria=categoria, descripcion=desc, tipo=tipo,
                valor_orig=valor, base=base,
                monto_sp=monto_sp, monto_cp=monto_cp,
                bt_sp=rpc_sp['bt'], bnt_sp=rpc_sp['bnt'], monr_sp=rpc_sp['monr'],
                monu_sp=rpc_sp['monu'], mos_sp=rpc_sp['mos'], moc_sp=rpc_sp['moc'],
                bt_cp=rpc_cp['bt'], bnt_cp=rpc_cp['bnt'], monr_cp=rpc_cp['monr'],
                monu_cp=rpc_cp['monu'], mos_cp=rpc_cp['mos'], moc_cp=rpc_cp['moc']
            ))

        adic_sp['total'] = sum(adic_sp.values())
        adic_cp['total'] = sum(adic_cp.values())

        return resultados, adic_sp, adic_cp

    def calcular_costos_completos(
        self,
        df_conceptos: pd.DataFrame,
        df_adicionales: pd.DataFrame,
        sup_sp: float = 1.0,
        sup_cp: float = 1.0,
        rend_sp: float = 0.0,
        rend_cp: float = 0.0,
        precio_ton: float = 0.0,
        perd_sp_pct: float = 0.0,
        perd_cp_pct: float = 0.0
    ) -> ResultadoCalculo:
        """
        Metodo principal: calcula costos completos (por hectarea y por superficie)
        incluyendo desglose RPC, ingresos, utilidades y relacion B/C.
        """
        # -- 1. Calcular conceptos detallados --
        det_conceptos, tot_sp, tot_cp = self.calcular_conceptos_detallado(df_conceptos)

        subtotal_directos_sp = tot_sp['total']
        subtotal_directos_cp = tot_cp['total']

        # -- 2. Calcular proporciones RPC para prorrateo de adicionales --
        def calcular_props(totales):
            total = totales['total']
            if total == 0:
                return {k: 0.0 for k in ['bt', 'bnt', 'monr', 'monu', 'mos', 'moc']}
            return {k: totales[k] / total for k in ['bt', 'bnt', 'monr', 'monu', 'mos', 'moc']}

        props_sp = calcular_props(tot_sp)
        props_cp = calcular_props(tot_cp)

        # -- 3. Calcular costos adicionales con prorrateo --
        det_adicionales, adic_sp, adic_cp = self.calcular_costos_adicionales_v2(
            subtotal_directos_sp, subtotal_directos_cp,
            props_sp, props_cp, df_adicionales
        )

        # -- 4. Sumar totales por hectarea --
        total_sp_ha = subtotal_directos_sp + adic_sp['total']
        total_cp_ha = subtotal_directos_cp + adic_cp['total']

        # Totales RPC por hectarea
        bt_sp_ha = tot_sp['bt'] + adic_sp['bt']
        bnt_sp_ha = tot_sp['bnt'] + adic_sp['bnt']
        monr_sp_ha = tot_sp['monr'] + adic_sp['monr']
        monu_sp_ha = tot_sp['monu'] + adic_sp['monu']
        mos_sp_ha = tot_sp['mos'] + adic_sp['mos']
        moc_sp_ha = tot_sp['moc'] + adic_sp['moc']

        bt_cp_ha = tot_cp['bt'] + adic_cp['bt']
        bnt_cp_ha = tot_cp['bnt'] + adic_cp['bnt']
        monr_cp_ha = tot_cp['monr'] + adic_cp['monr']
        monu_cp_ha = tot_cp['monu'] + adic_cp['monu']
        mos_cp_ha = tot_cp['mos'] + adic_cp['mos']
        moc_cp_ha = tot_cp['moc'] + adic_cp['moc']

        # -- 5. Escalar por superficie --
        total_proy_sp = total_sp_ha * sup_sp
        total_proy_cp = total_cp_ha * sup_cp

        # -- 6. Calcular ingresos --
        factor_kg_a_ton = 1000.0
        ingreso_sp = (rend_sp / factor_kg_a_ton) * (1.0 - perd_sp_pct / 100.0) * precio_ton * sup_sp
        ingreso_cp = (rend_cp / factor_kg_a_ton) * (1.0 - perd_cp_pct / 100.0) * precio_ton * sup_cp
        #ingreso_sp = rend_sp * (1.0 - perd_sp_pct / 100.0) * precio_ton * sup_sp
        #ingreso_cp = rend_cp * (1.0 - perd_cp_pct / 100.0) * precio_ton * sup_cp

        # -- 7. Utilidades y B/C --
        utilidad_sp = ingreso_sp - total_proy_sp
        utilidad_cp = ingreso_cp - total_proy_cp
        bc_sp = ingreso_sp / total_proy_sp if total_proy_sp > 0 else 0.0
        bc_cp = ingreso_cp / total_proy_cp if total_proy_cp > 0 else 0.0

        return ResultadoCalculo(
            conceptos_sp=subtotal_directos_sp,
            conceptos_cp=subtotal_directos_cp,
            adicionales_sp=adic_sp['total'],
            adicionales_cp=adic_cp['total'],
            total_sp=total_sp_ha,
            total_cp=total_cp_ha,
            bt_sp=bt_sp_ha, bnt_sp=bnt_sp_ha, monr_sp=monr_sp_ha,
            monu_sp=monu_sp_ha, mos_sp=mos_sp_ha, moc_sp=moc_sp_ha,
            bt_cp=bt_cp_ha, bnt_cp=bnt_cp_ha, monr_cp=monr_cp_ha,
            monu_cp=monu_cp_ha, mos_cp=mos_cp_ha, moc_cp=moc_cp_ha,
            sup_sp=sup_sp, sup_cp=sup_cp,
            total_proy_sp=total_proy_sp, total_proy_cp=total_proy_cp,
            ingreso_sp=ingreso_sp, ingreso_cp=ingreso_cp,
            utilidad_sp=utilidad_sp, utilidad_cp=utilidad_cp,
            relacion_bc_sp=bc_sp, relacion_bc_cp=bc_cp,
            detalle_conceptos=det_conceptos,
            detalle_adicionales=det_adicionales
        )

# ============================================================
# GENERADOR DE REPORTES (sin cambios, solo ajuste de importación)
# ============================================================
class ReporteGenerator:
    def generar_reporte_detallado_v2(self, cultivo_data, df_conceptos, df_adicionales,
                                      depto="No especificado", municipio="No especificado",
                                      proyecto_nombre="", tipo_cambio=6.96,
                                      variedad="", mes_siembra="", mes_cosecha=""):

        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Costos_Detallados"
        
        # Configurar anchos de columna optimizados
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 10
        ws.column_dimensions['N'].width = 10
        ws.column_dimensions['O'].width = 10
        ws.column_dimensions['P'].width = 10
        ws.column_dimensions['Q'].width = 10
        ws.column_dimensions['R'].width = 10
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=9)
        title_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True, size=11)
        label_font = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        double_border_top = Border(top=Side(style='double'))
        
        # ENCABEZADO PRINCIPAL
        ws.merge_cells('A1:R1')
        ws['A1'] = "COSTOS DE PRODUCCION POR HECTAREA (Bs.)"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        ws.merge_cells('A2:R2')
        ws['A2'] = proyecto_nombre
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].font = Font(italic=True, size=10)
        
        # INFORMACIÓN GENERAL — FILA 4
        ws['A4'] = "DEPARTAMENTO:"
        ws['B4'] = depto
        ws['B4'].font = label_font
        ws['D4'] = "MUNICIPIO:"
        ws['E4'] = municipio
        ws['E4'].font = label_font
        ws['G4'] = "CULTIVO:"
        ws['H4'] = cultivo_data.get('Cultivo', 'No especificado')
        ws['H4'].font = label_font
        ws['J4'] = "VARIEDAD:"
        ws['K4'] = variedad
        ws['K4'].font = label_font
        ws['M4'] = "SIEMBRA:"
        ws['N4'] = mes_siembra
        ws['N4'].font = label_font
        ws['P4'] = "COSECHA:"
        ws['Q4'] = mes_cosecha
        ws['Q4'].font = label_font
        
        # ENCABEZADOS DE TABLA
        row = 6
        
        # Título de secciones SP y CP
        ws.merge_cells(f'C{row}:E{row}')
        ws[f'C{row}'] = "SIN  PROYECTO"
        ws[f'C{row}'].font = subtitle_font
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        ws[f'C{row}'].fill = PatternFill(start_color="B4C7DC", fill_type="solid")

        ws.merge_cells(f'F{row}:H{row}')
        ws[f'F{row}'] = "CON  PROYECTO"
        ws[f'F{row}'].font = subtitle_font
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        ws[f'F{row}'].fill = PatternFill(start_color="B4C7DC", fill_type="solid")

        ws.merge_cells(f'I{row}:M{row}')
        ws[f'I{row}'] = "SIN PROYECTO - RPC"
        ws[f'I{row}'].font = subtitle_font
        ws[f'I{row}'].alignment = Alignment(horizontal='center')
        ws[f'I{row}'].fill = PatternFill(start_color="B4C7DC", fill_type="solid")
        
        ws.merge_cells(f'N{row}:R{row}')
        ws[f'N{row}'] = "CON PROYECTO - RPC"
        ws[f'N{row}'].font = subtitle_font
        ws[f'N{row}'].alignment = Alignment(horizontal='center')
        ws[f'N{row}'].fill = PatternFill(start_color="B4C7DC", fill_type="solid")
        
        #PROVISIONAL BORDES
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
            ws[f'{col}{row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))       

        row += 1        
        # Encabezados de columnas detalladas
        headers = [
            ("CONCEPTO", 38), ("UNID.", 10), ("CANTID.", 10), ("PRECIO\nUNIT.Bs.", 11), 
            ("TOTAL\nBs.", 12), ("CANTID.", 10), ("PRECIO\nUNIT.Bs.", 11), ("TOTAL\nBs.", 12),
            ("BIENES\nTRANSABLES", 12), ("MATERIAL\nLOCAL", 12), ("Mano de\nObra Local", 12), 
            ("Mano de\nObra Semical.", 12), ("Mano de\nObra Calific.", 12),
            ("BIENES\nTRANSABLES", 12), ("MATERIAL\nLOCAL", 12), ("Mano de\nObra Local", 12), 
            ("Mano de\nObra Semical.", 12), ("Mano de\nObra Calific.", 12)
        ]
        
        for idx, (header, width) in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        ws.row_dimensions[row].height = 35
        current_row = row + 1
        
        # PROCESAR CONCEPTOS desde df_conceptos
        if df_conceptos.empty:
            return None
        
        # Estructura para totales
        categorias_orden = ['MANO DE OBRA', 'SERVICIOS NO PERSONALES', 'INSUMOS', 'OTROS COSTOS DIRECTOS']
        totales_cat = {cat: {'sp': 0, 'cp': 0, 'bt_sp': 0, 'bnt_sp': 0, 'monr_sp': 0, 'mos_sp': 0, 'moc_sp': 0,
                            'bt_cp': 0, 'bnt_cp': 0, 'monr_cp': 0, 'mos_cp': 0, 'moc_cp': 0} for cat in categorias_orden}
        
        for categoria in categorias_orden:
            # Filtrar conceptos de esta categoría
            mask = df_conceptos['CATEGORIA'].str.upper().str.strip() == categoria
            df_cat = df_conceptos[mask]
            
            if df_cat.empty:
                continue
                
            # Encabezado de categoría
            ws.merge_cells(f'A{current_row}:R{current_row}')
            ws[f'A{current_row}'] = categoria
            ws[f'A{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'A{current_row}'].fill = PatternFill(start_color="5B9BD5", fill_type="solid")
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', indent=1)
            current_row += 1
            
            # Procesar cada concepto
            for _, row_data in df_cat.iterrows():
                concepto = row_data.get('CONCEPTO', '')
                unidad = row_data.get('UNIDAD', '')
                cantidad_base = float(row_data.get('CANTIDAD', 0) or 0)
                cantidad_cp = float(row_data.get('CANTIDAD_CP', cantidad_base) or cantidad_base)
                precio = float(row_data.get('PRECIO_UNITARIO', 0) or 0)
                clase_rpc = str(row_data.get('CLASE_RPC', 'bnt')).lower().strip()
                
                # CÁLCULO SIN PROYECTO
                total_sp = cantidad_base * precio
                
                # CÁLCULO CON PROYECTO
                total_cp = cantidad_cp * precio
                
                # Desglose RPC SP
                bt_sp = total_sp if clase_rpc == 'bt' else 0
                bnt_sp = total_sp if clase_rpc == 'bnt' else 0
                monr_sp = total_sp if clase_rpc == 'monr' else 0
                mos_sp = total_sp if clase_rpc == 'mos' else 0
                moc_sp = total_sp if clase_rpc == 'moc' else 0
                
                # Desglose RPC CP
                bt_cp = total_cp if clase_rpc == 'bt' else 0
                bnt_cp = total_cp if clase_rpc == 'bnt' else 0
                monr_cp = total_cp if clase_rpc == 'monr' else 0
                mos_cp = total_cp if clase_rpc == 'mos' else 0
                moc_cp = total_cp if clase_rpc == 'moc' else 0
                
                # Acumular totales
                totales_cat[categoria]['sp'] += total_sp
                totales_cat[categoria]['cp'] += total_cp
                totales_cat[categoria]['bt_sp'] += bt_sp
                totales_cat[categoria]['bnt_sp'] += bnt_sp
                totales_cat[categoria]['monr_sp'] += monr_sp
                totales_cat[categoria]['mos_sp'] += mos_sp
                totales_cat[categoria]['moc_sp'] += moc_sp
                totales_cat[categoria]['bt_cp'] += bt_cp
                totales_cat[categoria]['bnt_cp'] += bnt_cp
                totales_cat[categoria]['monr_cp'] += monr_cp
                totales_cat[categoria]['mos_cp'] += mos_cp
                totales_cat[categoria]['moc_cp'] += moc_cp
                
                # Escribir fila
                values = [
                    concepto, unidad, cantidad_base, precio, total_sp,
                    cantidad_cp, precio, total_cp,
                    bt_sp if bt_sp > 0 else None, 
                    bnt_sp if bnt_sp > 0 else None,
                    monr_sp if monr_sp > 0 else None,
                    mos_sp if mos_sp > 0 else None,
                    moc_sp if moc_sp > 0 else None,
                    bt_cp if bt_cp > 0 else None,
                    bnt_cp if bnt_cp > 0 else None,
                    monr_cp if monr_cp > 0 else None,
                    mos_cp if mos_cp > 0 else None,
                    moc_cp if moc_cp > 0 else None
                ]
                
                for idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=idx, value=val)
                    if idx > 2:
                        cell.number_format = '#,##0.00'
                    if idx >= 9:
                        cell.border = thin_border
                
                current_row += 1
            
            # Subtotal categoría
            ws[f'A{current_row}'] = f"Sub-total {categoria.title()}"
            ws[f'A{current_row}'].font = Font(bold=True)
            
            ws[f'E{current_row}'] = totales_cat[categoria]['sp']
            ws[f'H{current_row}'] = totales_cat[categoria]['cp']
            ws[f'I{current_row}'] = totales_cat[categoria]['bt_sp']
            ws[f'J{current_row}'] = totales_cat[categoria]['bnt_sp']
            ws[f'K{current_row}'] = totales_cat[categoria]['monr_sp']
            ws[f'L{current_row}'] = totales_cat[categoria]['mos_sp']
            ws[f'M{current_row}'] = totales_cat[categoria]['moc_sp']
            ws[f'N{current_row}'] = totales_cat[categoria]['bt_cp']
            ws[f'O{current_row}'] = totales_cat[categoria]['bnt_cp']
            ws[f'P{current_row}'] = totales_cat[categoria]['monr_cp']
            ws[f'Q{current_row}'] = totales_cat[categoria]['mos_cp']
            ws[f'R{current_row}'] = totales_cat[categoria]['moc_cp']
            
            for col in ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q','R']:
                ws[f'{col}{current_row}'].font = Font(bold=True)
                ws[f'{col}{current_row}'].number_format = '#,##0.00'
                ws[f'{col}{current_row}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
            
            current_row += 1
        
        # COSTOS INDIRECTOS (antes Adicionales)
        if not df_adicionales.empty:
            current_row += 1
            ws.merge_cells(f'A{current_row}:R{current_row}')
            ws[f'A{current_row}'] = "COSTOS INDIRECTOS"
            ws[f'A{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'A{current_row}'].fill = PatternFill(start_color="70AD47", fill_type="solid")
            current_row += 1
            
            for _, row_adic in df_adicionales.iterrows():
                categoria = row_adic.get('CATEGORIA', '')
                desc = row_adic.get('DESCRIPCION', '')
                monto_sp = float(row_adic.get('MONTO_SP', 0) or 0)
                monto_cp = float(row_adic.get('MONTO_CP', 0) or 0)
                
                ws[f'A{current_row}'] = f"{categoria} - {desc}"
                ws[f'E{current_row}'] = monto_sp
                ws[f'H{current_row}'] = monto_cp
                ws[f'E{current_row}'].number_format = '#,##0.00'
                ws[f'H{current_row}'].number_format = '#,##0.00'
                current_row += 1
        
        # TOTAL COSTOS
        current_row += 1
        subtotal_sp = sum(totales_cat[cat]['sp'] for cat in categorias_orden)
        subtotal_cp = sum(totales_cat[cat]['cp'] for cat in categorias_orden)
        
        # Sumar adicionales si existen
        if not df_adicionales.empty:
            total_adic_sp = df_adicionales['MONTO_SP'].sum() if 'MONTO_SP' in df_adicionales.columns else 0
            total_adic_cp = df_adicionales['MONTO_CP'].sum() if 'MONTO_CP' in df_adicionales.columns else 0
        else:
            total_adic_sp = 0
            total_adic_cp = 0
        
        total_costo_sp = subtotal_sp + total_adic_sp
        total_costo_cp = subtotal_cp + total_adic_cp
        
        ws[f'A{current_row}'] = "TOTAL  COSTOS (Bs/Ha)"
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        ws[f'E{current_row}'] = total_costo_sp
        ws[f'H{current_row}'] = total_costo_cp
        ws[f'E{current_row}'].font = Font(bold=True, size=11)
        ws[f'H{current_row}'].font = Font(bold=True, size=11)
        ws[f'E{current_row}'].fill = title_fill
        ws[f'H{current_row}'].fill = title_fill
        ws[f'E{current_row}'].border = thin_border
        ws[f'H{current_row}'].border = thin_border
        
        # INGRESOS Y UTILIDAD — con formato numérico corregido
        current_row += 2

        rend_sp = float(cultivo_data.get('Rendimiento_SP', 0))
        rend_cp = float(cultivo_data.get('Rendimiento_CP', 0))
        precio = float(cultivo_data.get('Precio_Bs_Ton', 0))
        perd_sp = float(cultivo_data.get('Perdidas_SP_pct', 0)) / 100
        perd_cp = float(cultivo_data.get('Perdidas_CP_pct', 0)) / 100       
        # CORREGIDO: rendimiento en kg/ha → ton/ha
        ingreso_sp = (rend_sp / 1000.0) * (1.0 - perd_sp) * precio
        ingreso_cp = (rend_cp / 1000.0) * (1.0 - perd_cp) * precio

        ws[f'A{current_row}'] = "Rendimiento/precio/ingreso"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'C{current_row}'] = rend_sp
        ws[f'D{current_row}'] = precio
        ws[f'E{current_row}'] = ingreso_sp
        ws[f'F{current_row}'] = rend_cp
        ws[f'G{current_row}'] = precio
        ws[f'H{current_row}'] = ingreso_cp
        for col in ['C','D','E','F','G','H']:
            ws[f'{col}{current_row}'].number_format = '#,##0.00'
        
        current_row += 1
        ws[f'A{current_row}'] = "TOTAL INGRESOS (Bs/Ha)"
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        ws[f'E{current_row}'] = ingreso_sp
        ws[f'H{current_row}'] = ingreso_cp
        for col in ['E','H']:
            cell = ws[f'{col}{current_row}']
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="E2EFDA", fill_type="solid")
            cell.number_format = '#,##0.00'
        
        current_row += 1
        utilidad_sp = ingreso_sp - total_costo_sp
        utilidad_cp = ingreso_cp - total_costo_cp
        
        ws[f'A{current_row}'] = "UTILIDAD (Bs/Ha)"
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        ws[f'E{current_row}'] = utilidad_sp
        ws[f'H{current_row}'] = utilidad_cp
        for col in ['E','H']:
            cell = ws[f'{col}{current_row}']
            cell.font = Font(bold=True, size=11)
            cell.number_format = '#,##0.00'
        
        current_row += 1
        bc_sp = ingreso_sp / total_costo_sp if total_costo_sp > 0 else 0
        bc_cp = ingreso_cp / total_costo_cp if total_costo_cp > 0 else 0
        
        ws[f'A{current_row}'] = "Relación B/C"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'E{current_row}'] = bc_sp
        ws[f'H{current_row}'] = bc_cp
        for col in ['E','H']:
            cell = ws[f'{col}{current_row}']
            cell.font = Font(bold=True)
            cell.number_format = '0.00'
        
        # Tipo de cambio
        current_row += 2
        ws[f'A{current_row}'] = f"TIPO DE CAMBIO:   1 $US = {tipo_cambio} Bs"
        ws[f'A{current_row}'].font = Font(italic=True)
        
        # Costos en USD (lateral)
        ws[f'D{current_row}'] = "Costo T. sin proy. ($US)"
        ws[f'E{current_row}'] = total_costo_sp / tipo_cambio
        ws[f'E{current_row}'].number_format = '#,##0.00'
        ws[f'G{current_row}'] = "Costo T. con proy.($US)"
        ws[f'H{current_row}'] = total_costo_cp / tipo_cambio
        ws[f'H{current_row}'].number_format = '#,##0.00'
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Informe_Detallado_{cultivo_data.get('Cultivo', 'Cultivo').replace(' ', '_')}_{timestamp}.xlsx"
        wb.save(output_path)
        return output_path
# ============================================================
# FUNCIÓN DE EXPORTACIÓN (con sincronización)
# ============================================================
def _agregar_cultivo_a_proyecto_db(proyecto_db, global_db, plantilla_id, resultado, sup_sp_calc, sup_cp_calc):
    """Agrega o actualiza el cultivo calculado en la tabla persistente proyecto_cultivos."""
    try:
        plantilla = global_db.obtener_plantilla_por_id(plantilla_id)
        if not plantilla:
            return False, "Plantilla no encontrada"
        cultivo_nombre = global_db.obtener_nombre_cultivo(plantilla['cultivo_id'])
        rend_sp   = float(plantilla.get('rendimiento_sp', 0) or 0)
        rend_cp   = float(plantilla.get('rendimiento_cp', 0) or 0)
        precio_ton = float(plantilla.get('precio_ref_bs_ton', 0) or 0)
        perd_sp   = float(plantilla.get('perdidas_sp_pct', 0) or 0)
        perd_cp   = float(plantilla.get('perdidas_cp_pct', 0) or 0)

        datos = {
            "nombre": cultivo_nombre,
            "codigo": "",
            "sup_sp_ha": float(sup_sp_calc),
            "sup_cp_ha": float(sup_cp_calc),
            "rend_sp": rend_sp,
            "rend_cp": rend_cp,
            "perd_sp_pct": perd_sp,
            "perd_cp_pct": perd_cp,
            "precio_bs_ton": precio_ton,
            "costo_total_sp": float(resultado.total_sp),
            "costo_total_cp": float(resultado.total_cp),
            "bt_sp": float(resultado.bt_sp), "bnt_sp": float(resultado.bnt_sp),
            "monr_sp": float(resultado.monr_sp), "monu_sp": float(resultado.monu_sp),
            "mos_sp": float(resultado.mos_sp), "moc_sp": float(resultado.moc_sp),
            "bt_cp": float(resultado.bt_cp), "bnt_cp": float(resultado.bnt_cp),
            "monr_cp": float(resultado.monr_cp), "monu_cp": float(resultado.monu_cp),
            "mos_cp": float(resultado.mos_cp), "moc_cp": float(resultado.moc_cp),
            # Esta parte de ingresos se debe verificar para considerar las Superficies reales de cultivo - ESTAS 2 LINEAS SON LA CLAVE
            "ingreso_sp": (rend_sp / 1000.0) * (1.0 - perd_sp / 100.0) * precio_ton,
            "ingreso_cp": (rend_cp / 1000.0) * (1.0 - perd_cp / 100.0) * precio_ton,
        }
        proyecto_db.agregar_cultivo_proyecto(datos)
        return True, cultivo_nombre
    except Exception as e:
        return False, str(e)

def exportar_cultivo_a_proyecto_db(proyecto_db: ProyectoDB, data_manager: DataManager,
                                   gestor_ref, cultivo_nombre: str, sup_sp: float, sup_cp: float):
    fila = gestor_ref.obtener_por_nombre(cultivo_nombre)
    if fila is None:
        return False

    datos = {
        "nombre": cultivo_nombre,
        "codigo": fila.get("Cod", ""),
        "sup_sp_ha": sup_sp,
        "sup_cp_ha": sup_cp,
        "rend_sp": float(fila.get("Rendimiento_SP", 0)),
        "rend_cp": float(fila.get("Rendimiento_CP", 0)),
        "perd_sp_pct": float(fila.get("Perdidas_SP_pct", 0)),
        "perd_cp_pct": float(fila.get("Perdidas_CP_pct", 0)),
        "precio_bs_ton": float(fila.get("Precio_Bs_Ton", 0)),
        "costo_total_sp": float(fila.get("CostoTotal_SP", 0)),
        "costo_total_cp": float(fila.get("CostoTotal_CP", 0)),
        "bt_sp": float(fila.get("BT_SP", 0)),
        "bnt_sp": float(fila.get("BNT_SP", 0)),
        "monr_sp": float(fila.get("MONR_SP", 0)),
        "monu_sp": float(fila.get("MONU_SP", 0)),
        "mos_sp": float(fila.get("MOS_SP", 0)),
        "moc_sp": float(fila.get("MOC_SP", 0)),
        "bt_cp": float(fila.get("BT_CP", 0)),
        "bnt_cp": float(fila.get("BNT_CP", 0)),
        "monr_cp": float(fila.get("MONR_CP", 0)),
        "monu_cp": float(fila.get("MONU_CP", 0)),
        "mos_cp": float(fila.get("MOS_CP", 0)),
        "moc_cp": float(fila.get("MOC_CP", 0)),
        "ingreso_sp": (float(fila.get("Rendimiento_SP", 0)) *
                      (1 - float(fila.get("Perdidas_SP_pct", 0))/100) *
                      float(fila.get("Precio_Bs_Ton", 0)) * sup_sp),
        "ingreso_cp": (float(fila.get("Rendimiento_CP", 0)) *
                      (1 - float(fila.get("Perdidas_CP_pct", 0))/100) *
                      float(fila.get("Precio_Bs_Ton", 0)) * sup_cp)
    }

    if proyecto_db.agregar_cultivo_proyecto(datos):
        data_manager.actualizar_desde_proyecto_db(proyecto_db)
        return True
    return False

# ============================================================
# INTERFAZ PRINCIPAL
# ============================================================
def inicializar_session_state():
    if 'global_db' not in st.session_state:
        st.session_state.global_db = GlobalDB(RUTA_GLOBAL_DB)
    if 'proyecto_db' not in st.session_state:
        st.session_state.proyecto_db = ProyectoDB(RUTA_PROYECTO_DB)
    if 'data_manager' not in st.session_state:
        st.session_state.data_manager = DataManager(RUTA_EXCEL_PUENTE)
    # ── NUEVO: exponer configuración del proyecto activo ──
    if 'config_proyecto' not in st.session_state:
        st.session_state.config_proyecto = ConfiguracionProyecto.cargar(RUTA_CONFIG)
    if 'calculadora' not in st.session_state:
        st.session_state.calculadora = CalculadoraCostos(
            st.session_state.global_db,
            st.session_state.config_proyecto
        )
    if 'reporte_gen' not in st.session_state:
        st.session_state.reporte_gen = ReporteGenerator()

def rerun():
    try:
        st.rerun()
    except AttributeError:
        st.experimental_rerun()

# ------------------------------------------------------------
# RENDER: Gestión de Conceptos
# ------------------------------------------------------------
def render_gestion_conceptos():
    st.header("📋 Gestión de Conceptos de Costo")
    global_db = st.session_state.global_db

    st.subheader("Lista de conceptos")
    df = global_db.listar_conceptos()
    if df.empty:
        st.info("No hay conceptos. Agregue usando el formulario.")
    else:
        edited_df = st.data_editor(
            df,
            use_container_width=True,
            num_rows="dynamic",
            key="editor_conceptos"
        )
        if st.button("💾 Guardar cambios en conceptos"):
            if global_db.guardar_conceptos_bulk(edited_df):
                st.success("Cambios guardados")
                rerun()
            else:
                st.error("Error al guardar")

    with st.expander("➕ Agregar nuevo concepto"):
        with st.form("nuevo_concepto"):
            concepto = st.text_input("Concepto")
            unidad = st.text_input("Unidad")
            cantidad = st.number_input("Cantidad", min_value=0.0, step=0.1)
            precio = st.number_input("Precio unitario (Bs)", min_value=0.0, step=1.0)
            cultivo = st.text_input("Cultivo")
            lugar = st.text_input("Lugar")
            categ = st.selectbox("Categoria", ["Mano de obra", "Insumos", "Servicios no personales", "Otros costos directos"])
            clasif = st.selectbox("Clasificación RPC", ["bt", "bnt", "monr", "monu", "mos", "moc"])
            obs = st.text_input("Observaciones")
            submit = st.form_submit_button("Agregar")
            if submit:
                nuevo = {
                    'CONCEPTO': concepto, 'UNIDAD': unidad, 'CANTIDAD': cantidad,
                    'PRECIO_UNITARIO': precio, 'CULTIVO': cultivo, 'CATEGORIA': categ,
                    'CLASE_RPC': clasif, 'LUGAR': lugar, 'OBSERVACIONES': obs
                }
                if global_db.agregar_concepto(nuevo):
                    st.success("Concepto agregado")
                    rerun()
                else:
                    st.error("Error al agregar")

# ------------------------------------------------------------
# RENDER: Gestión de Cultivos Referencia (Plantillas)
# ------------------------------------------------------------
def render_gestion_referencia():
    st.header("🌾 Gestión de Cultivos Referencia")
    st.markdown("Administre plantillas de costos. Puede escribir valores nuevos o reutilizar existentes.")
    global_db = st.session_state.global_db

    # ── Estado de la UI ──
    if 'ref_modo' not in st.session_state:
        st.session_state.ref_modo = 'listado'
    if 'ref_edit_id' not in st.session_state:
        st.session_state.ref_edit_id = None

    # ═══════════════════════════════════════════════════════
    # 1. MODO LISTADO
    # ═══════════════════════════════════════════════════════
    if st.session_state.ref_modo == 'listado':
        df_plantillas = global_db.listar_plantillas(solo_activas=False)

        if df_plantillas.empty:
            st.info("No hay plantillas registradas. Cree la primera usando el botón inferior.")
        else:
            st.subheader("📋 Catálogo de plantillas")
            cols_vis = ['CULTIVO', 'VARIEDAD', 'NIVEL_TECNOLOGICO', 'DEPARTAMENTO',
                        'MUNICIPIO', 'CAMPANIA', 'RENDIMIENTO_SP', 'RENDIMIENTO_CP',
                        'PRECIO_REF_BS_TON']
            cols_existen = [c for c in cols_vis if c in df_plantillas.columns]
            st.dataframe(df_plantillas[cols_existen], use_container_width=True, hide_index=True)

            c1, c2, c3 = st.columns(3)
            c1.metric("Total plantillas", len(df_plantillas))
            c2.metric("Activas", int((df_plantillas['ACTIVO'] == 1).sum()) if 'ACTIVO' in df_plantillas.columns else "—")
            c3.metric("Cultivos distintos", df_plantillas['CULTIVO'].nunique() if 'CULTIVO' in df_plantillas.columns else "—")

        st.markdown("---")
        col_a, col_b = st.columns([1, 3])
        with col_a:
            if st.button("➕ Nueva plantilla", type="primary", use_container_width=True):
                st.session_state.ref_modo = 'crear'
                st.session_state.ref_edit_id = None
                rerun()

        if not df_plantillas.empty:
            with col_b:
                opciones = {}
                for _, row in df_plantillas.iterrows():
                    label = f"{row['CULTIVO']} | {row['NIVEL_TECNOLOGICO']} | {row['DEPARTAMENTO']}"
                    if row.get('MUNICIPIO'):
                        label += f" - {row['MUNICIPIO']}"
                    if row.get('CAMPANIA'):
                        label += f" ({row['CAMPANIA']})"
                    opciones[label] = int(row['ID'])

                sel_label = st.selectbox("Seleccionar plantilla para editar o eliminar",
                                         list(opciones.keys()), key="ref_sel_plantilla")
                sel_id = opciones[sel_label]

                c1, c2 = st.columns(2)
                with c1:
                    if st.button("✏️ Editar seleccionada", use_container_width=True):
                        st.session_state.ref_modo = 'editar'
                        st.session_state.ref_edit_id = sel_id
                        rerun()
                with c2:
                    if st.button("🗑️ Eliminar seleccionada", use_container_width=True):
                        st.session_state.ref_modo = 'confirmar_eliminar'
                        st.session_state.ref_edit_id = sel_id
                        rerun()
        return

    # ═══════════════════════════════════════════════════════
    # 2. MODO CONFIRMAR ELIMINACIÓN
    # ═══════════════════════════════════════════════════════
    if st.session_state.ref_modo == 'confirmar_eliminar':
        plantilla_del = global_db.obtener_plantilla_por_id(st.session_state.ref_edit_id)
        nombre_del = plantilla_del.get('campania', f"ID {st.session_state.ref_edit_id}") if plantilla_del else "la plantilla"
        st.warning(f"⚠️ ¿Eliminar **{nombre_del}**? Esta acción no se puede deshacer.")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("✅ Sí, eliminar", type="primary", key="conf_del_si"):
                if global_db.eliminar_plantilla(st.session_state.ref_edit_id):
                    st.success("✅ Plantilla eliminada.")
                else:
                    st.error("❌ No se pudo eliminar.")
                st.session_state.ref_modo = 'listado'
                st.session_state.ref_edit_id = None
                rerun()
        with c2:
            if st.button("❌ Cancelar", key="conf_del_no"):
                st.session_state.ref_modo = 'listado'
                st.session_state.ref_edit_id = None
                rerun()
        return

    # ═══════════════════════════════════════════════════════
    # 3. MODO FORMULARIO (Crear o Editar)
    # ═══════════════════════════════════════════════════════
    plantilla = None
    if st.session_state.ref_modo == 'editar' and st.session_state.ref_edit_id:
        plantilla = global_db.obtener_plantilla_por_id(st.session_state.ref_edit_id)
        if not plantilla:
            st.error("Plantilla no encontrada.")
            st.session_state.ref_modo = 'listado'
            rerun()

    def val(campo, default=''):
        if plantilla and campo in plantilla and plantilla[campo] is not None:
            return plantilla[campo]
        return default

    def val_num(campo, default=0.0):
        if plantilla and campo in plantilla and plantilla[campo] is not None:
            return float(plantilla[campo])
        return default

    if st.session_state.ref_modo == 'crear':
        st.subheader("➕ Nueva plantilla de costos")
    else:
        st.subheader("✏️ Editar plantilla de costos")

    # --- Campos maestros LIBRES (texto) ---
    cultivo_default = global_db.obtener_nombre_cultivo(plantilla['cultivo_id']) if plantilla else ""
    variedad_default = global_db.obtener_nombre_variedad(plantilla.get('variedad_id')) if plantilla else ""
    nivel_default = global_db.obtener_nombre_nivel_tecnologico(plantilla['nivel_tecnologico_id']) if plantilla else ""
    depto_default = global_db.obtener_nombre_departamento(plantilla['departamento_id']) if plantilla else ""
    muni_default = global_db.obtener_nombre_municipio(plantilla.get('municipio_id')) if plantilla else ""

    cultivo_nombre = st.text_input("Cultivo *", value=cultivo_default, key="ref_cultivo")
    variedad_nombre = st.text_input("Variedad (opcional)", value=variedad_default, key="ref_variedad",
                                    help="Dejar en blanco si no aplica")
    nivel_nombre = st.text_input("Nivel tecnológico *", value=nivel_default, key="ref_nivel")
    depto_nombre = st.text_input("Departamento *", value=depto_default, key="ref_depto")
    muni_nombre = st.text_input("Municipio (opcional)", value=muni_default, key="ref_municipio",
                                 help="Dejar en blanco si no aplica")

    # --- Campos descriptivos y numéricos ---
    campania = st.text_input("Campaña (ej. 2022/2023)", value=val('campania', ''), key="ref_campania")

    c1, c2 = st.columns(2)
    with c1:
        rend_sp = st.number_input("Rendimiento SIN proyecto (kg/ha)", min_value=0.0, step=100.0,
                                   value=val_num('rendimiento_sp'), key="ref_rend_sp")
        perd_sp = st.number_input("% Pérdidas SIN proyecto", min_value=0.0, max_value=100.0, step=0.5,
                                   value=val_num('perdidas_sp_pct'), key="ref_perd_sp")
    with c2:
        rend_cp = st.number_input("Rendimiento CON proyecto (kg/ha)", min_value=0.0, step=100.0,
                                   value=val_num('rendimiento_cp'), key="ref_rend_cp")
        perd_cp = st.number_input("% Pérdidas CON proyecto", min_value=0.0, max_value=100.0, step=0.5,
                                   value=val_num('perdidas_cp_pct'), key="ref_perd_cp")

    precio = st.number_input("Precio referencial (Bs/ton)", min_value=0.0, step=100.0,
                              value=val_num('precio_ref_bs_ton'), key="ref_precio")
    fuente = st.text_input("Fuente de datos", value=val('fuente', 'Manual'), key="ref_fuente")

    activo = st.checkbox("Plantilla activa", value=bool(val('activo', 1)), key="ref_activo") if plantilla else True

    # --- Referencia rápida de existentes (colapsable) ---
    with st.expander("📎 Ver valores ya registrados (solo referencia)"):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.caption("Cultivos existentes")
            st.write(", ".join([c['nombre'] for c in global_db.obtener_cultivos()]) or "—")
        with c2:
            st.caption("Departamentos existentes")
            st.write(", ".join([d['nombre'] for d in global_db.obtener_departamentos()]) or "—")
        with c3:
            st.caption("Niveles tecnológicos existentes")
            st.write(", ".join([n['nombre'] for n in global_db.obtener_niveles_tecnologicos()]) or "—")

    # --- Botones ---
    btn_col1, btn_col2, btn_col3 = st.columns([1, 1, 1])
    with btn_col1:
        guardar = st.button("💾 Guardar", type="primary", use_container_width=True)
    with btn_col2:
        cancelar = st.button("❌ Cancelar", use_container_width=True)
    with btn_col3:
        eliminar = st.button("🗑️ Eliminar", type="secondary", use_container_width=True) if plantilla else False

    if cancelar:
        st.session_state.ref_modo = 'listado'
        st.session_state.ref_edit_id = None
        rerun()

    # --- Lógica GUARDAR ---
    if guardar:
        errores = []
        if not cultivo_nombre.strip():
            errores.append("Cultivo es obligatorio")
        if not nivel_nombre.strip():
            errores.append("Nivel tecnológico es obligatorio")
        if not depto_nombre.strip():
            errores.append("Departamento es obligatorio")
        if errores:
            for e in errores:
                st.error(f"❌ {e}")
            return

        # Resolver IDs (crear si no existen)
        try:
            cultivo_id = global_db.obtener_o_crear_cultivo(cultivo_nombre)
            variedad_id = global_db.obtener_o_crear_variedad(variedad_nombre, cultivo_id)
            nivel_id = global_db.obtener_o_crear_nivel_tecnologico(nivel_nombre)
            depto_id = global_db.obtener_o_crear_departamento(depto_nombre)
            municipio_id = global_db.obtener_o_crear_municipio(muni_nombre, depto_id)
        except Exception as e:
            st.error(f"Error creando catálogos maestros: {e}")
            return

        # Validar duplicados
        conn = global_db._get_conn()
        if variedad_id is None and municipio_id is None:
            query = """SELECT COUNT(*) FROM plantilla_costos
                       WHERE cultivo_id = ? AND nivel_tecnologico_id = ?
                         AND variedad_id IS NULL AND departamento_id = ?
                         AND municipio_id IS NULL AND campania = ?"""
            params = (cultivo_id, nivel_id, depto_id, campania.strip())
        elif variedad_id is None:
            query = """SELECT COUNT(*) FROM plantilla_costos
                       WHERE cultivo_id = ? AND nivel_tecnologico_id = ?
                         AND variedad_id IS NULL AND departamento_id = ?
                         AND municipio_id = ? AND campania = ?"""
            params = (cultivo_id, nivel_id, depto_id, municipio_id, campania.strip())
        elif municipio_id is None:
            query = """SELECT COUNT(*) FROM plantilla_costos
                       WHERE cultivo_id = ? AND nivel_tecnologico_id = ?
                         AND variedad_id = ? AND departamento_id = ?
                         AND municipio_id IS NULL AND campania = ?"""
            params = (cultivo_id, nivel_id, variedad_id, depto_id, campania.strip())
        else:
            query = """SELECT COUNT(*) FROM plantilla_costos
                       WHERE cultivo_id = ? AND nivel_tecnologico_id = ?
                         AND variedad_id = ? AND departamento_id = ?
                         AND municipio_id = ? AND campania = ?"""
            params = (cultivo_id, nivel_id, variedad_id, depto_id, municipio_id, campania.strip())

        if st.session_state.ref_modo == 'editar' and st.session_state.ref_edit_id:
            query += " AND id != ?"
            params += (st.session_state.ref_edit_id,)

        count = conn.execute(query, params).fetchone()[0]
        conn.close()

        if count > 0:
            st.error("⚠️ Ya existe una plantilla con la misma combinación de cultivo, variedad, nivel, ubicación y campaña.")
            return

        datos = {
            'cultivo_id': cultivo_id,
            'variedad_id': variedad_id,
            'nivel_tecnologico_id': nivel_id,
            'departamento_id': depto_id,
            'municipio_id': municipio_id,
            'campania': campania.strip(),
            'rendimiento_sp': rend_sp,
            'rendimiento_cp': rend_cp,
            'precio_ref_bs_ton': precio,
            'perdidas_sp_pct': perd_sp,
            'perdidas_cp_pct': perd_cp,
            'fuente': fuente,
            'activo': 1 if activo else 0
        }

        if st.session_state.ref_modo == 'crear':
            new_id = global_db.agregar_plantilla(datos)
            if new_id:
                st.success("✅ Plantilla creada correctamente.")
                st.session_state.ref_modo = 'listado'
                st.session_state.ref_edit_id = None
                rerun()
            else:
                st.error("❌ Error al crear la plantilla.")
        else:
            if global_db.actualizar_plantilla(st.session_state.ref_edit_id, datos):
                st.success("✅ Plantilla actualizada correctamente.")
                st.session_state.ref_modo = 'listado'
                st.session_state.ref_edit_id = None
                rerun()
            else:
                st.error("❌ Error al actualizar la plantilla.")

    if eliminar:
        st.session_state.ref_modo = 'confirmar_eliminar'
        rerun()
# ------------------------------------------------------------
# RENDER: Asignación y Cálculo (con filtrado inteligente por cultivo)
# ------------------------------------------------------------
def render_asignacion_calculo():
    st.header("⚙️ Asignación de Conceptos y Cálculo")
    global_db = st.session_state.global_db
    proyecto_db = st.session_state.proyecto_db
    calc = st.session_state.calculadora

    df_plantillas = global_db.listar_plantillas()
    if df_plantillas.empty:
        st.warning("No hay plantillas. Primero agregue una plantilla.")
        return

    plantilla_opts = df_plantillas.apply(
        lambda row: f"{row['CULTIVO']} ({row['NIVEL_TECNOLOGICO']}) - {row['DEPARTAMENTO']}",
        axis=1
    ).tolist()
    plantilla_ids = df_plantillas['ID'].tolist()
    plantilla_map = dict(zip(plantilla_opts, plantilla_ids))

    sel_label = st.selectbox("Seleccionar plantilla", plantilla_opts, key="calc_plantilla")
    plantilla_id = plantilla_map[sel_label]

    # Obtener nombre del cultivo para filtrado inteligente
    plantilla = global_db.obtener_plantilla_por_id(plantilla_id)
    cultivo_nombre = global_db.obtener_nombre_cultivo(plantilla['cultivo_id']) if plantilla else ""

    # ═══════════════════════════════════════════════════════
    # SECCIÓN 1: FILTRADO INTELIGENTE POR CULTIVO Y LUGAR
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("🔍 Conceptos sugeridos para esta plantilla")

    fc1, fc2, fc3 = st.columns([2, 2, 1])
    with fc1:
        filtrar_cultivo = st.checkbox(
            f"Mostrar solo conceptos de '{cultivo_nombre or '...'}'",
            value=True,
            key="filtrar_cultivo"
        )
    with fc2:
        lugar_filtro = st.text_input(
            "Filtrar por lugar (opcional)",
            value="",
            key="filtro_lugar",
            help="Ej: Ayata, Valle Central, Mizque, etc."
        )
    with fc3:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        buscar = st.button("🔎 Buscar", key="btn_buscar_conceptos", use_container_width=True)

    # Determinar dataset según filtros
    if filtrar_cultivo and cultivo_nombre:
        if lugar_filtro.strip():
            df_conceptos_filtrados = global_db.obtener_conceptos_por_cultivo_y_lugar(
                cultivo_nombre, lugar_filtro.strip()
            )
            st.caption(f"Filtrado por cultivo **{cultivo_nombre}** + lugar **{lugar_filtro.strip()}** — {len(df_conceptos_filtrados)} conceptos")
        else:
            df_conceptos_filtrados = global_db.obtener_conceptos_por_cultivo(cultivo_nombre)
            st.caption(f"Filtrado por cultivo **{cultivo_nombre}** — {len(df_conceptos_filtrados)} conceptos")
    else:
        df_conceptos_filtrados = global_db.listar_conceptos()
        st.caption(f"Mostrando todos los conceptos — {len(df_conceptos_filtrados)} conceptos")

    # Preview de conceptos sugeridos
    if not df_conceptos_filtrados.empty:
        with st.expander("📋 Ver lista de conceptos sugeridos", expanded=False):
            vis_cols = ['CONCEPTO', 'UNIDAD', 'PRECIO_UNITARIO', 'CATEGORIA', 'CLASE_RPC', 'LUGAR', 'CULTIVO']
            vis_exist = [c for c in vis_cols if c in df_conceptos_filtrados.columns]
            st.dataframe(df_conceptos_filtrados[vis_exist], use_container_width=True, hide_index=True)
    else:
        st.warning("No se encontraron conceptos con los filtros aplicados. Desactive 'Mostrar solo conceptos de este cultivo' para ver todos.")

    # ═══════════════════════════════════════════════════════
    # SECCIÓN 2: ASIGNACIÓN POR CATEGORÍA
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("✅ Asignar conceptos a la plantilla")

    # Conceptos ya asignados
    df_asignados = global_db.obtener_conceptos_de_plantilla(plantilla_id)
    conceptos_asignados = df_asignados['CONCEPTO'].tolist() if not df_asignados.empty else []

    # Construir opciones por categoría (filtradas o todas)
    categorias = [
        ('INSUMOS', '📦 Insumos'),
        ('MANO DE OBRA', '👷 Mano de Obra'),
        ('SERVICIOS NO PERSONALES', '🚜 Servicios No Personales'),
        ('OTROS COSTOS DIRECTOS', '📋 Otros Costos Directos')
    ]

    if filtrar_cultivo and not df_conceptos_filtrados.empty:
        conceptos_por_cat = {}
        for cat, _ in categorias:
            mask = df_conceptos_filtrados['CATEGORIA'].str.upper().str.strip() == cat
            conceptos_por_cat[cat] = sorted(df_conceptos_filtrados[mask]['CONCEPTO'].dropna().unique().tolist())
    else:
        conceptos_por_cat = global_db.get_categoria_conceptos_dict()

    col1, col2 = st.columns(2)
    seleccionados_totales = []
    resumen_cats = []

    with col1:
        for cat, label in categorias[:2]:
            opciones = conceptos_por_cat.get(cat, [])
            default = [c for c in conceptos_asignados if c in opciones]
            seleccionados = st.multiselect(
                label, options=opciones, default=default,
                key=f"ms_{cat}_{plantilla_id}"
            )
            seleccionados_totales.extend(seleccionados)
            if seleccionados:
                resumen_cats.append(f"{label.split()[0]}: {len(seleccionados)}")

    with col2:
        for cat, label in categorias[2:]:
            opciones = conceptos_por_cat.get(cat, [])
            default = [c for c in conceptos_asignados if c in opciones]
            seleccionados = st.multiselect(
                label, options=opciones, default=default,
                key=f"ms_{cat}_{plantilla_id}"
            )
            seleccionados_totales.extend(seleccionados)
            if seleccionados:
                resumen_cats.append(f"{label.split()[0]}: {len(seleccionados)}")

    if seleccionados_totales:
        st.success(f"**Total seleccionados:** {len(seleccionados_totales)} ({' | '.join(resumen_cats)})")
    else:
        st.warning("⚠️ No ha seleccionado ningún concepto.")

    if st.button("💾 Guardar asignación de conceptos", key="btn_guardar_asig"):
        ids_seleccionados = []
        for nombre in seleccionados_totales:
            concepto = global_db.obtener_concepto_por_nombre(nombre)
            if concepto:
                ids_seleccionados.append(concepto['id'])
        if global_db.asignar_conceptos_a_plantilla(plantilla_id, ids_seleccionados):
            st.success("✅ Asignación guardada correctamente")
            rerun()
        else:
            st.error("❌ Error al guardar")

    # ═══════════════════════════════════════════════════════
    # SECCIÓN 3: TABLA EDITABLE DE CONCEPTOS ASIGNADOS
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("Ajustar cantidades y precios de conceptos asignados")
    st.caption("Edite CANTIDAD (SP), CANTIDAD_CP y PRECIO_OVERRIDE. Deje PRECIO_OVERRIDE en 0 para usar el precio de catalogo.")

    df_asignados = global_db.obtener_conceptos_de_plantilla(plantilla_id)
    if not df_asignados.empty:
        cols_edit = ['REL_ID', 'CONCEPTO', 'UNIDAD', 'CANTIDAD', 'CANTIDAD_CP',
                     'PRECIO_UNITARIO', 'PRECIO_OVERRIDE', 'CATEGORIA', 'CLASE_RPC', 'OBSERVACIONES']
        cols_existen = [c for c in cols_edit if c in df_asignados.columns]
        df_edit = df_asignados[cols_existen].copy()

        # Asegurar que CANTIDAD_CP existe (puede no venir de la BD)
        if 'CANTIDAD_CP' not in df_edit.columns:
            df_edit['CANTIDAD_CP'] = pd.NA

        # Asegurar PRECIO_OVERRIDE
        if 'PRECIO_OVERRIDE' not in df_edit.columns:
            df_edit['PRECIO_OVERRIDE'] = 0.0
        else:
            df_edit['PRECIO_OVERRIDE'] = pd.to_numeric(df_edit['PRECIO_OVERRIDE'], errors='coerce').fillna(0)

        edited_df = st.data_editor(
            df_edit,
            use_container_width=True,
            num_rows="fixed",
            key=f"editor_asignados_{plantilla_id}",
            column_config={
                "REL_ID": st.column_config.NumberColumn("ID Rel", disabled=True),
                "CONCEPTO": st.column_config.TextColumn("Concepto", disabled=True),
                "UNIDAD": st.column_config.TextColumn("Unidad", disabled=True),
                "CANTIDAD": st.column_config.NumberColumn("Cantidad SP", min_value=0.0, step=0.1,
                    help="Cantidad SIN proyecto. Editar aqui modifica la base SP."),
                "CANTIDAD_CP": st.column_config.NumberColumn("Cantidad CP", min_value=0.0, step=0.1,
                    help="Cantidad CON proyecto. Dejar VACIO para usar factor automatico R-115. 0 = eliminar insumo."),
                "PRECIO_UNITARIO": st.column_config.NumberColumn("Precio Ref.", disabled=True, format="%.2f"),
                "PRECIO_OVERRIDE": st.column_config.NumberColumn("Precio Override", min_value=0.0, step=1.0,
                    help="Sobrescribe el precio de catalogo para este calculo. 0 = usar referencia."),
                "CATEGORIA": st.column_config.TextColumn("Categoria", disabled=True),
                "CLASE_RPC": st.column_config.TextColumn("RPC", disabled=True),
                "OBSERVACIONES": st.column_config.TextColumn("Observaciones", disabled=True)
            }
        )

        c1, c2 = st.columns([1, 3])
        with c1:
            if st.button("Guardar ajustes en plantilla", key="btn_guardar_ajustes", type="primary"):
                # Guardar CANTIDAD, CANTIDAD_CP y PRECIO_OVERRIDE en global_db
                if global_db.guardar_conceptos_plantilla_bulk(plantilla_id, edited_df):
                    st.success("Ajustes guardados en la plantilla global")
                    rerun()
                else:
                    st.error("Error al guardar ajustes")
        with c2:
            st.info("Los ajustes se guardan en la plantilla global y afectan todos los calculos futuros.")
    else:
        st.info("Esta plantilla aun no tiene conceptos asignados.")

# ------------------------------------------------------------------------------
# SECCION 3.5: COSTOS ADICIONALES Y VARIABLES
# ------------------------------------------------------------------------------
    st.markdown("---")
    st.subheader("Costos Indirectos")
    st.caption("Gastos generales, interes, depreciacion y puede agregar otros que considere necesario.")

    ss_key_ca = f"costos_adicionales_{plantilla_id}"
    if ss_key_ca not in st.session_state:
        st.session_state[ss_key_ca] = pd.DataFrame({
            'CATEGORIA': ['Gastos Generales', 'Interes',
                          'Depreciacion Herramientas'],
            'TIPO': ['porcentaje', 'porcentaje', 'fijo'],
            'VALOR': [5.0, 2.5, 110.5],
            'BASE_CALCULO': ['directos', 'directos', 'ninguna'],
            'DESCRIPCION': [
                'Gastos generales (5%)',
                'Interés(50% de Gastosgenerales)',
                'Depreciacion anual de herramientas'
            ]
        })

    df_ca = st.session_state[ss_key_ca]
    edited_ca = st.data_editor(
        df_ca,
        use_container_width=True,
        num_rows="dynamic",
        key=f"editor_ca_{plantilla_id}",
        column_config={
            'CATEGORIA': st.column_config.TextColumn('Categoria'),
            'TIPO': st.column_config.SelectboxColumn('Tipo', options=['fijo', 'porcentaje']),
            'VALOR': st.column_config.NumberColumn('Valor', min_value=0.0, step=0.01),
            'BASE_CALCULO': st.column_config.SelectboxColumn('Base', options=['ninguna', 'directos', 'total']),
            'DESCRIPCION': st.column_config.TextColumn('Descripcion')
        }
    )
    st.session_state[ss_key_ca] = edited_ca
    st.info("Tipo fijo -> Bs/Ha directos. Tipo porcentaje -> % sobre la base seleccionada. Los porcentajes se prorratean por RPC.")
    st.markdown("---")

# ------------------------------------------------------------------------------
# SECCION 4: CALCULO DE COSTOS
# ------------------------------------------------------------------------------
    st.subheader("Calcular costos de produccion")

    # Datos de la plantilla para ingresos
    plantilla = global_db.obtener_plantilla_por_id(plantilla_id)
    rend_sp = float(plantilla.get('rendimiento_sp', 0) or 0) if plantilla else 0
    rend_cp = float(plantilla.get('rendimiento_cp', 0) or 0) if plantilla else 0
    precio_ton = float(plantilla.get('precio_ref_bs_ton', 0) or 0) if plantilla else 0
    perd_sp = float(plantilla.get('perdidas_sp_pct', 0) or 0) if plantilla else 0
    perd_cp = float(plantilla.get('perdidas_cp_pct', 0) or 0) if plantilla else 0

    col_sup1, col_sup2 = st.columns(2)
    with col_sup1:
        sup_sp_calc = st.number_input("Superficie SP (Ha)", min_value=0.0, value=1.0, step=0.5, key="calc_sup_sp")
    with col_sup2:
        sup_cp_calc = st.number_input("Superficie CP (Ha)", min_value=0.0, value=1.0, step=0.5, key="calc_sup_cp")

    col_calc1, col_calc2 = st.columns(2)
    with col_calc1:
        btn_calcular = st.button("Calcular y mostrar resultados", type="primary", key="btn_calcular")
    with col_calc2:
        btn_guardar = st.button("Guardar calculo en proyecto", type="secondary", key="btn_guardar")

    def realizar_calculo_completo():
        # Obtener conceptos con ajustes guardados (incluye CANTIDAD, CANTIDAD_CP, PRECIO_OVERRIDE)
        df_calc = global_db.obtener_conceptos_de_plantilla_para_calculo(plantilla_id)
        if df_calc.empty:
            st.error("No hay conceptos asignados para calcular.")
            return None

        # Asegurar tipos numericos
        for col in ['CANTIDAD', 'PRECIO_UNITARIO', 'PRECIO_OVERRIDE']:
            if col in df_calc.columns:
                df_calc[col] = pd.to_numeric(df_calc[col], errors='coerce').fillna(0)
        
        # CANTIDAD_CP: NaN = automatico (NO hacer fillna)
        if 'CANTIDAD_CP' in df_calc.columns:
            df_calc['CANTIDAD_CP'] = pd.to_numeric(df_calc['CANTIDAD_CP'], errors='coerce')

        # Costos adicionales
        df_ca = st.session_state.get(ss_key_ca, pd.DataFrame())

        # Ejecutar calculo completo
        resultado = calc.calcular_costos_completos(
            df_conceptos=df_calc,
            df_adicionales=df_ca,
            sup_sp=sup_sp_calc,
            sup_cp=sup_cp_calc,
            rend_sp=rend_sp,
            rend_cp=rend_cp,
            precio_ton=precio_ton,
            perd_sp_pct=perd_sp,
            perd_cp_pct=perd_cp
        )
        return resultado

    def mostrar_resultados(res, guardado=False):
        if guardado:
            st.success("Calculo guardado correctamente en el proyecto")
        else:
            st.success("Calculo completado (vista previa)")

        # -- Metricas principales --
        st.markdown("### Resumen por Hectarea")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Costo Conceptos SP", f"{res.conceptos_sp:,.2f} Bs/Ha")
            st.metric("Costo Adicionales SP", f"{res.adicionales_sp:,.2f} Bs/Ha")
            st.metric("Costo Total SP", f"{res.total_sp:,.2f} Bs/Ha")
        with col2:
            st.metric("Costo Conceptos CP", f"{res.conceptos_cp:,.2f} Bs/Ha")
            st.metric("Costo Adicionales CP", f"{res.adicionales_cp:,.2f} Bs/Ha")
            st.metric("Costo Total CP", f"{res.total_cp:,.2f} Bs/Ha")
        with col3:
            delta_costo = res.total_cp - res.total_sp
            delta_pct = (delta_costo / res.total_sp * 100) if res.total_sp > 0 else 0
            st.metric("Delta Costo CP vs SP", f"{delta_costo:,.2f} Bs/Ha", f"{delta_pct:+.1f}%")

        # -- Desglose RPC --
        with st.expander("Desglose RPC por Hectarea"):
            rpc_data = {
                'Clasificacion': ['Bienes Transables (BT)', 'Bienes No Transables (BNT)', 
                                  'Mano de Obra No Calificada (MONR)', 'Mano de Obra No Calificada Urbana (MONU)',
                                  'Mano de Obra Semicalificada (MOS)', 'Mano de Obra Calificada (MOC)', 'TOTAL'],
                'SP (Bs/Ha)': [res.bt_sp, res.bnt_sp, res.monr_sp, res.monu_sp, res.mos_sp, res.moc_sp, res.total_sp],
                'CP (Bs/Ha)': [res.bt_cp, res.bnt_cp, res.monr_cp, res.monu_cp, res.mos_cp, res.moc_cp, res.total_cp],
                'Delta (Bs/Ha)': [res.bt_cp-res.bt_sp, res.bnt_cp-res.bnt_sp, res.monr_cp-res.monr_sp,
                              res.monu_cp-res.monu_sp, res.mos_cp-res.mos_sp, res.moc_cp-res.moc_sp, delta_costo]
            }
            st.dataframe(pd.DataFrame(rpc_data), use_container_width=True, hide_index=True)

        # -- Costos Indirectos Detalle --
        with st.expander("Detalle de Costos Indirectos"):
            if res.detalle_adicionales:
                ca_data = []
                for item in res.detalle_adicionales:
                    ca_data.append({
                        'Categoria': item.categoria,
                        'Descripcion': item.descripcion,
                        'Tipo': item.tipo.upper(),
                        'Valor': item.valor_orig,
                        'Base': item.base,
                        'SP (Bs/Ha)': item.monto_sp,
                        'CP (Bs/Ha)': item.monto_cp
                    })
                st.dataframe(pd.DataFrame(ca_data), use_container_width=True, hide_index=True)
            else:
                st.info("No hay costos adicionales configurados.")

        # -- Totales por Superficie --
        st.markdown("### Totales del Proyecto (por Superficie)")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric(f"Costo Total SP ({res.sup_sp} Ha)", f"{res.total_proy_sp:,.2f} Bs")
            st.metric("Ingreso SP", f"{res.ingreso_sp:,.2f} Bs")
            st.metric("Utilidad SP", f"{res.utilidad_sp:,.2f} Bs")
        with c2:
            st.metric(f"Costo Total CP ({res.sup_cp} Ha)", f"{res.total_proy_cp:,.2f} Bs")
            st.metric("Ingreso CP", f"{res.ingreso_cp:,.2f} Bs")
            st.metric("Utilidad CP", f"{res.utilidad_cp:,.2f} Bs")
        with c3:
            st.metric("Relacion B/C SP", f"{res.relacion_bc_sp:.2f}")
            st.metric("Relacion B/C CP", f"{res.relacion_bc_cp:.2f}")
            if res.relacion_bc_cp > res.relacion_bc_sp:
                st.success("El proyecto mejora la relacion B/C")
            elif res.relacion_bc_cp < res.relacion_bc_sp:
                st.warning("El proyecto reduce la relacion B/C")
        return res

    # --- Boton Calcular ---
    if btn_calcular:
        resultado = realizar_calculo_completo()
        if resultado:
            mostrar_resultados(resultado, guardado=False)

    # --- Boton Guardar ---
    if btn_guardar:
        resultado = realizar_calculo_completo()
        if resultado:
            res = mostrar_resultados(resultado, guardado=True)

            # Preparar datos para guardar en proyecto.db
            det_conceptos = [
                {
                    'CONCEPTO': c.concepto, 'UNIDAD': c.unidad, 'CATEGORIA': c.categoria,
                    'CLASE_RPC': c.clase_rpc, 'CANTIDAD': c.cantidad_sp, 'CANTIDAD_CP': c.cantidad_cp,
                    'PRECIO_UNITARIO': c.precio_efectivo, 'TOTAL_SP': c.total_sp, 'TOTAL_CP': c.total_cp,
                    'OBSERVACIONES': c.observaciones
                }
                for c in resultado.detalle_conceptos
            ]
            det_adicionales = [
                {
                    'CATEGORIA': a.categoria, 'DESCRIPCION': a.descripcion, 'TIPO': a.tipo,
                    'VALOR_ORIG': a.valor_orig, 'BASE': a.base,
                    'MONTO_SP': a.monto_sp, 'MONTO_CP': a.monto_cp
                }
                for a in resultado.detalle_adicionales
            ]

            datos_guardar = {
                'plantilla_id': plantilla_id,
                'sup_sp_ha': sup_sp_calc,
                'sup_cp_ha': sup_cp_calc,
                'costo_conceptos_sp': resultado.conceptos_sp,
                'costo_conceptos_cp': resultado.conceptos_cp,
                'costo_adicionales_sp': resultado.adicionales_sp,
                'costo_adicionales_cp': resultado.adicionales_cp,
                'costo_total_sp': resultado.total_sp,
                'costo_total_cp': resultado.total_cp,
                'bt_sp': resultado.bt_sp, 'bnt_sp': resultado.bnt_sp, 'monr_sp': resultado.monr_sp,
                'monu_sp': resultado.monu_sp, 'mos_sp': resultado.mos_sp, 'moc_sp': resultado.moc_sp,
                'bt_cp': resultado.bt_cp, 'bnt_cp': resultado.bnt_cp, 'monr_cp': resultado.monr_cp,
                'monu_cp': resultado.monu_cp, 'mos_cp': resultado.mos_cp, 'moc_cp': resultado.moc_cp,
                'total_proy_sp': resultado.total_proy_sp,
                'total_proy_cp': resultado.total_proy_cp,
                'ingreso_sp': resultado.ingreso_sp,
                'ingreso_cp': resultado.ingreso_cp,
                'utilidad_sp': resultado.utilidad_sp,
                'utilidad_cp': resultado.utilidad_cp,
                'relacion_bc_sp': resultado.relacion_bc_sp,
                'relacion_bc_cp': resultado.relacion_bc_cp,
                'detalle_conceptos': det_conceptos,
                'detalle_adicionales': det_adicionales
            }

            if proyecto_db.guardar_costos_calculados_v2(datos_guardar):
                # AUTO-EXPORTAR a tabla persistente de cultivos
                ok, msg = _agregar_cultivo_a_proyecto_db(
                    proyecto_db, global_db, plantilla_id, resultado, sup_sp_calc, sup_cp_calc
                )
                if ok:
                    st.info(f"🌾 '{msg}' agregado automáticamente a la tabla de cultivos del proyecto.")
                else:
                    st.warning(f"⚠️ No se pudo auto-agregar cultivo: {msg}")
                st.balloons()
            else:
                st.error("Error al guardar en la base de datos del proyecto")

    # ═══════════════════════════════════════════════════════
    # SECCIÓN 5: GENERACIÓN DE REPORTE EXCEL
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("📄 Generación de Reporte Excel")

    config = st.session_state.get('config_proyecto')

    with st.expander("⚙️ Configuración del Reporte", expanded=False):
        col_r1, col_r2 = st.columns(2)
        with col_r1:
            rpt_depto = st.text_input(
                "Departamento",
                value=config.depto if config else "",
                key="rpt_depto_ac"
            )
            rpt_proy = st.text_input(
                "Nombre del Proyecto",
                value=config.nombre if config else "",
                key="rpt_proy_ac"
            )
            rpt_variedad = st.text_input("Variedad", value="", key="rpt_variedad_ac")
        with col_r2:
            rpt_mun = st.text_input(
                "Municipio",
                value=config.municipio if config else "",
                key="rpt_mun_ac"
            )
            tc_default = 6.96
            if config and hasattr(config, 'rpc') and isinstance(config.rpc, dict):
                tc_default = config.rpc.get('divisa', 6.96)
            rpt_tc = st.number_input(
                "Tipo de Cambio (Bs/US$)",
                value=float(config.tipo_cambio if config else ""),   #Esta linea por confirmar
                step=0.01,
                format="%.2f",
                key="rpt_tc_ac"
            )
            c1, c2 = st.columns(2)
            with c1:
                rpt_mes_siembra = st.text_input("Mes Siembra", value="", key="rpt_siembra_ac")
            with c2:
                rpt_mes_cosecha = st.text_input("Mes Cosecha", value="", key="rpt_cosecha_ac")

    if st.button("📥 Generar Reporte Detallado", type="primary", key="btn_gen_rpt_ac"):
        with st.spinner("Generando reporte..."):
            costos_guardados = proyecto_db.obtener_costos_calculados(plantilla_id)
            if not costos_guardados:
                st.error("No hay cálculo guardado. Primero calcule y guarde en esta misma pestaña.")
            else:
                detalle_conceptos = costos_guardados.get('detalle_conceptos', [])
                detalle_adicionales = costos_guardados.get('detalle_adicionales', [])
                df_conceptos_rpt = pd.DataFrame(detalle_conceptos) if detalle_conceptos else pd.DataFrame()
                df_adicionales_rpt = pd.DataFrame(detalle_adicionales) if detalle_adicionales else pd.DataFrame()

                required_cols = ['CONCEPTO', 'UNIDAD', 'CANTIDAD', 'CANTIDAD_CP',
                                 'PRECIO_UNITARIO', 'CATEGORIA', 'CLASE_RPC', 'OBSERVACIONES']
                for col in required_cols:
                    if col not in df_conceptos_rpt.columns:
                        df_conceptos_rpt[col] = '' if col in ['CONCEPTO', 'UNIDAD', 'OBSERVACIONES'] else 0

                cultivo_nombre = global_db.obtener_nombre_cultivo(plantilla['cultivo_id'])
                cultivo_data_rpt = {
                    'Cultivo': cultivo_nombre,
                    'Rendimiento_SP': plantilla.get('rendimiento_sp', 0),
                    'Rendimiento_CP': plantilla.get('rendimiento_cp', 0),
                    'Precio_Bs_Ton': plantilla.get('precio_ref_bs_ton', 0),
                    'Perdidas_SP_pct': plantilla.get('perdidas_sp_pct', 0),
                    'Perdidas_CP_pct': plantilla.get('perdidas_cp_pct', 0),
                    'CostoTotal_SP': costos_guardados['costo_total_sp'],
                    'CostoTotal_CP': costos_guardados['costo_total_cp']
                }

                gen = st.session_state.reporte_gen
                ruta = gen.generar_reporte_detallado_v2(
                    cultivo_data=cultivo_data_rpt,
                    df_conceptos=df_conceptos_rpt,
                    df_adicionales=df_adicionales_rpt,
                    depto=rpt_depto,
                    municipio=rpt_mun,
                    proyecto_nombre=rpt_proy,
                    tipo_cambio=rpt_tc,
                    variedad=rpt_variedad,
                    mes_siembra=rpt_mes_siembra,
                    mes_cosecha=rpt_mes_cosecha
                )

                if ruta:
                    with open(ruta, "rb") as file:
                        st.download_button(
                            label="⬇️ Descargar Reporte Excel",
                            data=file,
                            file_name=os.path.basename(ruta),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="btn_download_rpt_ac"
                        )
                        # Auto-exportar a tabla de cultivos del proyecto
                        resultado_rpt = realizar_calculo_completo()
                        if resultado_rpt:
                            ok, msg = _agregar_cultivo_a_proyecto_db(
                                proyecto_db, global_db, plantilla_id, resultado_rpt, sup_sp_calc, sup_cp_calc
                            )
                            if ok:
                                st.info(f"🌾 '{msg}' agregado/actualizado en la tabla de cultivos del proyecto.")
                        # Hasta aqui Auto-exportar, se puede eliminar esta seccion

                    st.success(f"✅ Reporte generado: `{os.path.basename(ruta)}`")
                else:
                    st.error("No se pudo generar el reporte.")
# ------------------------------------------------------------
# Reporte Valor Neto
# ------------------------------------------------------------
def generar_reporte_valor_neto(df_cultivos, config, tipo_cambio=6.96):
    """Genera Excel de Valor Neto de la Producción con formato oficial."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Valor_Neto_Produccion"

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Título principal
    ws.merge_cells('A1:J1')
    ws['A1'] = "RESUMEN COSTOS DE PRODUCCION"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f'PROYECTO "{config.nombre if config else "Sin nombre"}"'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    def escribir_tabla(start_row, titulo, df, sup_col, rend_col, perd_col, costo_key, ingreso_key):
        ws.merge_cells(f'A{start_row}:J{start_row}')
        ws[f'A{start_row}'] = titulo
        ws[f'A{start_row}'].font = subtitle_font
        start_row += 2

        headers = [
            "Cultivos", "Superficie (ha)", "Rendimiento (tn/ha)",
            "Pérdida post cosecha (%)", "Precio de venta (Bs/Tn)",
            "Costo/ha (Bs)", "Costo Total (Bs)",
            "Ingreso/ha (Bs)", "Ingreso Total (Bs)",
            "Valor Neto Producción (Bs)"
        ]
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
        start_row += 1

        total_sup = 0.0
        total_costo_total = 0.0
        total_ingreso_total = 0.0
        total_valor_neto = 0.0

        for _, row in df.iterrows():
            sup = float(row[sup_col]) if sup_col in row else 0.0
            rend_kg = float(row[rend_col]) if rend_col in row else 0.0
            rend_tn = rend_kg / 1000.0
            perd = float(row[perd_col]) if perd_col in row else 0.0
            precio_bs = float(row['Precio_Bs_Ton']) if 'Precio_Bs_Ton' in row else 0.0
            costo_ha_bs = float(row[costo_key]) if costo_key in row else 0.0

            costo_total_bs = costo_ha_bs * sup
            ingreso_ha_bs = float(row[ingreso_key]) if ingreso_key in row else 0.0
            ingreso_total_bs = ingreso_ha_bs * sup

            # Lógica del modelo oficial
            valor_neto_bs = (ingreso_total_bs - costo_total_bs) * (1.0 - perd / 100.0) if sup > 0 else 0.0

            values = [
                row['Nombre'], sup, rend_tn, f"{perd:.0f}%", precio_bs,
                costo_ha_bs, costo_total_bs, ingreso_ha_bs, ingreso_total_bs, valor_neto_bs
            ]
            for c_idx, val in enumerate(values, 1):
                cell = ws.cell(row=start_row, column=c_idx, value=val)
                cell.border = thin_border
                if c_idx > 1 and isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                if c_idx == 4:
                    cell.alignment = Alignment(horizontal='center')
            start_row += 1

            if sup > 0:
                total_sup += sup
                total_costo_total += costo_total_bs
                total_ingreso_total += ingreso_total_bs
                total_valor_neto += valor_neto_bs

        # Fila TOTAL
        totales = ["TOTAL", total_sup, None, None, None, None,
                   total_costo_total, None, total_ingreso_total, total_valor_neto]
        for c_idx, val in enumerate(totales, 1):
            cell = ws.cell(row=start_row, column=c_idx, value=val)
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
            if c_idx in (2, 7, 9, 10) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
        return start_row + 1, total_valor_neto

    # Tabla SP
    row_actual = 4
    row_actual, vn_sp = escribir_tabla(
        row_actual,
        "DETERMINACIÓN DEL VALOR NETO DE LA PRODUCCIÓN AGRICOLA EN SITUACIÓN S/PROYECTO",
        df_cultivos, 'Sup_SP_Ha', 'Rend_SP', 'Perd_SP_%', 'CostoTotal_SP', 'Ingreso_SP'
    )

    row_actual += 2

    # Tabla CP
    row_actual, vn_cp = escribir_tabla(
        row_actual,
        "DETERMINACIÓN DEL VALOR NETO DE LA PRODUCCIÓN AGRICOLA EN SITUACIÓN C/PROYECTO",
        df_cultivos, 'Sup_CP_Ha', 'Rend_CP', 'Perd_CP_%', 'CostoTotal_CP', 'Ingreso_CP'
    )

    row_actual += 2
    incremento = vn_cp - vn_sp
    ws.merge_cells(f'A{row_actual}:F{row_actual}')
    ws[f'A{row_actual}'] = "Incremento del Valor Neto de Producción con proyecto"
    ws[f'A{row_actual}'].font = Font(bold=True, size=11)
    ws[f'G{row_actual}'] = "$US."
    ws[f'G{row_actual}'].font = Font(bold=True)
    ws[f'G{row_actual}'].alignment = Alignment(horizontal='right')
    ws[f'H{row_actual}'] = incremento / tipo_cambio
    ws[f'H{row_actual}'].number_format = '#,##0.00'
    ws[f'H{row_actual}'].font = Font(bold=True)
    ws[f'I{row_actual}'] = "Bs."
    ws[f'I{row_actual}'].font = Font(bold=True)
    ws[f'I{row_actual}'].alignment = Alignment(horizontal='right')
    ws[f'J{row_actual}'] = incremento
    ws[f'J{row_actual}'].number_format = '#,##0.00'
    ws[f'J{row_actual}'].font = Font(bold=True)

    # Anchos
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"Valor_Neto_Produccion_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path
# ------------------------------------------------------------
# RENDER: Exportar a Proyecto
def render_exportar_proyecto():
    st.header("🔄 Exportar a Proyecto")
    proyecto_db = st.session_state.proyecto_db
    data_manager = st.session_state.data_manager
    config = st.session_state.get('config_proyecto')
    pm = ProjectManager()
    rutas = pm.get_rutas_activas()

    # ═══════════════════════════════════════════════════════
    # 1. TABLA EDITABLE DE CULTIVOS EN EL PROYECTO
    # ═══════════════════════════════════════════════════════
    df_proyecto = proyecto_db.listar_cultivos_proyecto()

    st.subheader("📋 Cultivos en el Proyecto")
    st.caption("Edite superficies, rendimientos o elimine cultivos. Los cambios se guardan en la base de datos del proyecto.")

    if not df_proyecto.empty:
        cols_vis = [
            'Nombre', 'Codigo', 'Sup_SP_Ha', 'Sup_CP_Ha', 'Rend_SP', 'Rend_CP',
            'Perd_SP_%', 'Perd_CP_%', 'Precio_Bs_Ton', 'CostoTotal_SP', 'CostoTotal_CP',
            'BT_SP', 'BNT_SP', 'MONR_SP', 'MONU_SP', 'MOS_SP', 'MOC_SP',
            'BT_CP', 'BNT_CP', 'MONR_CP', 'MONU_CP', 'MOS_CP', 'MOC_CP',
            'Ingreso_SP', 'Ingreso_CP'
        ]

        edited_df = st.data_editor(
            df_proyecto[cols_vis],
            use_container_width=True,
            num_rows="dynamic",
            key="editor_cultivos_proyecto"
        )

        c1, c2 = st.columns(2)
        with c1:
            if st.button("💾 Guardar cambios en tabla", key="btn_guardar_tabla_exp"):
                try:
                    with proyecto_db._get_conn() as conn:
                        conn.execute("DELETE FROM proyecto_cultivos")
                        for _, row in edited_df.iterrows():
                            if pd.isna(row.get('Nombre')) or str(row.get('Nombre')).strip() == '':
                                continue
                            conn.execute("""
                                INSERT INTO proyecto_cultivos (
                                    nombre, codigo, sup_sp_ha, sup_cp_ha, rend_sp, rend_cp,
                                    perd_sp_pct, perd_cp_pct, precio_bs_ton,
                                    costo_total_sp, costo_total_cp,
                                    bt_sp, bnt_sp, monr_sp, monu_sp, mos_sp, moc_sp,
                                    bt_cp, bnt_cp, monr_cp, monu_cp, mos_cp, moc_cp,
                                    ingreso_sp, ingreso_cp, fecha_agregado
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (
                                str(row['Nombre']), str(row.get('Codigo','')),
                                float(row.get('Sup_SP_Ha',0)), float(row.get('Sup_CP_Ha',0)),
                                float(row.get('Rend_SP',0)), float(row.get('Rend_CP',0)),
                                float(row.get('Perd_SP_%',0)), float(row.get('Perd_CP_%',0)),
                                float(row.get('Precio_Bs_Ton',0)),
                                float(row.get('CostoTotal_SP',0)), float(row.get('CostoTotal_CP',0)),
                                float(row.get('BT_SP',0)), float(row.get('BNT_SP',0)),
                                float(row.get('MONR_SP',0)), float(row.get('MONU_SP',0)),
                                float(row.get('MOS_SP',0)), float(row.get('MOC_SP',0)),
                                float(row.get('BT_CP',0)), float(row.get('BNT_CP',0)),
                                float(row.get('MONR_CP',0)), float(row.get('MONU_CP',0)),
                                float(row.get('MOS_CP',0)), float(row.get('MOC_CP',0)),
                                float(row.get('Ingreso_SP',0)), float(row.get('Ingreso_CP',0)),
                                datetime.now().isoformat()
                            ))
                        conn.commit()
                    st.success("✅ Tabla de cultivos actualizada correctamente.")
                    rerun()
                except Exception as e:
                    st.error(f"❌ Error guardando cambios: {e}")

        with c2:
            if st.button("🗑️ Limpiar todos los cultivos", type="secondary"):
                try:
                    with proyecto_db._get_conn() as conn:
                        conn.execute("DELETE FROM proyecto_cultivos")
                        conn.commit()
                    st.success("🗑️ Tabla de cultivos limpiada.")
                    rerun()
                except Exception as e:
                    st.error(f"❌ Error: {e}")

        # Métricas de superficie acumulada
        total_sup_sp = float(df_proyecto['Sup_SP_Ha'].sum())
        total_sup_cp = float(df_proyecto['Sup_CP_Ha'].sum())
        m1, m2, m3 = st.columns(3)
        m1.metric("Cultivos registrados", len(df_proyecto))
        m2.metric("Superficie SP total", f"{total_sup_sp:.2f} Ha")
        m3.metric("Superficie CP total", f"{total_sup_cp:.2f} Ha")
    else:
        st.info("ℹ️ No hay cultivos en el proyecto. Vaya a **Asignación y Cálculo**, calcule y guarde para agregar cultivos automáticamente.")

    # ═══════════════════════════════════════════════════════
    # 2. REPORTE EXCEL DE VALOR NETO DE LA PRODUCCIÓN
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("📊 Reporte de Valor Neto de la Producción")

    if not df_proyecto.empty:
        tc_vn = st.number_input(
            "Tipo de Cambio (Bs/$US) para reporte",
            value=float(config.tipo_cambio), step=0.01, format="%.2f", key="vn_tc_exp"  #Esta linea por confirmar
        )
        if st.button("📥 Generar Reporte Valor Neto", type="primary"):
            with st.spinner("Generando reporte..."):
                ruta = generar_reporte_valor_neto(df_proyecto, config, tc_vn)
                if ruta:
                    with open(ruta, "rb") as f:
                        st.download_button(
                            label="⬇️ Descargar Reporte Valor Neto (.xlsx)",
                            data=f,
                            file_name=os.path.basename(ruta),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    st.success(f"✅ Reporte generado: `{os.path.basename(ruta)}`")
    else:
        st.warning("Agregue cultivos al proyecto para habilitar el reporte.")

    # ═══════════════════════════════════════════════════════
    # 3. EXPORTAR GLOBAL AL PUENTE EXCEL + ACTUALIZAR CONFIG
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("🌐 Exportar Global al Proyecto")
    st.caption("Esta acción sincroniza todos los cultivos listados con el archivo puente Excel y actualiza las superficies en la Configuración del proyecto.")

    if not df_proyecto.empty:
        total_sup_sp = float(df_proyecto['Sup_SP_Ha'].sum())
        total_sup_cp = float(df_proyecto['Sup_CP_Ha'].sum())
        st.write(f"**Superficies que se escribirán en Configuración:** SP = **{total_sup_sp:.2f} Ha** | CP = **{total_sup_cp:.2f} Ha**")

        if st.button("💾 Agregar todo al Proyecto (Puente Excel + Config)", type="primary"):
            # 3.1 Actualizar Configuración
            if config and rutas:
                config.superficie_actual = total_sup_sp
                config.superficie_proyecto = total_sup_cp
                config.guardar(rutas["config"])
                st.success(f"✅ Configuración actualizada: `superficie_actual={total_sup_sp:.2f} Ha`, `superficie_proyecto={total_sup_cp:.2f} Ha`")

            # 3.2 Sincronizar puente Excel (incluye hoja Configuracion)
            data_manager.actualizar_desde_proyecto_db(proyecto_db, config)
            st.success("✅ Puente Excel sincronizado con todos los cultivos del proyecto.")
            st.balloons()
    else:
        st.info("No hay cultivos para exportar.")
# ------------------------------------------------------------
# MAIN
# ------------------------------------------------------------
def main():
    st.title("🌾 Sistema de Costos - Gestión de Conceptos y Cultivos")
    st.markdown("**Compatible con:** Resolución Ministerial Nº 115/2015")

    inicializar_session_state()

    tabs = st.tabs([
        "📋 Conceptos de Costo",
        "🌾 Cultivos Referencia",
        "⚙️ Asignación y Cálculo",
        "🔄 Exportar a Proyecto"
    ])

    with tabs[0]:
        render_gestion_conceptos()
    with tabs[1]:
        render_gestion_referencia()
    with tabs[2]:
        render_asignacion_calculo()
    with tabs[3]:
        render_exportar_proyecto()

    st.markdown("---")

if __name__ == "__main__":
    main()

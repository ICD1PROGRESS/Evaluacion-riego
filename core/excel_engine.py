from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional, Tuple, Union
import pandas as pd
from io import BytesIO

from openpyxl.chart import BarChart, LineChart, Reference
#from openpyxl.chart.axis import CategoryAxis, ValueAxis
from openpyxl.chart.label import DataLabelList

class ExcelReportEngine:

    # ============================================================
    # PALETA DE COLORES Y ESTILOS (idénticos a los actuales)
    # ============================================================
    
    COLOR_TITULO = "1F4E78"
    COLOR_HEADER = "D9E1F2"
    COLOR_SUBHEADER = "5B9BD5"
    COLOR_TOTAL = "C6E0B4"
    COLOR_ALERTA = "FFC000"
    
    def __init__(self):
        self.wb = Workbook()
        self._remove_default_sheet()
        self.ws_active = None
    
    def _remove_default_sheet(self):
        """Elimina la hoja por defecto 'Sheet'."""
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
    
    # ============================================================
    # GESTIÓN DE HOJAS
    # ============================================================
    
    def crear_hoja(self, nombre: str) -> Any:
        """Crea o retorna una hoja existente."""
        if nombre in self.wb.sheetnames:
            return self.wb[nombre]
        return self.wb.create_sheet(nombre)
    
    def activar_hoja(self, nombre: str):
        """Activa una hoja para operaciones subsiguientes."""
        self.ws_active = self.crear_hoja(nombre)
        return self.ws_active
    
    # ============================================================
    # ESTILOS REUTILIZABLES
    # ============================================================
    
    def _font(self, bold: bool = False, size: int = 11, color: str = None) -> Font:
        """Crea fuente personalizada."""
        kwargs = {'bold': bold, 'size': size}
        if color:
            kwargs['color'] = color
        return Font(**kwargs)
    
    def _fill(self, color: str) -> PatternFill:
        """Crea relleno de color sólido."""
        return PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    def _border(self, style: str = 'thin') -> Border:
        """Crea borde delgado en todas las celdas."""
        side = Side(style=style)
        return Border(left=side, right=side, top=side, bottom=side)
    
    def _align(self, horizontal: str = 'center', vertical: str = 'center', 
               wrap: bool = True) -> Alignment:
        """Crea alineación personalizada."""
        return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)
    
    # ============================================================
    # PRIMITIVAS DE ESCRITURA
    # ============================================================
    
    def escribir_celda(self, fila: int, columna: int, valor: Any,
                       bold: bool = False, size: int = 11, color: str = None,
                       fill_color: str = None, border: bool = True,
                       align_h: str = 'center', number_format: str = None) -> Any:
        """Escribe una celda con formato completo."""
        ws = self.ws_active
        cell = ws.cell(row=fila, column=columna, value=valor)
        cell.font = self._font(bold=bold, size=size, color=color)
        if fill_color:
            cell.fill = self._fill(fill_color)
        if border:
            cell.border = self._border()
        cell.alignment = self._align(horizontal=align_h)
        if number_format and isinstance(valor, (int, float)) and not isinstance(valor, bool):
            cell.number_format = number_format
        return cell
    
    def escribir_titulo_seccion(self, fila: int, texto: str, 
                                columnas: int = 1, merge: bool = True) -> int:
        """
        Escribe un título de sección (subheader azul oscuro).
        Retorna la siguiente fila disponible.
        """
        ws = self.ws_active
        if merge and columnas > 1:
            ws.merge_cells(start_row=fila, start_column=1, 
                          end_row=fila, end_column=columnas)
        
        self.escribir_celda(fila, 1, texto, bold=True, size=12, 
                           color=self.COLOR_TITULO, align_h='left')
        return fila + 1
    
    def escribir_subtitulo(self, fila: int, texto: str, columnas: int = 1) -> int:
        """Escribe subtítulo en negrita."""
        ws = self.ws_active
        if columnas > 1:
            ws.merge_cells(start_row=fila, start_column=1,
                          end_row=fila, end_column=columnas)
        self.escribir_celda(fila, 1, texto, bold=True, size=11, 
                           color=self.COLOR_TITULO, align_h='left')
        return fila + 1
    
    def escribir_fila_datos(self, fila: int, valores: List[Any],
                           bold: bool = False, fill_color: str = None,
                           number_format: str = '#,##0.00') -> int:
        """
        Escribe una fila completa de datos con formato uniforme.
        Retorna la siguiente fila.
        """
        for col_idx, val in enumerate(valores, 1):
            nf = number_format if isinstance(val, (int, float)) and not isinstance(val, bool) else None
            align = 'right' if isinstance(val, (int, float)) and not isinstance(val, bool) else 'left'
            self.escribir_celda(fila, col_idx, val, bold=bold, 
                               fill_color=fill_color, align_h=align, 
                               number_format=nf)
        return fila + 1
    
    # ============================================================
    # TABLAS CON MULTIINDEX (idéntico a PREP, E_FIN, E_ECO)
    # ============================================================
    
    def escribir_tabla_multiindex(self, fila_inicio: int, df: pd.DataFrame,
                                  formato_numero: str = '#,##0.00',
                                  fila_totales: bool = True) -> int:
        """
        Escribe un DataFrame con MultiIndex de columnas (nivel1, nivel2).
        Formato idéntico a las hojas E_FIN y E_ECO actuales.
        Retorna la fila siguiente al final de la tabla.
        """
        ws = self.ws_active
        fila = fila_inicio
        
        # Validar que tenga MultiIndex
        if not isinstance(df.columns, pd.MultiIndex):
            # Si no es MultiIndex, escribir como tabla simple
            return self.escribir_tabla_simple(fila, df, formato_numero)
        
        nivel1 = [c[0] for c in df.columns]
        nivel2 = [c[1] for c in df.columns]
        
        # --- NIVEL 1 (merge de grupos) ---
        for col_idx, val in enumerate(nivel1, 1):
            self.escribir_celda(fila, col_idx, val if val else None,
                               bold=True, fill_color=self.COLOR_HEADER)
        
        # Merge celdas de nivel 1
        current_group = None
        start_col = 1
        for col_idx, val in enumerate(nivel1, 1):
            if val != current_group and val != '':
                if current_group is not None:
                    ws.merge_cells(start_row=fila, start_column=start_col,
                                  end_row=fila, end_column=col_idx-1)
                current_group = val
                start_col = col_idx
        if current_group:
            ws.merge_cells(start_row=fila, start_column=start_col,
                          end_row=fila, end_column=len(nivel1))
        
        fila += 1
        
        # --- NIVEL 2 (headers detallados) ---
        for col_idx, val in enumerate(nivel2, 1):
            self.escribir_celda(fila, col_idx, val,
                               bold=True, fill_color=self.COLOR_HEADER)
        fila += 1
        
        # --- DATOS ---
        for r_idx, row in enumerate(df.values, fila):
            for col_idx, val in enumerate(row, 1):
                is_total_row = (val == 'TOTAL')
                self.escribir_celda(r_idx, col_idx, val,
                                   bold=is_total_row,
                                   align_h='right' if isinstance(val, (int, float)) else 'left',
                                   number_format=formato_numero if isinstance(val, (int, float)) else None)
        fila = r_idx + 1
        
        # --- FILA TOTALES (opcional) ---
        if fila_totales and len(df) > 0:
            totales = {}
            for col_idx, (k, v) in enumerate(zip(df.columns, df.values[-1] if len(df) > 0 else []), 1):
                if k[1] == 'Cultivos' or k[1] == '':
                    totales[col_idx] = 'TOTAL'
                else:
                    # Sumar columna numérica
                    col_vals = df.iloc[:, col_idx-1]
                    numeric_vals = pd.to_numeric(col_vals, errors='coerce').fillna(0)
                    totales[col_idx] = numeric_vals.sum()
            
            for col_idx, val in totales.items():
                self.escribir_celda(fila, col_idx, val, bold=True,
                                   fill_color=self.COLOR_HEADER,
                                   number_format=formato_numero)
            fila += 1
        
        return fila
    
    def escribir_tabla_simple(self, fila_inicio: int, df: pd.DataFrame,
                              formato_numero: str = '#,##0.00') -> int:
        """Escribe tabla sin MultiIndex."""
        ws = self.ws_active
        fila = fila_inicio
        
        # Headers
        for col_idx, col in enumerate(df.columns, 1):
            self.escribir_celda(fila, col_idx, col, bold=True,
                               fill_color=self.COLOR_HEADER)
        fila += 1
        
        # Datos
        for r_idx, row in enumerate(df.values, fila):
            for col_idx, val in enumerate(row, 1):
                self.escribir_celda(r_idx, col_idx, val,
                                   align_h='right' if isinstance(val, (int, float)) else 'left',
                                   number_format=formato_numero if isinstance(val, (int, float)) else None)
        fila = r_idx + 1
        
        return fila
    
    # ============================================================
    # TABLAS DE FLUJOS (E_FIN, E_ECO)
    # ============================================================
    
    def escribir_flujo_horizontal(self, fila_inicio: int, 
                                  componentes: List[Tuple[str, List[float]]],
                                  anios: List[int],
                                  headers: List[str] = None,
                                  totales: bool = True) -> int:
        """
        Escribe tabla de flujos horizontales (años como columnas).
        Formato idéntico a E_FIN y E_ECO.
        """
        ws = self.ws_active
        fila = fila_inicio
        n = len(anios)
        
        # Headers
        if headers is None:
            headers = ["Componentes", "Años"] + [str(a) for a in anios] + ["Total"]
        
        for col_idx, h in enumerate(headers, 1):
            self.escribir_celda(fila, col_idx, h, bold=True,
                               fill_color=self.COLOR_HEADER)
        fila += 1
        
        # Filas de datos
        for label, valores in componentes:
            is_bold = label in ["Subtotal", "TOTAL COSTOS", "Flujo de Fondos Neto", 
                               "TOTAL", "Total"]
            fila_valores = [label, ""] + valores
            if totales:
                fila_valores.append(sum(valores))
            
            self.escribir_fila_datos(fila, fila_valores, bold=is_bold)
            fila += 1
        
        return fila
    
    # ============================================================
    # SECCIONES ESPECÍFICAS RM 115
    # ============================================================
    def escribir_datos_generales(self, fila_inicio: int, config) -> int:
        """Escribe sección 1. DATOS GENERALES DE DISEÑO (PREP) con campos completos."""
        fila = fila_inicio
        
        fila = self.escribir_titulo_seccion(fila, "1. DATOS GENERALES DE DISEÑO", 3)
        
        datos = [
            ("Código del Proyecto:", config.codigo),
            ("Nombre del Proyecto:", config.nombre),
            ("Departamento:", config.depto),
            ("Municipio:", config.municipio),
            ("Año de inicio del Proyecto:", config.anio_inicio),
            ("Período de Diseño (años):", config.periodo_diseno),
            ("Duración de la Inversión (años):", config.duracion_inversion),
            ("Población Base según Censo:", config.poblacion_base),
            ("Tasa de Crecimiento Poblacional (%):", f"{config.tasa_crecimiento*100:.4f}"),
            ("Promedio de Personas por Familia:", config.personas_por_familia),
            ("Total Familias beneficiadas:", config.total_familias),
            ("Índice de Impacto en el Primer Año (%):", config.indice_impacto),
            ("Superficie actual con riego (Ha):", config.superficie_actual),
            ("Superficie a regar con proyecto (Ha):", config.superficie_proyecto),
            ("Área Incremental (Ha):", config.area_incremental),
            ("Tipo de Cambio (Bs/USD):", config.tipo_cambio),
        ]
        
        for label, valor in datos:
            self.escribir_celda(fila, 1, label, align_h='left')
            self.escribir_celda(fila, 2, valor, align_h='left')
            fila += 1
        
        return fila + 1
    def escribir_rpc(self, fila_inicio: int, config) -> int:
        """Escribe sección de Razones Precio Cuenta (E_ECO)."""
        fila = fila_inicio
        
        fila = self.escribir_titulo_seccion(fila, "RAZONES PRECIO CUENTA DE EFICIENCIA", 2)
        
        rpc_data = [
            ("RPC DIVISA", config.rpc['divisa']),
            ("RPC MANO DE OBRA CALIFICADA", config.rpc['mo_calificada']),
            ("RPC MANO DE OBRA SEMICALIFICADA", config.rpc['mo_semicalificada']),
            ("RPC MANO DE OBRA NO CALIFICADA URBANA", config.rpc['mo_no_calif_urbana']),
            ("RPC MANO DE OBRA NO CALIFICADA RURAL", config.rpc['mo_no_calif_rural']),
            ("TASA SOCIAL DE DESCUENTO", config.tasa_social_descuento),
        ]
        
        for label, val in rpc_data:
            self.escribir_celda(fila, 1, label, bold=True, align_h='left')
            cell = self.escribir_celda(fila, 2, val, align_h='right')
            if isinstance(val, float):
                cell.number_format = '0.00%' if 'TASA' in label else '0.00'
            fila += 1
        
        return fila + 1
    
    def escribir_indicadores(self, fila_inicio: int, indicadores: Dict[str, Any],
                             titulo: str = "INDICADORES") -> int:
        """Escribe tabla de indicadores financieros o económicos."""
        fila = fila_inicio
        
        fila = self.escribir_titulo_seccion(fila, titulo, 2)
        
        for label, val in indicadores.items():
            self.escribir_celda(fila, 1, label, bold=True, align_h='left')
            cell = self.escribir_celda(fila, 2, val, align_h='right')
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = '#,##0.00'
            fila += 1
        
        return fila
    
    # ============================================================
    # UTILIDADES
    # ============================================================    
    def ajustar_anchos(self, ws=None, max_width: int = 50, padding: int = 2):
        """Ajusta anchos de columna automáticamente."""
        target_ws = ws or self.ws_active
        for idx in range(1, target_ws.max_column + 1):
            col_letter = get_column_letter(idx)
            max_len = 0
            for row in target_ws.iter_rows(min_col=idx, max_col=idx):
                cell = row[0]
                try:
                    if cell.value is not None:
                        val_len = len(str(cell.value))
                        if val_len > max_len:
                            max_len = val_len
                except Exception:
                    pass
            target_ws.column_dimensions[col_letter].width = min(max_len + padding, max_width)


    def insertar_grafico_barras(self, fila_inicio: int, columna_inicio: int,
                               datos: pd.DataFrame, x_col: str, y_col: str,
                               titulo: str = "", ancho: int = 15, alto: int = 10,
                               etiquetas: bool = False) -> int:
        ws = self.ws_active
        
        # Escribir los datos en la hoja en una ubicación temporal
        start_row = fila_inicio
        start_col = columna_inicio
        for i, row in datos.iterrows():
            ws.cell(row=start_row + i, column=start_col, value=row[x_col])
            ws.cell(row=start_row + i, column=start_col + 1, value=row[y_col])
        
        # Crear referencia a los datos
        data = Reference(ws, min_col=start_col + 1, min_row=start_row, max_row=start_row + len(datos) - 1)
        categories = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=start_row + len(datos) - 1)
        
        chart = BarChart()
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        chart.title = titulo
        chart.x_axis.title = x_col
        chart.y_axis.title = y_col
        chart.width = ancho
        chart.height = alto
        
        if etiquetas:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
        
        # Posicionar el gráfico debajo de los datos
        chart_pos_row = start_row + len(datos) + 2
        ws.add_chart(chart, f"{get_column_letter(start_col)}{chart_pos_row}")
        
        return chart_pos_row + 1

    def insertar_grafico_lineas(self, fila_inicio: int, columna_inicio: int,
                                datos: pd.DataFrame, x_col: str, y_col: str,
                                titulo: str = "", ancho: int = 15, alto: int = 10) -> int:
        """Inserta un gráfico de líneas."""
        ws = self.ws_active
        start_row = fila_inicio
        start_col = columna_inicio
        for i, row in datos.iterrows():
            ws.cell(row=start_row + i, column=start_col, value=row[x_col])
            ws.cell(row=start_row + i, column=start_col + 1, value=row[y_col])
        
        data = Reference(ws, min_col=start_col + 1, min_row=start_row, max_row=start_row + len(datos) - 1)
        categories = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=start_row + len(datos) - 1)
        
        chart = LineChart()
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        chart.title = titulo
        chart.x_axis.title = x_col
        chart.y_axis.title = y_col
        chart.width = ancho
        chart.height = alto
        
        chart_pos_row = start_row + len(datos) + 2
        ws.add_chart(chart, f"{get_column_letter(start_col)}{chart_pos_row}")
        return chart_pos_row + 1
    
    def to_bytes(self) -> bytes:
        """Retorna el workbook como bytes para descarga."""
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def guardar_archivo(self, ruta: str):
        self.wb.save(ruta)
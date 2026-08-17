# core/data_manager.py
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Optional, Tuple
from .database import ProyectoDB 
from .schema import HOJAS_PUENTE, normalizar_columna, validar_hoja, VALIDATION_SCHEMA

class DataManager:
    """
    Capa de acceso al puente Excel / SQLite con validación de schema.
    Si la ruta termina en .db, opera directamente sobre ProyectoDB.
    """

    RUTA_DEFAULT = "proyecto.db"   # ← CAMBIADO: antes era "proyecto_activo.xlsx"

    def __init__(self, ruta: str = None):
        self.ruta = ruta or self.RUTA_DEFAULT
        self._db: Optional[ProyectoDB] = None
        if str(self.ruta).endswith('.db'):
            self._db = ProyectoDB(self.ruta)

    def _es_db(self) -> bool:
        return self._db is not None    
    # ============================================================
    # LECTURA VALIDADA
    # ============================================================
    #Leer hoja es temporal por compatibilidad con Excel
    def leer_hoja(self, nombre_hoja: str, normalizar: bool = True) -> Tuple[pd.DataFrame, List[str]]:
        """
        Lee una hoja del puente con validación de schema.
        Si la ruta es .db, mapea la hoja a la tabla SQLite equivalente.
        """
        if self._es_db():
            return self._leer_hoja_db(nombre_hoja, normalizar)

        # -------- Modo Excel (código original) --------
        if not os.path.exists(self.ruta):
            return pd.DataFrame(), [f"Archivo no encontrado: {self.ruta}"]

        try:
            df = pd.read_excel(self.ruta, sheet_name=nombre_hoja)
        except Exception as e:
            return pd.DataFrame(), [f"Error leyendo '{nombre_hoja}': {str(e)}"]

        advertencias = []
        ok, errores = validar_hoja(df, nombre_hoja)
        if not ok:
            advertencias.extend(errores)

        if normalizar:
            df.columns = [normalizar_columna(c) for c in df.columns]

        if nombre_hoja in HOJAS_PUENTE:
            tipos = HOJAS_PUENTE[nombre_hoja].get("tipos", {})
            for col, tipo in tipos.items():
                if col in df.columns:
                    if tipo == float:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
                    elif tipo == str:
                        df[col] = df[col].astype(str).replace('nan', '')

        return df, advertencias

    def _leer_hoja_db(self, nombre_hoja: str, normalizar: bool) -> Tuple[pd.DataFrame, List[str]]:
        """Backend SQLite para leer_hoja."""
        try:
            if nombre_hoja == "Cultivos":
                df = self._db.listar_cultivos_proyecto()
            elif nombre_hoja == "Obras_Detalle":
                df = self._db.listar_obras()
            elif nombre_hoja in ("Servicios", "servicios"):
                df = self._db.listar_servicios()
            elif nombre_hoja in ("APU_Componentes", "apu_componentes"):
                df = self._db.listar_apu_componentes()
            elif nombre_hoja == "Inversion_Resumen":
                df = self._db.obtener_resumen_inversion()
            else:
                return pd.DataFrame(), [f"Hoja '{nombre_hoja}' no tiene mapeo en la base de datos"]

            if normalizar and not df.empty:
                df.columns = [normalizar_columna(c) for c in df.columns]
            return df, []
        except Exception as e:
            return pd.DataFrame(), [f"Error leyendo '{nombre_hoja}' de DB: {str(e)}"]
    
    # Leer cultivos es temporal por compatibilidad con Excel
    def leer_cultivos(self) -> Dict:
        """
        Lee cultivos desde Excel o desde proyecto.db.
        Retorna el mismo diccionario que espera app_evaluacion.py.
        """
        if self._es_db():
            return self._leer_cultivos_db()

        # -------- Modo Excel (código original) --------
        df, adv = self.leer_hoja("Cultivos")

        if df.empty:
            return {'ok': False, 'errores': adv}

        df = df[df['nombre'].notna() & (df['nombre'] != 'Nombre') & (df['nombre'] != 'nan')].copy()

        if df.empty:
            return {'ok': False, 'errores': ['No hay cultivos válidos'] + adv}

        # ... cálculos idénticos al original ...
        ingreso_inc_por_cultivo = {}
        for _, row in df.iterrows():
            nombre = str(row['nombre']).strip()
            if not nombre or nombre.lower() == 'nan':
                continue
            ing_cp = float(row.get('ingreso_cp', 0))
            ing_sp = float(row.get('ingreso_sp', 0))
            sup_cp = float(row.get('sup_cp_ha', 0))
            sup_sp = float(row.get('sup_sp_ha', 0))
            ingreso_inc_por_cultivo[nombre] = (ing_cp * sup_cp) - (ing_sp * sup_sp)

        prod_rpc = {
            'BT': ((df['bt_cp'] * df['sup_cp_ha']) - (df['bt_sp'] * df['sup_sp_ha'])).sum(),
            'BNT': ((df['bnt_cp'] * df['sup_cp_ha']) - (df['bnt_sp'] * df['sup_sp_ha'])).sum(),
            'MOC': ((df['moc_cp'] * df['sup_cp_ha']) - (df['moc_sp'] * df['sup_sp_ha'])).sum(),
            'MOS': ((df['mos_cp'] * df['sup_cp_ha']) - (df['mos_sp'] * df['sup_sp_ha'])).sum(),
            'MONU': ((df['monu_cp'] * df['sup_cp_ha']) - (df['monu_sp'] * df['sup_sp_ha'])).sum(),
            'MONR': ((df['monr_cp'] * df['sup_cp_ha']) - (df['monr_sp'] * df['sup_sp_ha'])).sum(),
        }
        prod_rpc['TOTAL'] = sum(prod_rpc.values())

        prod_cp_abs = {
            'BT': (df['bt_cp'] * df['sup_cp_ha']).sum(),
            'BNT': (df['bnt_cp'] * df['sup_cp_ha']).sum(),
            'MOC': (df['moc_cp'] * df['sup_cp_ha']).sum(),
            'MOS': (df['mos_cp'] * df['sup_cp_ha']).sum(),
            'MONU': (df['monu_cp'] * df['sup_cp_ha']).sum(),
            'MONR': (df['monr_cp'] * df['sup_cp_ha']).sum(),
            'TOTAL': (df['costototal_cp'] * df['sup_cp_ha']).sum(),
        }

        prod_sp_abs = {
            'BT': (df['bt_sp'] * df['sup_sp_ha']).sum(),
            'BNT': (df['bnt_sp'] * df['sup_sp_ha']).sum(),
            'MOC': (df['moc_sp'] * df['sup_sp_ha']).sum(),
            'MOS': (df['mos_sp'] * df['sup_sp_ha']).sum(),
            'MONU': (df['monu_sp'] * df['sup_sp_ha']).sum(),
            'MONR': (df['monr_sp'] * df['sup_sp_ha']).sum(),
            'TOTAL': (df['costototal_sp'] * df['sup_sp_ha']).sum(),
        }

        return {
            'ok': True,
            'df': df,
            'ingreso_incremental_anual': sum(ingreso_inc_por_cultivo.values()),
            'ingreso_incremental_por_cultivo': ingreso_inc_por_cultivo,
            'produccion_rpc': prod_rpc,
            'produccion_cp_abs': prod_cp_abs,
            'produccion_sp_abs': prod_sp_abs,
            'num_cultivos': len(df),
            'sup_total_cp': df['sup_cp_ha'].sum(),
            'sup_total_sp': df['sup_sp_ha'].sum(),
            'advertencias': adv
        }

    def _leer_cultivos_db(self) -> Dict:
        """Backend SQLite para leer_cultivos."""
        df = self._db.listar_cultivos_proyecto()
        if df.empty:
            return {'ok': False, 'errores': ['No hay cultivos registrados en la base de datos']}

        # Normalizar nombres de columnas al formato que espera el resto del sistema
        df.columns = [normalizar_columna(c) for c in df.columns]

        # Renombrar columnas que quedaron con '%' para mantener compatibilidad
        rename_map = {}
        for col in df.columns:
            if 'perd_sp_' in col and 'pct' not in col:
                rename_map[col] = 'perd_sp_pct'
            elif 'perd_cp_' in col and 'pct' not in col:
                rename_map[col] = 'perd_cp_pct'
        if rename_map:
            df.rename(columns=rename_map, inplace=True)

        # Asegurar columnas numéricas
        numeric_cols = ['sup_sp_ha', 'sup_cp_ha', 'rend_sp', 'rend_cp', 'perd_sp_pct',
                        'perd_cp_pct', 'precio_bs_ton', 'costototal_sp', 'costototal_cp',
                        'bt_sp', 'bnt_sp', 'monr_sp', 'monu_sp', 'mos_sp', 'moc_sp',
                        'bt_cp', 'bnt_cp', 'monr_cp', 'monu_cp', 'mos_cp', 'moc_cp',
                        'ingreso_sp', 'ingreso_cp']
        for col in numeric_cols:
            if col not in df.columns:
                df[col] = 0.0
            else:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

        # Reutilizar la misma lógica de cálculo del modo Excel
        ingreso_inc_por_cultivo = {}
        for _, row in df.iterrows():
            nombre = str(row.get('nombre', '')).strip()
            if not nombre or nombre.lower() == 'nan':
                continue
            ing_cp = float(row.get('ingreso_cp', 0))
            ing_sp = float(row.get('ingreso_sp', 0))
            sup_cp = float(row.get('sup_cp_ha', 0))
            sup_sp = float(row.get('sup_sp_ha', 0))
            ingreso_inc_por_cultivo[nombre] = (ing_cp * sup_cp) - (ing_sp * sup_sp)

        prod_rpc = {
            'BT': ((df['bt_cp'] * df['sup_cp_ha']) - (df['bt_sp'] * df['sup_sp_ha'])).sum(),
            'BNT': ((df['bnt_cp'] * df['sup_cp_ha']) - (df['bnt_sp'] * df['sup_sp_ha'])).sum(),
            'MOC': ((df['moc_cp'] * df['sup_cp_ha']) - (df['moc_sp'] * df['sup_sp_ha'])).sum(),
            'MOS': ((df['mos_cp'] * df['sup_cp_ha']) - (df['mos_sp'] * df['sup_sp_ha'])).sum(),
            'MONU': ((df['monu_cp'] * df['sup_cp_ha']) - (df['monu_sp'] * df['sup_sp_ha'])).sum(),
            'MONR': ((df['monr_cp'] * df['sup_cp_ha']) - (df['monr_sp'] * df['sup_sp_ha'])).sum(),
        }
        prod_rpc['TOTAL'] = sum(prod_rpc.values())

        prod_cp_abs = {
            'BT': (df['bt_cp'] * df['sup_cp_ha']).sum(),
            'BNT': (df['bnt_cp'] * df['sup_cp_ha']).sum(),
            'MOC': (df['moc_cp'] * df['sup_cp_ha']).sum(),
            'MOS': (df['mos_cp'] * df['sup_cp_ha']).sum(),
            'MONU': (df['monu_cp'] * df['sup_cp_ha']).sum(),
            'MONR': (df['monr_cp'] * df['sup_cp_ha']).sum(),
            'TOTAL': (df['costototal_cp'] * df['sup_cp_ha']).sum(),
        }

        prod_sp_abs = {
            'BT': (df['bt_sp'] * df['sup_sp_ha']).sum(),
            'BNT': (df['bnt_sp'] * df['sup_sp_ha']).sum(),
            'MOC': (df['moc_sp'] * df['sup_sp_ha']).sum(),
            'MOS': (df['mos_sp'] * df['sup_sp_ha']).sum(),
            'MONU': (df['monu_sp'] * df['sup_sp_ha']).sum(),
            'MONR': (df['monr_sp'] * df['sup_sp_ha']).sum(),
            'TOTAL': (df['costototal_sp'] * df['sup_sp_ha']).sum(),
        }

        return {
            'ok': True,
            'df': df,
            'ingreso_incremental_anual': sum(ingreso_inc_por_cultivo.values()),
            'ingreso_incremental_por_cultivo': ingreso_inc_por_cultivo,
            'produccion_rpc': prod_rpc,
            'produccion_cp_abs': prod_cp_abs,
            'produccion_sp_abs': prod_sp_abs,
            'num_cultivos': len(df),
            'sup_total_cp': df['sup_cp_ha'].sum(),
            'sup_total_sp': df['sup_sp_ha'].sum(),
            'advertencias': []
        }
    # Leer inversion_resumen es temporal por compatibilidad con Excel
    def leer_inversion_resumen(self) -> Dict:
        """
        Lee resumen de inversión desde Excel o desde proyecto.db.
        """
        if self._es_db():
            return self._leer_inversion_resumen_db()

        # -------- Modo Excel (código original) --------
        if not os.path.exists(self.ruta):
            return {'ok': False, 'errores': [f"Archivo no encontrado: {self.ruta}"]}

        try:
            df = pd.read_excel(self.ruta, sheet_name="Inversion_Resumen", header=None)
        except Exception as e:
            return {'ok': False, 'errores': [f"Error leyendo Inversion_Resumen: {str(e)}"]}

        header_row_idx = None
        for i, row in df.iterrows():
            if row.astype(str).str.contains('Categoría', case=False, na=False).any():
                header_row_idx = i
                break

        if header_row_idx is None:
            try:
                df = pd.read_excel(self.ruta, sheet_name="Inversion_Resumen")
            except Exception as e:
                return {'ok': False, 'errores': [f"No se encontró fila de encabezados: {str(e)}"]}
        else:
            df = pd.read_excel(self.ruta, sheet_name="Inversion_Resumen", header=header_row_idx)

        # ... resto del procesamiento Excel idéntico ...
        col_map = {}
        for c in df.columns:
            c_str = str(c).strip().upper()
            if 'CATEGOR' in c_str:
                col_map[c] = 'CATEGORIA'
            elif c_str == 'BT':
                col_map[c] = 'BT'
            elif c_str in ['BNT', 'BIENES NO TRANSABLES', 'MATERIALES LOCALES', 'ML']:
                col_map[c] = 'BNT'
            elif c_str == 'MOC':
                col_map[c] = 'MOC'
            elif c_str == 'MOS':
                col_map[c] = 'MOS'
            elif c_str == 'MONU':
                col_map[c] = 'MONU'
            elif c_str == 'MONR' or c_str == 'MOL':
                col_map[c] = 'MONR'
            elif 'TOTAL' in c_str:
                col_map[c] = 'TOTAL'

        df.rename(columns=col_map, inplace=True)
        for c in ['CATEGORIA', 'BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR', 'TOTAL']:
            if c not in df.columns:
                df[c] = 0

        df = df[df['CATEGORIA'].notna() & (df['CATEGORIA'] != 'Categoría') & (df['CATEGORIA'] != 'nan')].copy()
        for c in ['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR', 'TOTAL']:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

        # Separar obras civiles y servicios (compatible con la mejora anterior)
        cats_obras_civiles = ['OBRAS CIVILES']
        cats_servicios = ['Asistencia Técnica Integral', 'Supervisión de Obras']

        obras_civiles = df[df['CATEGORIA'].isin(cats_obras_civiles)]
        servicios_inv = df[df['CATEGORIA'].isin(cats_servicios)]

        inv_obras_rpc = obras_civiles[['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR']].sum().to_dict()
        inv_servicios_rpc = servicios_inv[['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR']].sum().to_dict()
        inv_total = obras_civiles['TOTAL'].sum() + servicios_inv['TOTAL'].sum()
        inv_rpc = {k: inv_obras_rpc.get(k, 0) + inv_servicios_rpc.get(k, 0)
                   for k in ['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR']}

        om_row = df[df['CATEGORIA'].str.contains('Operación', case=False, na=False)]
        om_rpc, om_total = {}, 0.0
        if not om_row.empty:
            om_total = float(om_row['TOTAL'].values[0])
            if om_total > 0:
                om_rpc = om_row[['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR']].iloc[0].to_dict()

        mit_row = df[df['CATEGORIA'].str.contains('Mitigación', case=False, na=False)]
        mit_rpc, mit_total = {}, 0.0
        if not mit_row.empty:
            mit_total = float(mit_row['TOTAL'].values[0])
            if mit_total > 0:
                mit_rpc = mit_row[['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR']].iloc[0].to_dict()

        return {
            'ok': True,
            'df': df,
            'inversion_rpc': inv_rpc,
            'inversion_total': inv_total,
            'inversion_obras_rpc': inv_obras_rpc,
            'inversion_servicios_rpc': inv_servicios_rpc,
            'om_rpc': om_rpc,
            'om_total': om_total,
            'mitigacion_rpc': mit_rpc,
            'mitigacion_total': mit_total,
            'advertencias': []
        }

    def _leer_inversion_resumen_db(self) -> Dict:
        """Backend SQLite para leer_inversion_resumen."""
        df = self._db.obtener_resumen_inversion()
        if df.empty:
            return {'ok': False, 'errores': ['No hay datos de inversión en la base de datos']}

        for c in ['CATEGORIA', 'BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR', 'TOTAL']:
            if c not in df.columns:
                df[c] = 0

        for c in ['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR', 'TOTAL']:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

        cats_obras_civiles = ['OBRAS CIVILES']
        cats_servicios = ['Asistencia Técnica Integral', 'Supervisión de Obras']

        obras_civiles = df[df['CATEGORIA'].isin(cats_obras_civiles)]
        servicios_inv = df[df['CATEGORIA'].isin(cats_servicios)]

        inv_obras_rpc = obras_civiles[['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR']].sum().to_dict()
        inv_servicios_rpc = servicios_inv[['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR']].sum().to_dict()
        inv_total = obras_civiles['TOTAL'].sum() + servicios_inv['TOTAL'].sum()
        inv_rpc = {k: inv_obras_rpc.get(k, 0) + inv_servicios_rpc.get(k, 0)
                   for k in ['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR']}

        om_row = df[df['CATEGORIA'].str.contains('Operación', case=False, na=False)]
        om_rpc, om_total = {}, 0.0
        if not om_row.empty:
            om_total = float(om_row['TOTAL'].values[0])
            if om_total > 0:
                om_rpc = om_row[['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR']].iloc[0].to_dict()

        mit_row = df[df['CATEGORIA'].str.contains('Mitigación', case=False, na=False)]
        mit_rpc, mit_total = {}, 0.0
        if not mit_row.empty:
            mit_total = float(mit_row['TOTAL'].values[0])
            if mit_total > 0:
                mit_rpc = mit_row[['BT', 'BNT', 'MOC', 'MOS', 'MONU', 'MONR']].iloc[0].to_dict()

        return {
            'ok': True,
            'df': df,
            'inversion_rpc': inv_rpc,
            'inversion_total': inv_total,
            'inversion_obras_rpc': inv_obras_rpc,
            'inversion_servicios_rpc': inv_servicios_rpc,
            'om_rpc': om_rpc,
            'om_total': om_total,
            'mitigacion_rpc': mit_rpc,
            'mitigacion_total': mit_total,
            'advertencias': []
        }

    # ============================================================
    # ESCRITURA VALIDADA
    # ============================================================
    def crear_proyecto_nuevo(self) -> str:
        """Crea proyecto_activo.xlsx o proyecto.db según la extensión."""
        if self._es_db():
            # ProyectoDB ya inicializa todas las tablas en su constructor
            ProyectoDB(self.ruta)
            return self.ruta

        # -------- Modo Excel (código original) --------
        wb = Workbook()
        ws_cult = wb.active
        ws_cult.title = "Cultivos"
        headers = HOJAS_PUENTE["Cultivos"]["columnas"]
        ws_cult.append(headers)

        ws_inv = wb.create_sheet("Inversion_Resumen")
        ws_inv.append(["RESUMEN DE INVERSIÓN POR RPC - RM 115/2015"])
        ws_inv.append([])
        ws_inv.append(["Categoría", "BT", "BNT", "MOC", "MOS", "MONU", "MONR", "TOTAL"])
        categorias = [
            ["OBRAS CIVILES", 0, 0, 0, 0, 0, 0, 0],
            ["Asistencia Técnica Integral", 0, 0, 0, 0, 0, 0, 0],
            ["Supervisión de Obras", 0, 0, 0, 0, 0, 0, 0],
            ["Operación y Mantenimiento", 0, 0, 0, 0, 0, 0, 0],
            ["Mitigación Ambiental", 0, 0, 0, 0, 0, 0, 0],
            ["TOTAL INVERSIÓN", 0, 0, 0, 0, 0, 0, 0]
        ]
        for cat in categorias:
            ws_inv.append(cat)

        wb.save(self.ruta)
        return self.ruta

    # guardar_hoja es temporal por compatibilidad con Excel
    def guardar_hoja(self, nombre_hoja: str, df: pd.DataFrame,
                     validar: bool = True) -> Tuple[bool, List[str]]:
        """
        Guarda un DataFrame en una hoja del puente.
        Si la ruta es .db, sincroniza con la tabla SQLite equivalente.
        """
        if self._es_db():
            return self._guardar_hoja_db(nombre_hoja, df, validar)

        # -------- Modo Excel (código original) --------
        if validar:
            ok, errores = validar_hoja(df, nombre_hoja)
            if not ok:
                return False, errores

        try:
            if os.path.exists(self.ruta):
                wb = load_workbook(self.ruta)
                if nombre_hoja in wb.sheetnames:
                    del wb[nombre_hoja]
            else:
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])

            ws = wb.create_sheet(nombre_hoja)
            for col, header in enumerate(df.columns, 1):
                ws.cell(row=1, column=col, value=header)
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(self.ruta)
            return True, []
        except Exception as e:
            return False, [f"Error guardando '{nombre_hoja}': {str(e)}"]

    def _guardar_hoja_db(self, nombre_hoja: str, df: pd.DataFrame,
                         validar: bool) -> Tuple[bool, List[str]]:
        """Backend SQLite para guardar_hoja."""
        try:
            if nombre_hoja == "Cultivos":
                # Vaciar tabla e insertar desde DataFrame
                with self._db._get_conn() as conn:
                    conn.execute("DELETE FROM proyecto_cultivos")
                    conn.commit()
                for _, row in df.iterrows():
                    datos = {
                        'nombre': str(row.get('nombre', row.get('Nombre', ''))),
                        'codigo': str(row.get('codigo', row.get('Codigo', ''))),
                        'sup_sp_ha': row.get('sup_sp_ha', row.get('Sup_SP_Ha', 0)),
                        'sup_cp_ha': row.get('sup_cp_ha', row.get('Sup_CP_Ha', 0)),
                        'rend_sp': row.get('rend_sp', row.get('Rend_SP', 0)),
                        'rend_cp': row.get('rend_cp', row.get('Rend_CP', 0)),
                        'perd_sp_pct': row.get('perd_sp_pct', row.get('Perd_SP_%', 0)),
                        'perd_cp_pct': row.get('perd_cp_pct', row.get('Perd_CP_%', 0)),
                        'precio_bs_ton': row.get('precio_bs_ton', row.get('Precio_Bs_Ton', 0)),
                        'costo_total_sp': row.get('costototal_sp', row.get('CostoTotal_SP', 0)),
                        'costo_total_cp': row.get('costototal_cp', row.get('CostoTotal_CP', 0)),
                        'bt_sp': row.get('bt_sp', row.get('BT_SP', 0)),
                        'bnt_sp': row.get('bnt_sp', row.get('BNT_SP', 0)),
                        'monr_sp': row.get('monr_sp', row.get('MONR_SP', 0)),
                        'monu_sp': row.get('monu_sp', row.get('MONU_SP', 0)),
                        'mos_sp': row.get('mos_sp', row.get('MOS_SP', 0)),
                        'moc_sp': row.get('moc_sp', row.get('MOC_SP', 0)),
                        'bt_cp': row.get('bt_cp', row.get('BT_CP', 0)),
                        'bnt_cp': row.get('bnt_cp', row.get('BNT_CP', 0)),
                        'monr_cp': row.get('monr_cp', row.get('MONR_CP', 0)),
                        'monu_cp': row.get('monu_cp', row.get('MONU_CP', 0)),
                        'mos_cp': row.get('mos_cp', row.get('MOS_CP', 0)),
                        'moc_cp': row.get('moc_cp', row.get('MOC_CP', 0)),
                        'ingreso_sp': row.get('ingreso_sp', row.get('Ingreso_SP', 0)),
                        'ingreso_cp': row.get('ingreso_cp', row.get('Ingreso_CP', 0)),
                    }
                    self._db.agregar_cultivo_proyecto(datos)
                return True, []

            elif nombre_hoja == "Obras_Detalle":
                df_db = df.copy()
                expected = ['codigo_item', 'descripcion', 'descripcion_norm', 'unidad',
                            'cantidad', 'precio_unitario', 'parcial_directo', 'tipo_hoja',
                            'subcategoria', 'tipo_rpc', 'fecha_clasificacion',
                            'precio_unitario_final', 'parcial_real', 'factor_apu',
                            'indirectos_asignados', 'parcial']
                for col in expected:
                    if col not in df_db.columns:
                        for c in df_db.columns:
                            if c.lower().replace(' ', '_') == col.lower():
                                df_db.rename(columns={c: col}, inplace=True)
                                break
                self._db.guardar_obras(df_db)
                return True, []

            elif nombre_hoja in ("Servicios", "servicios"):
                self._db.guardar_servicios(df)
                return True, []

            elif nombre_hoja in ("APU_Componentes", "apu_componentes"):
                self._db.guardar_apu_componentes(df)
                return True, []

            elif nombre_hoja == "Inversion_Resumen":
                return True, ["Advertencia: Inversion_Resumen se calcula automáticamente desde obras y servicios; no se guarda directamente en DB"]

            else:
                return False, [f"Hoja '{nombre_hoja}' no tiene mapeo a tabla SQLite"]

        except Exception as e:
            return False, [f"Error guardando '{nombre_hoja}' en DB: {str(e)}"]

    def actualizar_desde_proyecto_db(self, proyecto_db, config=None):
        """
        Sincroniza el archivo Excel puente con los datos de proyecto.db.
        Si se pasa 'config' (ConfiguracionProyecto), también escribe la hoja Configuracion.
        """
        if self._es_db():
            return True  # Ya es la misma base de datos

        ruta = self.ruta

        # Crear archivo base si no existe
        if not os.path.exists(ruta):
            with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
                pd.DataFrame().to_excel(writer, sheet_name='Cultivos', index=False)
                pd.DataFrame().to_excel(writer, sheet_name='Obras_Detalle', index=False)
                pd.DataFrame().to_excel(writer, sheet_name='APU_Componentes', index=False)
                pd.DataFrame().to_excel(writer, sheet_name='Inversion_Resumen', index=False)
                if config is not None:
                    pd.DataFrame().to_excel(writer, sheet_name='Configuracion', index=False)
        try:
            book = load_workbook(ruta)

            # 1. HOJA CULTIVOS
            df_cultivos = proyecto_db.listar_cultivos_proyecto()
            if 'Cultivos' in book.sheetnames:
                del book['Cultivos']
            ws = book.create_sheet('Cultivos', 0)
            for r in dataframe_to_rows(df_cultivos, index=False, header=True):
                ws.append(r)

            # 2. HOJA OBRAS_DETALLE
            df_obras = proyecto_db.listar_obras()
            if 'Obras_Detalle' in book.sheetnames:
                del book['Obras_Detalle']
            ws = book.create_sheet('Obras_Detalle', 1)
            for r in dataframe_to_rows(df_obras, index=False, header=True):
                ws.append(r)

            # 3. HOJA INVERSION_RESUMEN
            df_resumen = proyecto_db.obtener_resumen_inversion()
            if 'Inversion_Resumen' in book.sheetnames:
                del book['Inversion_Resumen']
            ws = book.create_sheet('Inversion_Resumen', 2)
            for r in dataframe_to_rows(df_resumen, index=False, header=True):
                ws.append(r)

            # 4. HOJA APU_COMPONENTES
            df_apu = proyecto_db.listar_apu_componentes()
            if 'APU_Componentes' in book.sheetnames:
                del book['APU_Componentes']
            ws = book.create_sheet('APU_Componentes', 3)
            if not df_apu.empty:
                for r in dataframe_to_rows(df_apu, index=False, header=True):
                    ws.append(r)

            # 5. HOJA CONFIGURACION (NUEVO)
            if config is not None:
                if 'Configuracion' in book.sheetnames:
                    del book['Configuracion']
                ws = book.create_sheet('Configuracion', 4)

                # Aplanar el diccionario (especialmente el sub-dict 'rpc')
                config_dict = config.to_dict() if hasattr(config, 'to_dict') else dict(config)
                if 'rpc' in config_dict and isinstance(config_dict['rpc'], dict):
                    for k, v in config_dict['rpc'].items():
                        config_dict[f'rpc_{k}'] = v
                    del config_dict['rpc']

                df_config = pd.DataFrame([
                    {'Parametro': k, 'Valor': v}
                    for k, v in config_dict.items()
                ])
                for r in dataframe_to_rows(df_config, index=False, header=True):
                    ws.append(r)

            book.save(ruta)
            return True

        except Exception as e:
            print(f"Error al actualizar el puente Excel: {e}")
            return False